import argparse
import json
import os
import re
from datetime import datetime, timedelta

import pandas as pd
import openpyxl


def _ttl_prefix(now: datetime) -> str:
    ttl = now + timedelta(hours=24)
    return ttl.strftime("[TTL_%y%m%d_%H%M]")


def load_csv(csv_path: str) -> pd.DataFrame:
    # dtype=str + keep_default_na=False: prevent big integers turning into floats,
    # and preserve empty strings.
    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)

    # Normalize all string cells: strip whitespace.
    for col in df.columns:
        df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)

    return df


def filter_and_clean(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    required_cols = ["判罚原因", "判罚内容", "影响及结果"]
    for c in required_cols:
        if c not in df.columns:
            raise ValueError(f"CSV 缺少必需列: {c}")

    def _is_none_or_empty(v: str) -> bool:
        if v is None:
            return True
        s = str(v).strip()
        if s == "":
            return True
        return s.lower() == "none"

    before = len(df)

    mask_bad = df["判罚原因"].map(_is_none_or_empty) | df["判罚内容"].map(_is_none_or_empty)
    df2 = df.loc[~mask_bad].copy()

    # Remove trace markers like: [原申诉字段=否No] / [原申诉字段=是Yes]
    marker_re = re.compile(r"\s*\[原申诉字段=[^\]]*\]")

    def _clean_impact(v: str) -> str:
        s = "" if v is None else str(v)
        s = marker_re.sub("", s)
        # Normalize whitespace
        s = re.sub(r"\s+", " ", s).strip()
        return s

    df2["影响及结果"] = df2["影响及结果"].map(_clean_impact)

    after = len(df2)
    stats = {
        "rows_before": before,
        "rows_after": after,
        "rows_dropped": before - after,
    }

    return df2, stats


def extract_dropdowns_from_template_xlsx(template_xlsx_path: str, sheet_name: str) -> dict:
    wb = openpyxl.load_workbook(template_xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"模板 xlsx 中找不到工作表: {sheet_name}，现有: {wb.sheetnames}")

    ws = wb[sheet_name]

    dropdowns = {}
    for dv in ws.data_validations.dataValidation:
        # We only care about list type validations.
        if dv.type != "list":
            continue
        # dv.sqref could be a multi-range; cast to str for matching.
        sqref = str(dv.sqref)
        # Example: E2:E200 / I2:I200
        if sqref.startswith("E"):
            dropdowns["E"] = [x.strip() for x in str(dv.formula1).strip('"').split(",") if x.strip()]
        if sqref.startswith("I"):
            dropdowns["I"] = [x.strip() for x in str(dv.formula1).strip('"').split(",") if x.strip()]

    return dropdowns


def validate_dropdown_values(df: pd.DataFrame, dropdowns: dict):
    # Column mapping:
    #   E = 类型
    #   I = 是否申诉成功
    if "类型" not in df.columns:
        raise ValueError("CSV 缺少列: 类型（用于校验 E 列下拉）")
    if "是否申诉成功" not in df.columns:
        raise ValueError("CSV 缺少列: 是否申诉成功（用于校验 I 列下拉）")

    allowed_e = set(dropdowns.get("E", []))
    allowed_i = set(dropdowns.get("I", []))

    if not allowed_e:
        raise ValueError("未能从模板中解析到 E 列下拉列表")
    if not allowed_i:
        raise ValueError("未能从模板中解析到 I 列下拉列表")

    bad_e = sorted(set(df["类型"].tolist()) - allowed_e)
    bad_i = sorted(set(df["是否申诉成功"].tolist()) - allowed_i)

    if bad_e:
        raise ValueError(f"类型 列存在不在下拉范围内的值（E列）：{bad_e}；允许值：{sorted(allowed_e)}")
    if bad_i:
        raise ValueError(f"是否申诉成功 列存在不在下拉范围内的值（I列）：{bad_i}；允许值：{sorted(allowed_i)}")


def append_to_xlsx(
    template_xlsx_path: str,
    sheet_name: str,
    df: pd.DataFrame,
    out_xlsx_path: str,
) -> dict:
    wb = openpyxl.load_workbook(template_xlsx_path)
    ws = wb[sheet_name]

    # Determine append start row:
    # - openpyxl max_row is safe here because we downloaded the current sheet.
    start_row = ws.max_row + 1

    columns = [
        "判罚时间",
        "UID",
        "Handle",
        "Room id / Video id",
        "类型",
        "判罚原因",
        "判罚内容",
        "判罚截图",
        "是否申诉成功",
        "影响及结果",
        "提出人",
        "跟进人",
    ]
    for c in columns:
        if c not in df.columns:
            raise ValueError(f"清洗后数据缺少列: {c}")

    # Write rows
    for i, row in enumerate(df[columns].itertuples(index=False, name=None), start=0):
        r = start_row + i
        for c_idx, value in enumerate(row, start=1):
            # Ensure everything written as string to avoid Excel scientific notation.
            v = "" if value is None else str(value)
            ws.cell(row=r, column=c_idx).value = v

    wb.save(out_xlsx_path)

    return {
        "sheet_name": sheet_name,
        "append_start_row": start_row,
        "append_rows": len(df),
        "append_end_row": start_row + len(df) - 1 if len(df) else start_row - 1,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv", required=True)
    parser.add_argument("--template_xlsx", required=True)
    parser.add_argument("--sheet_name", required=True)
    parser.add_argument("--out_dir", default=".ephemeral_pool")
    args = parser.parse_args()

    now = datetime.now()
    ttl = _ttl_prefix(now)

    os.makedirs(args.out_dir, exist_ok=True)

    cleaned_csv_path = os.path.join(args.out_dir, f"{ttl}_ultimate_violations_cleaned.csv")
    out_xlsx_path = os.path.join(args.out_dir, f"{ttl}_sheet_append_payload.xlsx")

    df0 = load_csv(args.csv)
    df1, stats = filter_and_clean(df0)

    # Always write a cleaned snapshot for audit (DLQ-like local trace)
    df1.to_csv(cleaned_csv_path, index=False)

    dropdowns = extract_dropdowns_from_template_xlsx(args.template_xlsx, args.sheet_name)
    validate_dropdown_values(df1, dropdowns)

    append_meta = append_to_xlsx(args.template_xlsx, args.sheet_name, df1, out_xlsx_path)

    result = {
        "cleaned_csv_path": os.path.abspath(cleaned_csv_path),
        "out_xlsx_path": os.path.abspath(out_xlsx_path),
        "stats": stats,
        "dropdowns": dropdowns,
        "append_meta": append_meta,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
