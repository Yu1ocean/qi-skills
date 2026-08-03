#!/usr/bin/env python3
"""Upload task-flow-engine notification logs (JSONL) into a Lark Sheet.

目标：把 task_patrol_notify.py 产出的 notification_logs/notify_<YYYY-MM-DD>.jsonl
单向同步到飞书底表的 `Task_Notify_Logs` sheet 中，形成可视化审计链路。

实现约束（来自用户规范）：
- 不允许覆盖已有数据（幂等、只追加）
- 首列 Run_ID 必须全局唯一
- 写入后必须执行“写→等2s→读回校验”的 RAW 原子锁

注意：当前 MCP 工具集没有“按行 append”接口，因此这里采用 **MCP 方案的 append 语义模拟**：
1) MCP 下载整份 spreadsheet 到本地 xlsx
2) 在本地把新增日志插入到表头下方（第 2 行起）
3) MCP 以 update_sheet 的方式回写该 sheet（保持旧行不变 + 仅新增）

该方案在语义上等价于 append，但物理写入是“全量回写同一份数据”。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook


TZ_SHANGHAI = timezone(timedelta(hours=8))


def _repo_root() -> Path:
    # user_skills/task-flow-engine/scripts/upload_notify_logs_to_sheet.py
    return Path(__file__).resolve().parents[1]


def _workspace_root() -> Path:
    # parents[0]=scripts, [1]=task-flow-engine, [2]=user_skills, [3]=workspace root
    return Path(__file__).resolve().parents[3]


def _validate_safe_path_under_repo(path: Path, *, repo_root: Path, arg_name: str) -> Path:
    repo_root = repo_root.resolve()
    resolved = path.resolve()
    try:
        resolved.relative_to(repo_root)
    except ValueError:
        raise ValueError(f"{arg_name} 必须位于任务目录内：{resolved} (repo_root={repo_root})")
    return resolved


def _spreadsheet_to_url(url_or_token: str) -> str:
    s = (url_or_token or "").strip()
    if not s:
        raise ValueError("spreadsheet 不能为空")
    if s.startswith("http"):
        return s
    return f"https://bytedance.larkoffice.com/sheets/{s}"


@dataclass
class CmdResult:
    returncode: int
    stdout: str
    stderr: str


def _run_cmd(cmd: Sequence[str], *, cwd: Optional[Path] = None) -> CmdResult:
    p = subprocess.run(
        list(cmd),
        cwd=str(cwd) if cwd else None,
        capture_output=True,
        text=True,
    )
    return CmdResult(returncode=p.returncode, stdout=p.stdout, stderr=p.stderr)


def _require_ok(res: CmdResult, *, what: str) -> None:
    if res.returncode == 0:
        return
    detail = (res.stderr or res.stdout or "").strip()
    raise RuntimeError(f"{what} 失败：returncode={res.returncode}\n{detail}")


def _bytedcli_auth() -> None:
    """执行 bytedcli-auth，确保后续 MCP / CLI 都以用户身份可用。

    NOTE: 某些运行环境未下发 bytedcli-auth（或路径不同）。为避免每日巡检链路因鉴权脚本缺失而硬阻断，
    这里做安全降级：缺失则 WARN 并继续。
    """
    ws = _workspace_root()
    sh = ws / "inner_skills" / "bytedcli-auth" / "scripts" / "bytedcli_auth.sh"
    if not sh.exists():
        print(f"[WARN] bytedcli-auth not found, skip: {sh}", file=sys.stderr)
        return

    res = _run_cmd(["bash", str(sh)], cwd=sh.parent)
    _require_ok(res, what="bytedcli-auth")


def _parse_mcp_output_file_path(text: str) -> str:
    # MCP wrapper 输出通常包含：file_path: "/abs/path.xlsx"
    m = re.search(r"file_path:\s*\"([^\"]+)\"", text)
    if m:
        return m.group(1)

    m = re.search(r"file_path:\s*([^\s]+)", text)
    if m:
        return m.group(1)

    raise RuntimeError(f"无法从 MCP 输出中解析 file_path。raw=\n{text}\n")


def _mcp_lark_download_spreadsheet(*, spreadsheet_url: str) -> Path:
    ws = _workspace_root()
    tool = ws / "inner_skills" / "lark" / "mcp_lark_lark_download.py"
    if tool.exists():
        payload = json.dumps({"document_url": spreadsheet_url}, ensure_ascii=False)
        res = _run_cmd([sys.executable, str(tool), payload], cwd=ws)
        _require_ok(res, what="mcp:lark_lark_download")

        file_path = _parse_mcp_output_file_path(res.stdout + "\n" + res.stderr)
        p = Path(file_path)
        if not p.exists():
            raise FileNotFoundError(f"MCP 下载返回的文件不存在：{p}")
        return p

    # 2026-08-02: mcp_lark_lark_download 在当前运行时已下线。
    # 保持 MCP-only / lark-cli 用户身份链路，降级使用定制版 lark-cli 的只读 workbook export。
    export_dir = _repo_root() / ".runtime" / "downloads"
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = f"notify_log_source_{datetime.now(TZ_SHANGHAI).strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = export_dir / filename
    output_rel = output_path.relative_to(ws)
    res = _run_cmd(
        [
            "lark-cli",
            "sheets",
            "+workbook-export",
            "--url",
            spreadsheet_url,
            "--output-path",
            str(output_rel),
        ],
        cwd=ws,
    )
    _require_ok(res, what="lark-cli sheets +workbook-export")
    if not output_path.exists():
        raise FileNotFoundError(f"lark-cli 导出成功但本地文件不存在：{output_path}")
    return output_path


def _mcp_update_sheet(*, spreadsheet_url: str, sheet_name: str, source_file_path: Path) -> None:
    ws = _workspace_root()
    tool = ws / "inner_skills" / "lark_sheets_update" / "mcp_lark_sheets_update_lark_update_sheet.py"
    if not tool.exists():
        raise FileNotFoundError(f"找不到 mcp_lark_sheets_update_lark_update_sheet：{tool}")

    payload = json.dumps(
        {
            "document_url": spreadsheet_url,
            "sheet_name": sheet_name,
            "source_file_path": str(source_file_path.resolve()),
        },
        ensure_ascii=False,
    )
    res = _run_cmd([sys.executable, str(tool), payload], cwd=ws)
    _require_ok(res, what="mcp:lark_sheets_update_lark_update_sheet")


def _iter_jsonl(path: Path) -> Iterable[Dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        for line_no, raw in enumerate(f, start=1):
            s = raw.strip()
            if not s:
                continue
            try:
                obj = json.loads(s)
            except Exception as e:
                raise ValueError(f"notify log JSONL 第 {line_no} 行不是合法 JSON：{e}\nraw={s[:200]}") from e
            if not isinstance(obj, dict):
                continue
            yield obj


def _safe_one_line(s: str, *, max_len: int) -> str:
    s = (s or "").replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", "\\n")
    if len(s) > max_len:
        return s[: max_len - 3] + "..."
    return s


def _parse_iso(dt: str) -> Optional[datetime]:
    if not dt:
        return None
    try:
        # Python 3.11 支持解析带 offset 的 ISO
        return datetime.fromisoformat(dt)
    except Exception:
        return None


def _format_dt_shanghai(dt: Optional[datetime]) -> str:
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ_SHANGHAI).strftime("%Y-%m-%d %H:%M:%S")


def _make_unique_row_id(record: Dict[str, Any]) -> str:
    run_id = str(record.get("run_id") or "").strip()
    receiver = record.get("receiver") or {}
    email = str((receiver.get("email") if isinstance(receiver, dict) else "") or "").strip()
    created_at = str(record.get("created_at") or "").strip()
    payload_path = str(record.get("payload_path") or "").strip()
    mode = str(record.get("mode") or "").strip()
    route_key = str(record.get("route_key") or "").strip()

    salt = f"{email}|{created_at}|{payload_path}|{mode}|{route_key}"
    h = hashlib.sha1(salt.encode("utf-8")).hexdigest()[:10]

    # Run_ID 列：短、稳定、可读
    if run_id:
        return f"{run_id}_{h}"
    return f"NO_RUN_{h}"


def _record_to_row(record: Dict[str, Any], *, preview_max_chars: int) -> List[Any]:
    receiver = record.get("receiver") or {}
    email = ""
    if isinstance(receiver, dict):
        email = str(receiver.get("email") or "").strip()

    created_at = _parse_iso(str(record.get("created_at") or ""))

    mode = str(record.get("mode") or "").strip()
    result = str(record.get("result") or "").strip()
    status = "|".join([x for x in [mode, result] if x])

    count = record.get("count")
    if count is None:
        count_str = ""
    else:
        count_str = str(count)

    preview = _safe_one_line(str(record.get("message_preview") or ""), max_len=preview_max_chars)
    err = str(record.get("error") or "").strip()

    row_id = _make_unique_row_id(record)

    return [
        row_id,
        _format_dt_shanghai(created_at),
        email,
        status,
        count_str,
        preview,
        err,
    ]


def _load_sheet_rows(xlsx: Path, *, sheet_name: str) -> Tuple[List[str], List[List[Any]]]:
    wb = load_workbook(xlsx)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"下载的表格中找不到工作表：{sheet_name}；现有={wb.sheetnames}")

    ws = wb[sheet_name]

    # header
    headers: List[str] = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip())

    # rows
    data_rows: List[List[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 全空行跳过
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        data_rows.append(list(row))

    return headers, data_rows


def _write_single_sheet_xlsx(*, sheet_name: str, headers: List[str], rows: List[List[Any]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(headers)
    for r in rows:
        # pad
        padded = list(r) + [""] * max(0, len(headers) - len(r))
        ws.append(padded[: len(headers)])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--spreadsheet",
        required=True,
        help="要写入日志的飞书表格 URL 或 token（即主底表）。",
    )
    ap.add_argument(
        "--sheet-name",
        default="Task_Notify_Logs",
        help="日志工作表名称，默认 Task_Notify_Logs",
    )
    ap.add_argument(
        "--log-file",
        default=None,
        help="notification_logs/notify_<YYYY-MM-DD>.jsonl 路径（相对 repo_root）；不填则默认使用 notify_<UTC日期>.jsonl",
    )
    ap.add_argument(
        "--preview-max-chars",
        type=int,
        default=240,
        help="Preview 字段最大字符数（单元格内换行会被转义为\\n）",
    )
    ap.add_argument(
        "--skip-auth",
        action="store_true",
        help="跳过 bytedcli-auth（仅用于已在外层完成鉴权的场景）",
    )

    args = ap.parse_args()

    repo_root = _repo_root()

    if args.log_file:
        log_path = Path(args.log_file)
        if not log_path.is_absolute():
            log_path = repo_root / log_path
        log_path = _validate_safe_path_under_repo(log_path, repo_root=repo_root, arg_name="log-file")
    else:
        utc_day = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        log_path = repo_root / "notification_logs" / f"notify_{utc_day}.jsonl"

    if not log_path.exists():
        raise FileNotFoundError(f"找不到通知日志文件：{log_path}")

    spreadsheet_url = _spreadsheet_to_url(args.spreadsheet)

    if not args.skip_auth:
        _bytedcli_auth()

    # 1) 下载当前表格
    local_xlsx = _mcp_lark_download_spreadsheet(spreadsheet_url=spreadsheet_url)

    # 2) 读取现有 rows + 已写入 Run_ID
    headers, existing_rows = _load_sheet_rows(local_xlsx, sheet_name=args.sheet_name)

    expected_headers = ["Run_ID", "Date", "Receiver", "Status", "Count", "Preview", "Error"]
    if [h.strip() for h in headers[: len(expected_headers)]] != expected_headers:
        raise ValueError(
            "Task_Notify_Logs 表头不符合预期（为保证不写坏表，已熔断）。\n"
            f"expected={expected_headers}\n"
            f"got={headers}\n"
        )

    existing_ids = set()
    for r in existing_rows:
        if r and len(r) >= 1 and r[0] is not None:
            s = str(r[0]).strip()
            if s:
                existing_ids.add(s)

    # 3) 解析 JSONL → 新 rows
    new_rows: List[List[Any]] = []
    for rec in _iter_jsonl(log_path):
        row = _record_to_row(rec, preview_max_chars=args.preview_max_chars)
        rid = str(row[0]).strip()
        if not rid or rid in existing_ids:
            continue
        new_rows.append(row)

    if not new_rows:
        print(json.dumps({"ok": True, "uploaded": 0, "reason": "no_new_rows"}, ensure_ascii=False))
        return 0

    # 4) 生成仅包含目标 sheet 的 xlsx：header + 新增（插入到顶部） + 旧数据
    merged_rows = new_rows + existing_rows

    tmp_dir = repo_root / ".tmp"
    tmp_xlsx = tmp_dir / f"task_notify_logs_upload_{int(time.time())}.xlsx"
    _write_single_sheet_xlsx(sheet_name=args.sheet_name, headers=expected_headers, rows=merged_rows, out_path=tmp_xlsx)

    # 5) 回写（update_sheet）
    _mcp_update_sheet(spreadsheet_url=spreadsheet_url, sheet_name=args.sheet_name, source_file_path=tmp_xlsx)

    # 6) RAW 原子锁：等 2s → 下载读回 → 校验 Run_ID 均存在
    time.sleep(2)
    verify_xlsx = _mcp_lark_download_spreadsheet(spreadsheet_url=spreadsheet_url)
    _, verify_rows = _load_sheet_rows(verify_xlsx, sheet_name=args.sheet_name)
    verify_ids = set()
    for r in verify_rows:
        if r and len(r) >= 1 and r[0] is not None:
            s = str(r[0]).strip()
            if s:
                verify_ids.add(s)

    inserted = [str(r[0]).strip() for r in new_rows]
    missing = [rid for rid in inserted if rid not in verify_ids]
    if missing:
        raise RuntimeError(
            "写后读校验失败：部分 Run_ID 未在表中读回（已熔断）。\n"
            f"missing={missing[:10]} (total_missing={len(missing)})"
        )

    print(json.dumps({"ok": True, "uploaded": len(new_rows), "log_file": str(log_path)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
