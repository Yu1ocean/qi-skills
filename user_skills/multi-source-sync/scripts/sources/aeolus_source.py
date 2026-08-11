#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Aeolus (风神) data source adapter for multi-source-sync.

底层调用 inner_skills/aeolus-platform-analysis 的 url_query.py / download_dashboard_data.py。
Region 从 URL 自动推断（CN/SG/VA/MYBD）。
"""

from __future__ import annotations

import csv
import json
import os
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse


WORKSPACE_ROOT = Path(__file__).resolve().parents[4]
AEOLUS_SKILL_DIR = WORKSPACE_ROOT / "inner_skills" / "aeolus-platform-analysis"


class AeolusSourceError(RuntimeError):
    """Raised when Aeolus data fetch fails or contract violated."""


def _infer_region_and_base_url(url: str, region_hint: Optional[str] = None) -> Tuple[str, str]:
    """Infer region + base URL from URL domain.

    Supported:
    - data.bytedance.net              → CN
    - aeolus-sg.bytedance.net         → SG
    - aeolus-va.bytedance.net         → VA
    - aeolus-va.tiktok-row.net       → VA
    - aeolus-mybd.bytedance.net       → MYBD
    """
    if region_hint:
        region = region_hint.upper().strip()
    else:
        host = urlparse(url).netloc.lower()
        if "aeolus-sg" in host:
            region = "SG"
        elif "aeolus-va" in host or "-va" in host or "virginia" in host or "tiktok-row.net" in host:
            region = "VA"
        elif "aeolus-mybd" in host or "-mybd" in host:
            region = "MYBD"
        else:
            region = "CN"

    region_to_base = {
        "CN": "https://data.bytedance.net",
        "SG": "https://aeolus-sg.bytedance.net",
        "VA": "https://aeolus-va.tiktok-row.net",
        "MYBD": "https://aeolus-mybd.sinf.net",
    }
    if region not in region_to_base:
        raise AeolusSourceError(f"Unsupported aeolus region: {region}")

    return region, region_to_base[region]


def _run_aeolus_script(script_name: str, args: List[str]) -> Tuple[str, str, int]:
    """Run an aeolus platform script and capture stdout/stderr/exit_code."""
    script_path = AEOLUS_SKILL_DIR / "scripts" / script_name
    if not script_path.exists():
        raise AeolusSourceError(f"Aeolus script not found: {script_path}")
    cmd = ["python3", str(script_path), *args]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    return proc.stdout, proc.stderr, proc.returncode


def _parse_url_query_output(stdout: str) -> Dict[str, Any]:
    """Parse url_query.py JSON output (may include rawFile pointer if truncated)."""
    text = stdout.strip()
    start = text.find("{")
    if start < 0:
        raise AeolusSourceError(f"url_query.py produced no JSON output: {stdout[:500]}")
    try:
        return json.loads(text[start:])
    except json.JSONDecodeError:
        # Try to find a complete JSON object from the end
        end = text.rfind("}")
        if end > start:
            try:
                return json.loads(text[start:end + 1])
            except json.JSONDecodeError as exc:
                raise AeolusSourceError(f"Unable to parse url_query.py JSON: {exc}\nstdout={stdout[:1000]}")
        raise AeolusSourceError(f"Unable to parse url_query.py JSON. stdout={stdout[:1000]}")


def _read_csv_rows(csv_path: Path) -> Tuple[List[str], List[List[str]]]:
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _read_xlsx_rows(xlsx_path: Path) -> Tuple[List[str], List[List[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency should exist in workspace
        raise AeolusSourceError(f"openpyxl is required to read xlsx exports: {exc}")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = ["" if v is None else str(v) for v in rows[0]]
    data_rows: List[List[str]] = []
    for raw_row in rows[1:]:
        data_rows.append(["" if v is None else str(v) for v in raw_row])
    return header, data_rows


def _read_tabular_rows(file_path: Path) -> Tuple[List[str], List[List[str]]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(file_path)
    if suffix == ".xlsx":
        return _read_xlsx_rows(file_path)
    raise AeolusSourceError(f"Unsupported download file format: {file_path}")


def fetch(source_config: Dict[str, Any]) -> Dict[str, Any]:
    """Fetch data from an Aeolus source configuration.

    Args:
        source_config: dict with keys:
            - url: aeolus URL (required)
            - region: optional (auto-inferred from URL)
            - download_full: bool, default False → 走 download_dashboard_data.py
            - filters: list of "字段=值" strings, optional
            - report_id: optional (dashboard mode)
            - id: optional readable identifier

    Returns:
        dict:
          - columns: List[str]
          - rows: List[List[Any]]        (2D data rows)
          - records_fetched: int
          - source_meta: dict            (region, url, mode)
    """
    if source_config.get("type") != "aeolus":
        raise AeolusSourceError(f"Not an aeolus source: type={source_config.get('type')}")

    url = source_config.get("url", "").strip()
    if not url:
        raise AeolusSourceError("aeolus source missing 'url'")

    region, base_url = _infer_region_and_base_url(url, source_config.get("region"))
    filters = source_config.get("filters") or []
    download_full = bool(source_config.get("download_full", False))
    report_id = source_config.get("report_id")
    field_map = source_config.get("field_map") or {}
    source_id = source_config.get("id", "aeolus")

    if download_full:
        try:
            downloaded = _fetch_via_download(
                url=url,
                base_url=base_url,
                region=region,
                filters=filters,
                report_id=report_id,
                source_id=source_id,
            )
            required_fields = [str(k) for k in field_map.keys() if str(k)]
            if required_fields and any(name not in downloaded.get("columns", []) for name in required_fields):
                missing = [name for name in required_fields if name not in downloaded.get("columns", [])]
                fallback = _fetch_via_url_query(
                    url=url,
                    base_url=base_url,
                    region=region,
                    filters=filters,
                    report_id=report_id,
                    source_id=source_id,
                )
                fallback.setdefault("source_meta", {})["download_full_requested"] = True
                fallback["source_meta"]["download_full_fallback_reason"] = (
                    f"download output missing declared field_map columns: {missing}"
                )
                fallback["source_meta"]["download_file"] = downloaded.get("source_meta", {}).get("file")
                fallback["source_meta"]["mode"] = "url_query_fallback"
                return fallback
            return downloaded
        except AeolusSourceError as exc:
            fallback = _fetch_via_url_query(
                url=url,
                base_url=base_url,
                region=region,
                filters=filters,
                report_id=report_id,
                source_id=source_id,
            )
            fallback.setdefault("source_meta", {})["download_full_requested"] = True
            fallback["source_meta"]["download_full_fallback_reason"] = str(exc)
            fallback["source_meta"]["mode"] = "url_query_fallback"
            return fallback
    return _fetch_via_url_query(
        url=url,
        base_url=base_url,
        region=region,
        filters=filters,
        report_id=report_id,
        source_id=source_id,
    )


def _normalize_column_names(columns: List[Any]) -> List[str]:
    normalized: List[str] = []
    for col in columns:
        if isinstance(col, dict):
            normalized.append(str(col.get("label") or col.get("name") or col.get("uniqueId") or ""))
        else:
            normalized.append(str(col))
    return normalized


def _fetch_via_url_query(
    url: str, base_url: str, region: str, filters: List[str], report_id: Optional[str], source_id: str
) -> Dict[str, Any]:
    args = ["--url", url, "--base-url", base_url]
    if report_id:
        args += ["--report-id", str(report_id)]
    for f in filters:
        args += ["--filters", f]

    stdout, stderr, exit_code = _run_aeolus_script("url_query.py", args)
    # Even on exit_code=1 (data_truncated), stdout is still valid JSON with rawFile pointer
    if not stdout.strip():
        raise AeolusSourceError(
            f"url_query.py returned no stdout. exit={exit_code}, stderr={stderr[:500]}"
        )
    payload = _parse_url_query_output(stdout)

    columns = _normalize_column_names(payload.get("columns") or [])
    rows = payload.get("rows") or []
    data_truncated = payload.get("data_truncated", False)
    data_file = payload.get("dataFile")
    server_total = payload.get("total")

    # If truncated, prefer the processed dataFile emitted by url_query.py
    if data_truncated and data_file and Path(data_file).exists():
        try:
            with Path(data_file).open("r", encoding="utf-8") as f:
                full_payload = json.load(f)
            columns = _normalize_column_names(full_payload.get("columns") or columns)
            rows = full_payload.get("rows") or rows
            server_total = full_payload.get("total", server_total)
        except (json.JSONDecodeError, OSError) as exc:
            raise AeolusSourceError(
                f"Failed to read dataFile for full data: {exc}. data_file={data_file}"
            )

    if not columns:
        raise AeolusSourceError(
            f"aeolus url_query returned no columns. exit={exit_code}, payload keys={list(payload.keys())}"
        )

    # Safety guard: for paginated/pivot DataQuery responses, url_query may expose
    # server_total > extracted rows. Do not silently write partial data.
    try:
        server_total_int = int(server_total) if server_total is not None else None
    except (TypeError, ValueError):
        server_total_int = None
    if server_total_int is not None and len(rows) < server_total_int:
        raise AeolusSourceError(
            f"Aeolus pagination incomplete: fetched_rows={len(rows)} < server_total={server_total_int}. "
            f"data_file={data_file or ''}. Refusing to write partial data."
        )

    return {
        "columns": columns,
        "rows": rows,
        "records_fetched": len(rows),
        "source_meta": {
            "id": source_id,
            "region": region,
            "url": url,
            "mode": "url_query",
            "data_truncated": data_truncated,
            "data_file": data_file,
            "server_total": server_total_int,
        },
    }


def _fetch_via_download(
    url: str, base_url: str, region: str, filters: List[str], report_id: Optional[str], source_id: str
) -> Dict[str, Any]:
    # Prefer the single-viz `--chart-id` path over `--url` because:
    #   - `--url` auto-extracts dashboard_id + rid → drives the dashboard route, which
    #     requires dashboardAndSheet / sheetReport permissions. When only chart-level
    #     access is granted (common for shared dataQuery links), the dashboard path
    #     fails with 403 while the single-viz path (dataMart/report + vizQuery/download)
    #     succeeds.
    #   - The pivot_table server export via vizQuery/download returns the full row set
    #     (e.g. 542 real data rows + Sum), whereas `url_query.py`'s
    #     `extract_vizquery_data` only flattens partial pivot cells (~350 duplicates in
    #     the observed case). This is the fix for the 350→542 pagination gap.
    effective_chart_id = report_id
    if not effective_chart_id and url:
        try:
            from url_query import parse_aeolus_url  # type: ignore
        except ImportError:
            import sys as _sys
            _sys.path.insert(0, str(AEOLUS_SKILL_DIR / "scripts"))
            from url_query import parse_aeolus_url  # type: ignore
        parsed = parse_aeolus_url(url)
        rid = parsed.get("reportId") or 0
        if rid:
            effective_chart_id = str(rid)

    args: List[str] = ["--base-url", base_url]
    if effective_chart_id:
        args += ["--chart-id", str(effective_chart_id)]
    else:
        # Fall back to raw URL parsing inside download script (dashboard route).
        args += ["--url", url]
    for f in filters:
        args += ["--filters", f]

    # The Aeolus vizQuery/download endpoint is asynchronously polled; it occasionally
    # returns `aeolus/unknown` after 3 poll cycles (~6s). Retry the whole download a
    # couple of times before giving up, since the backend can recover on the next call.
    max_attempts = 3
    last_stdout = last_stderr = ""
    last_exit = 0
    for attempt in range(1, max_attempts + 1):
        stdout, stderr, exit_code = _run_aeolus_script("download_dashboard_data.py", args)
        last_stdout, last_stderr, last_exit = stdout, stderr, exit_code
        if exit_code == 0 and stdout.strip():
            break
        if attempt < max_attempts:
            import time as _time
            print(
                f"[aeolus_source] download attempt {attempt}/{max_attempts} failed "
                f"(exit={exit_code}); retrying in 5s",
                flush=True,
            )
            _time.sleep(5)

    stdout, stderr, exit_code = last_stdout, last_stderr, last_exit
    if exit_code != 0:
        raise AeolusSourceError(
            f"download_dashboard_data.py failed. exit={exit_code}\nstdout={stdout[:800]}\nstderr={stderr[:800]}"
        )
    # download_dashboard_data.py outputs JSON with a `file` path pointing to CSV
    text = stdout.strip()
    start = text.find("{")
    if start < 0:
        raise AeolusSourceError(f"download_dashboard_data.py produced no JSON: {stdout[:500]}")
    try:
        payload = json.loads(text[start:])
    except json.JSONDecodeError:
        end = text.rfind("}")
        payload = json.loads(text[start:end + 1])

    data_file = payload.get("file") or payload.get("csvFile") or payload.get("filePath")
    if not data_file or not Path(data_file).exists():
        raise AeolusSourceError(
            f"download_dashboard_data.py did not produce a readable file. payload={payload}"
        )
    columns, rows = _read_tabular_rows(Path(data_file))
    if not columns:
        raise AeolusSourceError(f"Downloaded file has no header: {data_file}")

    return {
        "columns": columns,
        "rows": rows,
        "records_fetched": len(rows),
        "source_meta": {
            "id": source_id,
            "region": region,
            "url": url,
            "mode": "download_full",
            "file": data_file,
            "download_payload": {
                "rowCount": payload.get("rowCount"),
                "displayType": payload.get("displayType"),
                "limitReached": payload.get("limitReached"),
            },
        },
    }
