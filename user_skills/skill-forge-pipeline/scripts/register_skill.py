#!/usr/bin/env python3
import argparse
import datetime
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
import zipfile
from pathlib import Path
from typing import Any, Callable, Dict, Optional, Tuple
from urllib.parse import quote, urlparse

import requests

# --- 三分区（Zone）策略执行器（V5.23） ---
# 飞书说明文档写入必须区分 Overwrite / Preserve / Append 三个区域：
# 人工沉淀的使用案例与踩坑记录（Preserve Zone）永不被 forge 覆盖。
sys.path.insert(0, str(Path(__file__).resolve().parent))
from doc_zone_manager import (  # noqa: E402
    APPEND_ANCHOR_TITLE,
    PRESERVE_ANCHOR_TITLE,
    build_changelog_entry_from_skill_md,
    build_new_doc_markdown,
    sync_doc_zones,
)

# --- Wiki「技能存量清单」Upsert 钩子（V5.24） ---
# 历史缺口：forge 只同步「专属技能清单」Sheet，Wiki 上给人看的存量清单表无人维护
# （三个月未更新、覆盖率一度只有 31.4%）。本钩子把该表纳入 forge 闭环。
from wiki_skill_list_sync import (  # noqa: E402
    DEFAULT_WIKI_URL as DEFAULT_WIKI_REGISTRY_URL,
    WikiSyncError,
    sync_wiki_skill_list,
)

DEFAULT_EMAIL = "yuqinan@bytedance.com"
DEFAULT_OPEN_API_BASE = "https://fsopen.bytedance.net"
DEFAULT_UPLOAD_API_BASE = "https://open.feishu.cn"
DEFAULT_SKILL_INVENTORY_URL = "https://bytedance.sg.larkoffice.com/sheets/ECQ0sDwmbhDex9tcUSjlkU7Bgdh"
DEFAULT_SKILL_INVENTORY_SHEET_NAME = "专属技能清单"
DEFAULT_SKILL_INVENTORY_UPDATED_AT_FORMATTER = "yyyy/MM/dd"

# --- SSOT (Single Source of Truth) Version Sync Bus ---
# We treat `version:` in the target skill's SKILL.md frontmatter as SSOT.
# During Archive, we bump version (Major: +1.0 / Minor: +0.1) and sync it to:
# 1) local SKILL.md
# 2) Feishu skill inventory sheet (full-row upsert via omni-asset-archiver)
# 3) (optional) Feishu skill doc version marker
#
# 【Initial Version Policy】
# Per user (奇楠) directive (2026-05-20):
# Brand-new skills MUST debut at version 1.1, not 0.x.
# Scaffolding tools (e.g. aime-skill-creator) typically seed SKILL.md with
# `version: 0.x` placeholders. When register_skill.py detects this
# "fresh skill" state during its first run, it will force-set the
# published version to DEFAULT_INITIAL_VERSION (1.1) instead of
# performing the usual +0.1 bump within the 0.x range.
DEFAULT_INITIAL_VERSION = "1.1"

# --- Cloud Publish (V5.19) 合规默认值 ---
# 可见性默认「个人可见」，绝不默认扩大到空间可见；
# enable-by-default 默认 false，只有用户显式要求全员可用时才允许开启。
DEFAULT_CLOUD_SCOPE = "user"
DEFAULT_CLOUD_ENABLE_BY_DEFAULT = False
DEFAULT_CLOUD_PUBLISH_DLQ = ".ephemeral_pool/cloud_publish_failures.jsonl"
DEFAULT_SSOT_SPREADSHEET_TOKEN = "ECQ0sDwmbhDex9tcUSjlkU7Bgdh"
DEFAULT_SSOT_SHEET_NAME = "专属技能清单"
DEFAULT_SSOT_SHEET_NAME_FALLBACKS = ["专属技能清单_Sheet"]
DEFAULT_WIKI_NODE_TOKEN = "GU0ewkyaGi4i5nkwBtNcM3aPn9g"
DEFAULT_WIKI_URL = f"https://bytedance.larkoffice.com/wiki/{DEFAULT_WIKI_NODE_TOKEN}"

DEFAULT_SSOT_VERSION_HEADERS = ["版本号", "版本", "version", "Version"]
DEFAULT_SSOT_ID_HEADERS = ["技能编号", "Skill ID", "skill_id", "SkillID", "ID", "编号", "包 ID"]
DEFAULT_SSOT_NAME_HEADERS = ["技能名称", "技能名", "name", "Name"]

REQUIRED_DOC_TEMPLATE_MARKERS = [
    "🔑 触发词",
    "📖 案例实录",
]


class GuardrailViolation(RuntimeError):
    """Hard-stop exception for guardrail violations."""


def validate_doc_template_markers(doc_url: str) -> None:
    """L1/L2 doc-structure validator for skill docs.

    We validate markers AFTER downloading the doc via Lark MCP.
    """

    markdown_path = download_doc_markdown(doc_url)
    content = markdown_path.read_text(encoding="utf-8", errors="ignore")

    missing = [m for m in REQUIRED_DOC_TEMPLATE_MARKERS if m not in content]
    if missing:
        raise GuardrailViolation(
            "Skill doc template markers missing: " + ", ".join(missing)
        )


def run_cda_guardrails_selfcheck(skill_dir: Path, risk: str = "auto") -> str:
    """CDA Guardrails checkpoint: fail-fast before side effects."""

    selfcheck_script = Path(__file__).resolve().parent / "cda_guardrails_selfcheck.py"
    if not selfcheck_script.exists():
        raise FileNotFoundError(f"CDA selfcheck script not found: {selfcheck_script}")

    return run_subprocess(
        [
            "python3",
            str(selfcheck_script),
            "--skill-dir",
            str(skill_dir),
            "--risk",
            risk,
        ],
        "CDA-Guardrails-Selfcheck",
    )


def run_post_forge_git_push(workspace_root: Path, skill_name: str, skill_version: str) -> str:
    """Run the qi-skills post-forge git sync hook after a successful archive."""

    if os.environ.get("SKIP_POST_FORGE_GIT_PUSH") == "1":
        return "⏭️ post-forge git push skipped by SKIP_POST_FORGE_GIT_PUSH=1"

    hook_path = workspace_root / "user_skills" / "scripts" / "post_forge_git_push.sh"
    if not hook_path.exists():
        raise FileNotFoundError(f"Post-forge git push hook not found: {hook_path}")

    return run_subprocess(
        [
            "bash",
            str(hook_path),
            skill_name or "unknown-skill",
            skill_version or "latest",
        ],
        "post-forge git push hook",
    )


def run_cloud_publish(
    skill_dir: Path,
    *,
    version: str,
    cloud_scope: str = DEFAULT_CLOUD_SCOPE,
    enable_by_default: bool = False,
    workspace_root: Optional[Path] = None,
) -> Dict[str, Any]:
    """Cloud Publish（第四步）：把技能真正上传到 Aime 云端。

    根因：forge 流水线此前只 push 到 GitHub，技能始终停留在本地草稿态，
    每次都要人工去点「上传到云端」按钮。本步骤把它固化进流水线。

    走 scripts/cloud_publish.py：
      draft 前置 -> `aime skill upload` -> `aime skill list` 云端回读断言。
    失败不静默：标记 SKILL.md「需手动上传」+ 落死信队列 + 输出醒目 ERROR。
    """

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from cloud_publish import cloud_publish_with_fallback  # noqa: PLC0415

    return cloud_publish_with_fallback(
        skill_dir,
        version=version,
        cloud_scope=cloud_scope,
        enable_by_default=enable_by_default,
        workspace_root=workspace_root,
    )


def get_workspace_root() -> Path:
    env_path = os.environ.get("IRIS_WORKSPACE_PATH")
    if env_path:
        return Path(env_path).resolve()
    return Path(__file__).resolve().parents[3]


FEISHU_DATE_EPOCH = datetime.datetime(1899, 12, 30)
FEISHU_DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%d",
    "%Y/%m/%d",
)


def maybe_to_feishu_datetime_serial(value: Any) -> Any:
    if isinstance(value, (int, float)):
        return value
    text = str(value or "").strip()
    if not text:
        return value
    for fmt in FEISHU_DATETIME_FORMATS:
        try:
            parsed = datetime.datetime.strptime(text, fmt)
            return round((parsed - FEISHU_DATE_EPOCH).total_seconds() / 86400, 10)
        except ValueError:
            continue
    return value


def get_raw_cloud_jwt() -> str:
    raw = (os.environ.get("AIME_USER_CLOUD_JWT") or "").strip()
    if not raw:
        raise RuntimeError(
            "Missing env var AIME_USER_CLOUD_JWT. Run this script with include_secrets=true."
        )
    return raw


def ensure_bearer_token() -> str:
    raw = get_raw_cloud_jwt()
    if raw.startswith("Bearer "):
        return raw
    return f"Bearer {raw}"


def build_headers(bearer_token: str, content_type: Optional[str] = None) -> Dict[str, str]:
    headers = {"Authorization": bearer_token}
    if content_type:
        headers["Content-Type"] = content_type
    if bearer_token.startswith("Bearer "):
        headers["Cookie"] = f"session={bearer_token.split(' ', 1)[1]}"
    return headers


def call_with_retry(action: str, callback: Callable[[], Any], attempts: int = 3, delay: float = 2.0) -> Any:
    last_error: Optional[Exception] = None
    for attempt in range(1, attempts + 1):
        try:
            return callback()
        except Exception as exc:  # noqa: PERF203
            last_error = exc
            if attempt == attempts:
                break
            print(f"⚠️ {action} failed on attempt {attempt}/{attempts}: {exc}")
            time.sleep(delay)
    raise RuntimeError(f"{action} failed after {attempts} attempts: {last_error}")


ZIP_SKIP_DIR_NAMES = {
    ".git",
    "__pycache__",
    ".DS_Store",
    ".tmp",
    ".runtime",
    "downloads",
    "snapshots",
    "output",
    "outputs",
}
ZIP_SKIP_SUFFIXES = {".zip", ".mp4", ".part", ".pyc"}
# `download_doc_markdown()` 会把 <token>.lark.md 落到 cwd（常常就是技能目录），不入包。
ZIP_SKIP_NAME_SUFFIXES = (".lark.md",)
ZIP_SIZE_WARN_BYTES = 50 * 1024 * 1024


def create_skill_zip(skill_dir: Path, output_zip: Path) -> Path:
    if not skill_dir.is_dir():
        raise FileNotFoundError(f"Skill directory not found: {skill_dir}")

    output_zip.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(output_zip, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(skill_dir.rglob("*")):
            if any(part in ZIP_SKIP_DIR_NAMES for part in path.parts):
                continue
            if path == output_zip:
                continue
            if path.is_file() and path.name.endswith(ZIP_SKIP_NAME_SUFFIXES):
                continue
            if path.is_file() and path.suffix.lower() in ZIP_SKIP_SUFFIXES:
                # 运行时产物（缓存 zip / 媒体 / 下载残片）一律不入包，
                # 否则技能包会被历史缓存拖成几百 MB（真机曾出现 245MB ZIP 被 push 拦截）。
                continue
            arcname = Path(skill_dir.name) / path.relative_to(skill_dir)
            archive.write(path, arcname)

    size = output_zip.stat().st_size
    if size > ZIP_SIZE_WARN_BYTES:
        print(
            f"⚠️ 技能包体积异常：{size / 1024 / 1024:.1f}MB > 50MB，"
            "请检查是否有运行时产物混入（.tmp / downloads / snapshots / 媒体文件）。"
        )
    return output_zip.resolve()


def get_root_folder_token(bearer_token: str, open_api_base: str) -> str:
    url = f"{open_api_base}/open-apis/drive/explorer/v2/root_folder/meta"
    response = requests.get(url, headers=build_headers(bearer_token), timeout=30)
    data = response.json()
    if response.status_code != 200 or data.get("code") != 0:
        raise RuntimeError(f"get root folder failed: http={response.status_code}, resp={data}")
    token = data.get("data", {}).get("token")
    if not token:
        raise RuntimeError(f"root folder token missing in response: {data}")
    return token


def upload_zip_to_drive(
    bearer_token: str,
    open_api_base: str,
    file_path: Path,
    parent_node: str = "",
) -> str:
    url = f"{open_api_base}/open-apis/drive/v1/files/upload_all"
    with file_path.open("rb") as file_obj:
        response = requests.post(
            url,
            headers=build_headers(bearer_token),
            data={
                "file_name": file_path.name,
                "parent_type": "explorer",
                "parent_node": parent_node,
                "size": str(file_path.stat().st_size),
            },
            files={"file": (file_path.name, file_obj)},
            timeout=300,
        )
    data = response.json()
    if response.status_code != 200 or data.get("code") != 0:
        raise RuntimeError(f"drive upload failed: http={response.status_code}, resp={data}")
    file_token = data.get("data", {}).get("file_token")
    if not file_token:
        raise RuntimeError(f"file_token missing after drive upload: {data}")
    return file_token


def add_permission_member(
    bearer_token: str,
    open_api_base: str,
    file_token: str,
    email: str,
    perm: str = "full_access",
) -> Dict[str, Any]:
    url = f"{open_api_base}/open-apis/drive/v1/permissions/{file_token}/members"
    response = requests.post(
        url,
        params={"type": "file", "need_notification": "false"},
        headers=build_headers(bearer_token, "application/json; charset=utf-8"),
        json={
            "member_type": "email",
            "member_id": email,
            "perm": perm,
            "type": "user",
        },
        timeout=30,
    )
    return {"http_status": response.status_code, "data": response.json()}


def update_permission_member(
    bearer_token: str,
    open_api_base: str,
    file_token: str,
    email: str,
    perm: str = "full_access",
) -> Dict[str, Any]:
    member_id = quote(email, safe="")
    url = f"{open_api_base}/open-apis/drive/v1/permissions/{file_token}/members/{member_id}"
    response = requests.put(
        url,
        params={"type": "file", "need_notification": "false"},
        headers=build_headers(bearer_token, "application/json; charset=utf-8"),
        json={"member_type": "email", "perm": perm, "type": "user"},
        timeout=30,
    )
    return {"http_status": response.status_code, "data": response.json()}


def ensure_drive_permission(
    bearer_token: str,
    open_api_base: str,
    file_token: str,
    email: str,
    perm: str = "full_access",
) -> None:
    """Legacy compatibility shim for older callers.

    Deprecated: do not use JWT/OpenAPI permission grants. Prefer
    `ensure_drive_asset_access_via_mcp()` with a concrete /file/ URL.
    """

    del bearer_token, open_api_base
    drive_file_url = build_drive_file_url("https://bytedance.larkoffice.com/docx/placeholder", file_token)
    ensure_drive_asset_access_via_mcp(drive_file_url, email=email, perm=perm)


def build_drive_file_url(doc_url: str, file_token: str) -> str:
    parsed = urlparse(doc_url)
    scheme = parsed.scheme or "https"
    netloc = parsed.netloc or "bytedance.larkoffice.com"
    return f"{scheme}://{netloc}/file/{file_token}"


def ensure_drive_asset_access_via_mcp(drive_file_url: str, email: str, perm: str = "full_access") -> str:
    workspace_root = get_workspace_root()
    repair_script = workspace_root / "user_skills/feishu-doc-writing-guide/scripts/grant_doc_permissions.py"
    if not repair_script.exists():
        raise FileNotFoundError(f"grant_doc_permissions wrapper not found: {repair_script}")

    # V5.18: NO silent-WARNING fallback. The legacy `move_lark_doc` shim was removed from the
    # runtime, and swallowing its FileNotFoundError turned this step into a permanent fake
    # success (users received ZIP assets they could view but not manage). The grant now runs
    # through `lark-cli drive +member-add` + `+member-list` RAW assertion; any failure must
    # circuit-break the release per the anti-fake-success rule.
    return run_subprocess(
        [
            "python3",
            str(repair_script),
            drive_file_url,
            "--email",
            email,
            "--perm",
            perm,
        ],
        "grant drive asset access via lark-cli drive +member-add (with member-list RAW assertion)",
    )


def run_subprocess(command: list[str], action: str) -> str:
    result = subprocess.run(command, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(
            f"{action} failed with exit code {result.returncode}\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )
    return result.stdout.strip()


# -----------------------------
# SSOT: Version Sync Bus
# -----------------------------

def ensure_bytedcli_auth() -> None:
    """Authenticate via bytedcli (required before MCP-based Feishu ops).

    NOTE: this expects the script to be executed with `include_secrets=true`
    so that IRIS_USER_CLOUD_JWT is available.
    """

    workspace_root = get_workspace_root()
    auth_script = workspace_root / "inner_skills/bytedcli-auth/scripts/bytedcli_auth.sh"
    if not auth_script.exists():
        raise FileNotFoundError(f"bytedcli auth script not found: {auth_script}")

    print("🔐 Ensuring bytedcli login (bytedcli-auth)...")
    run_subprocess(["bash", str(auth_script)], "bytedcli-auth")


def _parse_version(raw: str) -> Tuple[int, int, Optional[int]]:
    """Parse a version string into (major, minor, patch|None).

    V5.20 fix: the legacy `_normalize_version_to_int_pair()` truncated the patch
    segment (`v1.6.1` -> `1.6`), so every forge run silently downgraded
    three-segment versions. Patch is now preserved verbatim; `None` means the
    source version was two-segment and must stay two-segment.
    """

    text = (raw or "").strip().lstrip("vV").strip()
    m = re.match(r"^(\d+)\.(\d+)(?:\.(\d+))?$", text)
    if not m:
        raise ValueError(f"Unsupported version format: {raw!r} (expected X.Y or X.Y.Z)")
    patch = int(m.group(3)) if m.group(3) is not None else None
    return int(m.group(1)), int(m.group(2)), patch


def _format_version(major: int, minor: int, patch: Optional[int] = None) -> str:
    if patch is None:
        return f"{major}.{minor}"
    return f"{major}.{minor}.{patch}"


def normalize_version_text(raw: str) -> str:
    """Normalize any version literal while PRESERVING the patch segment."""

    return _format_version(*_parse_version(raw))


def read_skill_version_from_skill_md(skill_dir: Path) -> str:
    """Read SSOT version from target skill's SKILL.md frontmatter."""

    skill_md = skill_dir / "SKILL.md"
    if not skill_md.exists():
        raise FileNotFoundError(f"SKILL.md not found: {skill_md}")

    content = skill_md.read_text(encoding="utf-8", errors="ignore")
    m = re.search(r"(?m)^version:\s*([^\n]+)\s*$", content)
    if not m:
        raise ValueError(f"version field not found in SKILL.md: {skill_md}")

    return normalize_version_text(m.group(1).strip())


def bump_version(current_version: str, bump_type: str) -> str:
    """Bump version while preserving the three-segment shape when present.

    - major: X.Y[.Z] -> X+1.0[.0]
    - minor: X.Y[.Z] -> X.Y+1[.0]
    - patch: X.Y[.Z] -> X.Y.Z+1  (two-segment input is treated as .0)
    """

    major, minor, patch = _parse_version(current_version)
    bump_type = (bump_type or "").strip().lower()

    if bump_type == "major":
        return _format_version(major + 1, 0, 0 if patch is not None else None)
    if bump_type == "minor":
        return _format_version(major, minor + 1, 0 if patch is not None else None)
    if bump_type == "patch":
        return _format_version(major, minor, (patch or 0) + 1)

    raise ValueError(f"Unsupported bump type: {bump_type!r} (expected major/minor/patch)")


def is_initial_version(current_version: str) -> bool:
    """Return True if the current SKILL.md version is still in the
    0.x scaffold range, meaning this is the very first publish of the skill.
    """

    try:
        major, _minor, _patch = _parse_version(current_version)
    except ValueError:
        return False
    return major == 0


def write_skill_md_version(skill_dir: Path, new_version: str) -> None:
    """Sync point #1: overwrite SSOT version in local SKILL.md."""

    skill_md = skill_dir / "SKILL.md"
    content = skill_md.read_text(encoding="utf-8", errors="ignore")

    # only replace the frontmatter version key; keep other content intact.
    replaced, n = re.subn(
        r"(?m)^version:\s*([^\n]+)\s*$",
        f"version: {new_version}",
        content,
        count=1,
    )
    if n != 1:
        raise RuntimeError(f"Unable to update version field in SKILL.md: {skill_md}")

    skill_md.write_text(replaced, encoding="utf-8")


def build_ssot_spreadsheet_url(spreadsheet_token: str) -> str:
    token = (spreadsheet_token or "").strip()
    if token.startswith("http://") or token.startswith("https://"):
        return token
    return f"https://bytedance.larkoffice.com/sheets/{token}"


def mcp_lark_download(document_url: str) -> list[Path]:
    workspace_root = get_workspace_root()
    download_script = workspace_root / "inner_skills/lark_download/lark_download.py"
    if not download_script.exists():
        raise FileNotFoundError(f"Lark download script not found: {download_script}")

    try:
        output = run_subprocess(
            [
                "python3",
                str(download_script),
                json.dumps({"document_url": document_url}, ensure_ascii=False),
            ],
            "lark download",
        )
    except RuntimeError as exc:
        if "toolset lark_download not found" not in str(exc):
            raise
        token = document_url.rstrip("/").split("/")[-1].split("?")[0]
        # 落点固定为本技能目录自身（由脚本位置推导），避免硬编码技能名导致改名后失效
        fallback_path = Path(__file__).resolve().parents[1] / f"{token}.lark.md"
        fetch_output = run_subprocess(
            [
                "lark-cli",
                "docs",
                "+fetch",
                "--api-version",
                "v2",
                "--as",
                "user",
                "--doc",
                document_url,
                "--doc-format",
                "markdown",
                "--format",
                "json",
            ],
            "lark-cli docs fetch fallback",
        )
        try:
            payload = json.loads(fetch_output[fetch_output.find("{"):])
            content = payload["data"]["document"]["content"]
        except (json.JSONDecodeError, KeyError, TypeError) as parse_exc:
            raise RuntimeError(f"Unable to parse lark-cli docs +fetch output: {fetch_output}") from parse_exc
        fallback_path.write_text(content, encoding="utf-8")
        return [fallback_path.resolve()]

    paths = [Path(p).resolve() for p in re.findall(r'file_path:\s*"([^"]+)"', output)]
    if not paths:
        paths = [Path(p).resolve() for p in re.findall(r'(/[^\s\"]+\.(?:lark\.md|xlsx|xls|md))', output)]

    if not paths:
        try:
            parsed = json.loads(output)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, list):
            paths = [Path(str(p)).resolve() for p in parsed if str(p).strip()]
        elif isinstance(parsed, str) and parsed.strip():
            paths = [Path(parsed).resolve()]

    if not paths:
        raise RuntimeError(f"Unable to parse downloaded file paths from output: {output}")

    return paths


def _pick_first_xlsx(paths: list[Path]) -> Path:
    for p in paths:
        if p.suffix.lower() in {".xlsx", ".xls"}:
            return p
    raise RuntimeError(f"No xlsx file found in downloaded paths: {paths}")


def _normalize_header_value(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_skill_name_cell(v: Any) -> str:
    """Normalize skill name cell.

    In the inventory sheet, "技能名称" is often a HYPERLINK formula like:
    =HYPERLINK("url", "name")
    We extract the visible name part for matching.
    """

    raw = _normalize_header_value(v)
    if not raw:
        return ""

    if raw.lstrip().upper().startswith("=HYPERLINK"):
        parts = re.findall(r'"([^"]+)"', raw)
        if len(parts) >= 2:
            return parts[-1].strip()
    return raw


def _find_header_col_index(headers: list[str], candidates: list[str]) -> Optional[int]:
    lowered = [h.lower() for h in headers]
    for c in candidates:
        if c.lower() in lowered:
            return lowered.index(c.lower())
    return None


def _choose_sheet_name(workbook_sheetnames: list[str], preferred: str) -> str:
    if preferred in workbook_sheetnames:
        return preferred
    for alt in DEFAULT_SSOT_SHEET_NAME_FALLBACKS:
        if alt in workbook_sheetnames:
            return alt
    raise RuntimeError(
        "SSOT sheet not found. "
        f"preferred={preferred!r}, available={workbook_sheetnames!r}, fallbacks={DEFAULT_SSOT_SHEET_NAME_FALLBACKS!r}"
    )


def update_xlsx_version_cell(
    *,
    source_xlsx: Path,
    sheet_name: str,
    skill_id: str,
    skill_name: str,
    new_version: str,
) -> Tuple[Path, str]:
    """Update the Version cell in a local xlsx copy.

    Returns: (updated_xlsx_path, debug_location)
    """

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: PERF203
        raise RuntimeError(
            "Missing dependency openpyxl. Please install it (pip install openpyxl) before running this script."
        ) from exc

    wb = load_workbook(source_xlsx)
    real_sheet_name = _choose_sheet_name(wb.sheetnames, sheet_name)
    ws = wb[real_sheet_name]

    # assume the first row is header
    header_row = [
        _normalize_header_value(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)
    ]

    id_col = _find_header_col_index(header_row, DEFAULT_SSOT_ID_HEADERS)
    name_col = _find_header_col_index(header_row, DEFAULT_SSOT_NAME_HEADERS)
    version_col = _find_header_col_index(header_row, DEFAULT_SSOT_VERSION_HEADERS)

    if version_col is None:
        raise RuntimeError(
            f"Unable to locate version column by headers {DEFAULT_SSOT_VERSION_HEADERS!r}. "
            f"sheet={real_sheet_name!r}, header_row={header_row!r}"
        )

    # openpyxl column index is 1-based
    version_col_1b = version_col + 1
    id_col_1b = id_col + 1 if id_col is not None else None
    name_col_1b = name_col + 1 if name_col is not None else None

    matched_row: Optional[int] = None
    matched_by = ""
    for r in range(2, ws.max_row + 1):
        rid = _normalize_header_value(ws.cell(row=r, column=id_col_1b).value) if id_col_1b else ""
        rname = _normalize_skill_name_cell(ws.cell(row=r, column=name_col_1b).value) if name_col_1b else ""

        if skill_id and rid and rid.strip() == skill_id.strip():
            matched_row = r
            matched_by = "skill_id"
            break
        if skill_name and rname and rname.strip() == skill_name.strip():
            matched_row = r
            matched_by = "skill_name"
            break

    if not matched_row:
        raise RuntimeError(
            "Unable to locate target skill row in SSOT sheet. "
            f"skill_id={skill_id!r}, skill_name={skill_name!r}, sheet={real_sheet_name!r}"
        )

    ws.cell(row=matched_row, column=version_col_1b).value = new_version

    updated_path = source_xlsx.parent / f"ssot_updated_{uuid.uuid4().hex}.xlsx"
    wb.save(updated_path)

    debug_loc = f"{real_sheet_name}!R{matched_row}C{version_col_1b} (matched_by={matched_by})"
    return updated_path, debug_loc


def sync_skill_inventory_via_omni(metadata: Dict[str, Any]) -> Dict[str, Any]:
    """Sync point #2: upsert the full skill row into Feishu inventory via omni-asset-archiver."""

    workspace_root = get_workspace_root()
    archiver_script = workspace_root / "user_skills/omni-asset-archiver/scripts/archiver_driver.py"
    if not archiver_script.exists():
        raise FileNotFoundError(f"Omni asset archiver script not found: {archiver_script}")

    payload = {
        "asset_type": "skill_inventory",
        "target_route_key": "skill_inventory",
        "skill_id": metadata.get("skill_id", ""),
        "title": metadata.get("name", ""),
        "doc_url": metadata.get("doc_link", ""),
        "description": metadata.get("description", ""),
        "version": metadata.get("version", ""),
        "updated_at": maybe_to_feishu_datetime_serial(metadata.get("updated_at", "")),
    }

    required_fields = ["skill_id", "title", "doc_url", "description", "version", "updated_at"]
    missing_fields = [field for field in required_fields if not str(payload.get(field, "")).strip()]
    if missing_fields:
        raise RuntimeError(
            "Omni skill inventory sync payload missing required fields: " + ", ".join(missing_fields)
        )

    print("🚌 Syncing skill inventory row via omni-asset-archiver...")
    output = run_subprocess(
        [
            "python3",
            str(archiver_script),
            "--payload-json",
            json.dumps(payload, ensure_ascii=False),
        ],
        "omni-asset-archiver skill inventory upsert",
    )

    json_match = re.search(r"(\{[\s\S]*\})\s*$", output.strip())
    if not json_match:
        raise RuntimeError(f"Unable to parse omni-asset-archiver output: {output}")

    result = json.loads(json_match.group(1))
    status = str(result.get("status", "")).strip().lower()
    if status not in {"appended", "updated", "dry_run"}:
        raise RuntimeError(f"Unexpected omni-asset-archiver status: {result}")

    print(
        "✅ Skill inventory sync finished via omni-asset-archiver: "
        f"status={result.get('status')}, sheet={result.get('sheet_name')}, row={result.get('row_number')}"
    )
    return result


def run_wiki_skill_list_sync(metadata: Dict[str, Any], wiki_registry_url: str = "") -> Dict[str, Any]:
    """Sync point #3（V5.24）：把本次 forge 的技能 upsert 进 Wiki「技能存量清单」表格。

    定位：**增强项**。Wiki 表是给人扫读的导航索引，权威版本号仍在 SKILL.md（SSOT）
    与「专属技能清单」Sheet。因此整段用 try/except 包裹，失败只打 WARNING 不阻断 forge。

    ⚠️ 边界：这里的「不阻断」仅限本步骤。Wiki 同步**内部**的写后回读断言失败会 raise
    WikiSyncError，并在此被如实降级为 WARNING —— 绝不允许把断言失败包装成 success。
    """

    report: Dict[str, Any] = {"status": "", "detail": {}}
    try:
        detail = sync_wiki_skill_list(
            metadata.get("name", ""),
            doc_url=metadata.get("doc_link", ""),
            desc=metadata.get("description", ""),
            version=metadata.get("version", ""),
            wiki_url=wiki_registry_url or DEFAULT_WIKI_REGISTRY_URL,
        )
        report["status"] = "success"
        report["detail"] = detail
        print(
            "✅ Wiki 技能存量清单同步完成："
            f"action={detail.get('action')}, rows={detail.get('rows_after')}, "
            f"assert={detail.get('assert')}"
        )
    except (WikiSyncError, Exception) as exc:  # noqa: BLE001 - 增强项，禁止阻断主流程
        report["status"] = "failed"
        report["error"] = str(exc)
        print(f"⚠️ WARNING: Wiki sync failed: {exc}")
    return report


def _run_lark_sheets_json(args: list[str], step_name: str) -> dict:
    workspace_root = get_workspace_root()
    cli_path = workspace_root / "inner_skills/lark-sheets/bin/lark-sheets-cli"
    if not cli_path.exists():
        raise FileNotFoundError(f"lark-sheets cli not found: {cli_path}")
    output = run_subprocess([str(cli_path), *args], step_name)
    json_match = re.search(r"(\{[\s\S]*\})", output)
    if not json_match:
        raise RuntimeError(f"Unable to parse lark-sheets output for {step_name}: {output}")
    try:
        return json.loads(json_match.group(1))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Unable to decode lark-sheets json for {step_name}: {output}") from exc


def _resolve_skill_inventory_sheet_id() -> str:
    info = _run_lark_sheets_json(
        ["sheets", "+info", "--url", DEFAULT_SKILL_INVENTORY_URL],
        "resolve skill inventory sheet id",
    )
    sheets = info.get("data", {}).get("sheets", {}).get("sheets", [])
    for sheet in sheets:
        if sheet.get("title") == DEFAULT_SKILL_INVENTORY_SHEET_NAME:
            return str(sheet.get("sheet_id") or "").strip()
    raise RuntimeError(f"Unable to resolve sheet id for {DEFAULT_SKILL_INVENTORY_SHEET_NAME}")


def ensure_skill_inventory_updated_at_formatter(row_number: int, expected_text: str) -> None:
    if row_number <= 0:
        raise ValueError(f"Invalid row_number: {row_number}")
    expected_display = str(expected_text or "").strip()[:10].replace("-", "/")
    if not expected_display:
        raise ValueError("expected_text is empty for skill inventory updated_at formatter verification")
    sheet_id = _resolve_skill_inventory_sheet_id()
    range_str = f"{sheet_id}!E{row_number}:E{row_number}"
    _run_lark_sheets_json(
        [
            "sheets",
            "+set-style",
            "--url",
            DEFAULT_SKILL_INVENTORY_URL,
            "--range",
            range_str,
            "--style",
            json.dumps({"formatter": DEFAULT_SKILL_INVENTORY_UPDATED_AT_FORMATTER}, ensure_ascii=False),
        ],
        "set updated_at formatter on skill inventory row",
    )
    time.sleep(2)
    readback = _run_lark_sheets_json(
        [
            "sheets",
            "+read",
            "--url",
            DEFAULT_SKILL_INVENTORY_URL,
            "--sheet-id",
            sheet_id,
            "--range",
            f"E{row_number}:E{row_number}",
            "--value-render-option",
            "FormattedValue",
        ],
        "verify updated_at formatter on skill inventory row",
    )
    values = readback.get("data", {}).get("valueRange", {}).get("values", [])
    actual = ""
    if values and values[0]:
        actual = str(values[0][0] or "").strip()
    if actual != expected_display:
        raise RuntimeError(
            "Skill inventory updated_at formatter verification failed: "
            f"expected={expected_display!r}, actual={actual!r}, row={row_number}"
        )


def _parse_block_id_from_attach_output(attach_output: str) -> str:
    """Extract the freshly created file block id from `docs +media-insert` output."""

    for match in re.finditer(r'"block_id"\s*:\s*"([^"]+)"', attach_output or ""):
        return match.group(1)
    return ""


def move_block_to_doc_begin(doc_url: str, block_id: str) -> None:
    """Relocate a top-level block to BLOCK_BEGIN (index 0, i.e. right below the title).

    `lark-cli docs +media-insert` appends to the document end by default, which
    silently breaks the pipeline contract "ZIP 文件块必须挂在标题正下方".
    Anchoring `block_move_after` on the document root token moves the block to
    index 0. Failure raises: no silent drift allowed.
    """

    document_id = parse_doc_token(doc_url)
    run_subprocess(
        [
            "lark-cli",
            "docs",
            "+update",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--command",
            "block_move_after",
            "--block-id",
            document_id,
            "--src-block-ids",
            block_id,
        ],
        "move zip file block to doc begin",
    )


def assert_zip_block_at_doc_begin(doc_url: str, block_id: str) -> None:
    """Runtime gate: the first body block MUST be the newly attached file block."""

    output = run_subprocess(
        [
            "lark-cli",
            "docs",
            "+fetch",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--doc-format",
            "xml",
            "--detail",
            "with-ids",
            "-q",
            ".data.document.content",
        ],
        "verify zip file block position",
    )
    body = re.sub(r"^\s*<title\b[^>]*>.*?</title>", "", output.strip(), count=1, flags=re.S)
    first_block = re.search(r'<(\w+)\s+id="([^"]+)"', body.strip())
    if not first_block or first_block.group(2) != block_id:
        actual = first_block.group(2) if first_block else "<none>"
        raise RuntimeError(
            "ZIP file block position verification FAILED: expected the first body block "
            f"to be {block_id} (BLOCK_BEGIN, right below the title), got {actual}. "
            "文件块回挂位置不符合契约，拒绝宣称发布成功。"
        )


ZIP_FIGURE_RE = re.compile(
    r'<figure\s+id="(?P<block_id>[^"]+)"[^>]*>\s*<source\s+id="[^"]*"\s+name="(?P<file_name>[^"]*)"'
    r'[^>]*?token="(?P<file_token>[^"]+)"',
    re.S,
)


def list_doc_zip_file_blocks(doc_url: str) -> list[Dict[str, str]]:
    """Enumerate every ZIP file block (figure/source) in the doc via `docs +fetch`.

    The former `docx.v1.document_block.list` internal proxy path is no longer
    supported (`unsupported lark method_name`), and it degraded *silently* to an
    empty list — which is exactly how 8 stale ZIP blocks could pile up unnoticed.
    Parsing the authoritative `--doc-format xml --detail with-ids` payload keeps
    the enumeration on the same channel we already use for the position gate.
    """

    output = run_subprocess(
        [
            "lark-cli",
            "docs",
            "+fetch",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--doc-format",
            "xml",
            "--detail",
            "with-ids",
            "-q",
            ".data.document.content",
        ],
        "enumerate doc zip file blocks",
    )
    blocks: list[Dict[str, str]] = []
    for match in ZIP_FIGURE_RE.finditer(output):
        file_name = match.group("file_name") or ""
        if not file_name.lower().endswith(".zip"):
            continue
        blocks.append(
            {
                "block_id": match.group("block_id"),
                "file_name": file_name,
                "file_token": match.group("file_token"),
            }
        )
    return blocks


def is_own_skill_zip(file_name: str, skill_name: str) -> bool:
    """Does this ZIP file name belong to the skill currently being published?

    V5.22 (tightened): the remainder after stripping the `<skill_name>` prefix
    must be one of:
      * empty                      -> `skill-x.zip`
      * a dotted numeric version   -> `skill-x_5.21.zip`, `skill-x-v1.2.zip`
      * a drive dedup suffix       -> `skill-x (1).zip`
      * the explicit `_latest` alias

    Any suffix carrying alphabetic semantics (`-v4`, `_v4`, `-beta`, `_old`)
    is NOT treated as our own stale block — it may belong to an independent
    sibling skill sharing this doc, so it is reported as a foreign block
    instead of being silently deleted (cross-skill deletion accident guard).
    """

    name = (file_name or "").strip()
    if not name.lower().endswith(".zip") or not skill_name:
        return False
    stem = name[: -len(".zip")]
    if not stem.startswith(skill_name):
        return False
    suffix = stem[len(skill_name):].strip()
    if suffix == "" or suffix.lower() in {"_latest", "-latest"}:
        return True
    if re.fullmatch(r"\(\d+\)", suffix):
        return True
    # dotted numeric version only: `v4` / `4` (no dot) is ambiguous with a
    # sibling skill name suffix such as `<parent>-v4`, so it is rejected.
    return bool(re.fullmatch(r"[-_]?[vV]?[0-9]+(?:\.[0-9]+)+", suffix))


def delete_doc_blocks(doc_url: str, block_ids: list[str]) -> str:
    """Physically delete top-level blocks by id (`block_delete` supports batch)."""

    if not block_ids:
        return ""
    return run_subprocess(
        [
            "lark-cli",
            "docs",
            "+update",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--command",
            "block_delete",
            "--block-id",
            ",".join(block_ids),
        ],
        "delete stale zip file blocks",
    )


def prune_stale_zip_blocks(
    doc_url: str,
    skill_name: str,
    new_block_id: str,
) -> Dict[str, Any]:
    """Idempotent replacement: keep exactly ONE ZIP block for this skill.

    Ordering is deliberate — the caller must have already inserted, relocated and
    asserted the NEW block before pruning, so a failed insert can never leave the
    doc without any package. Foreign ZIP blocks (other skills) are reported only,
    never auto-deleted: deleting someone else's asset is destructive and needs a
    human call.

    V5.21 — every failure path here RAISES GuardrailViolation. There is no
    silent-WARNING degradation: enumerate failure, delete failure, readback
    failure and a uniqueness count != 1 are all hard circuit breaks. The previous
    report-only behaviour is what allowed ghost ZIP blocks to accumulate
    unnoticed (info-miner reached 14 stacked packages).
    """

    report: Dict[str, Any] = {
        "enumerated": [],
        "deleted": [],
        "foreign": [],
        "degraded": "",
        "unique_ok": None,
        "residual": [],
    }
    # V5.21: NO silent-WARNING degradation. Every failure below is a hard circuit
    # break. Rationale: this function runs AFTER the new block is inserted and
    # asserted, so raising can never leave the doc without a package — while a
    # soft WARNING is exactly what let 22 ghost ZIP blocks pile up across 8 docs.
    try:
        blocks = list_doc_zip_file_blocks(doc_url)
    except Exception as exc:  # noqa: BLE001
        report["degraded"] = f"enumerate failed: {exc}"
        raise GuardrailViolation(
            "ZIP 文件块 upsert 失败：无法枚举文档现有 ZIP 文件块，因此无法保证唯一性。"
            "严禁降级为「只插入不删除」（该降级正是幽灵安装包堆积的根因）。"
            f"原因：{exc}"
        ) from exc

    report["enumerated"] = blocks
    own_stale = [
        b for b in blocks if is_own_skill_zip(b["file_name"], skill_name) and b["block_id"] != new_block_id
    ]
    report["foreign"] = [b for b in blocks if not is_own_skill_zip(b["file_name"], skill_name)]

    if report["foreign"]:
        print("⚠️ 检测到非本技能的 ZIP 文件块（异物块，仅报告不自动删除，需人工确认）：")
        for b in report["foreign"]:
            print(f"   - block_id={b['block_id']} file={b['file_name']}")

    if own_stale:
        print(f"🧹 Deleting {len(own_stale)} stale ZIP file block(s) of {skill_name}...")
        try:
            delete_doc_blocks(doc_url, [b["block_id"] for b in own_stale])
            report["deleted"] = own_stale
        except Exception as exc:  # noqa: BLE001
            report["degraded"] = f"delete failed: {exc}"
            raise GuardrailViolation(
                f"ZIP 文件块 upsert 失败：{len(own_stale)} 个旧 ZIP 文件块删除失败，"
                f"文档将残留幽灵安装包。原因：{exc}"
            ) from exc
    else:
        print("ℹ️ No stale ZIP file block of this skill found; nothing to prune.")

    # RAW read-after-write uniqueness assertion — hard gate, never report-only.
    time.sleep(2)
    try:
        after = list_doc_zip_file_blocks(doc_url)
    except Exception as exc:  # noqa: BLE001
        report["degraded"] = f"{report['degraded']}; readback failed: {exc}".strip("; ")
        raise GuardrailViolation(
            "ZIP 文件块 upsert 失败：删除后 RAW 回读失败，无法断言唯一性。"
            f"原因：{exc}"
        ) from exc

    mine = [b for b in after if is_own_skill_zip(b["file_name"], skill_name)]
    report["residual"] = mine
    report["unique_ok"] = len(mine) == 1 and mine[0]["block_id"] == new_block_id
    if report["unique_ok"]:
        print(f"✅ ZIP block uniqueness verified: exactly 1 block ({new_block_id}) for {skill_name}.")
    else:
        residual_desc = "; ".join(
            f"block_id={b['block_id']} file={b['file_name']} token={b['file_token']}" for b in mine
        ) or "(none)"
        raise GuardrailViolation(
            f"ZIP 文件块唯一性断言失败：期望 {skill_name} 在文档中恰好 1 个 ZIP 文件块"
            f"（block_id={new_block_id}），实际 {len(mine)} 个。残留清单：{residual_desc}"
        )
    return report


def attach_zip_to_doc_via_mcp(doc_url: str, zip_path: Path, skill_name: str = "") -> str:
    result = subprocess.run(
        [
            "lark-cli",
            "docs",
            "+media-insert",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--file",
            zip_path.name,
            "--type",
            "file",
        ],
        cwd=str(zip_path.resolve().parent),
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"insert native file block via lark-cli failed with exit code {result.returncode}\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )

    attach_output = result.stdout.strip()

    # `+media-insert` appends to the document END; relocate to BLOCK_BEGIN and verify.
    block_id = _parse_block_id_from_attach_output(attach_output)
    if not block_id:
        raise RuntimeError(
            "Unable to parse the new file block_id from lark-cli media-insert output; "
            f"cannot guarantee BLOCK_BEGIN placement.\nSTDOUT:\n{attach_output}"
        )

    print(f"🚚 Relocating zip file block {block_id} to BLOCK_BEGIN (below title)...")
    move_block_to_doc_begin(doc_url, block_id)
    time.sleep(2)
    assert_zip_block_at_doc_begin(doc_url, block_id)
    print("✅ Zip file block verified at BLOCK_BEGIN (right below the doc title).")

    # Idempotent replacement instead of blind append: only after the NEW block is
    # verified in place do we prune this skill's stale ZIP blocks.
    effective_skill_name = skill_name or zip_path.stem
    prune_stale_zip_blocks(doc_url, effective_skill_name, block_id)

    return attach_output


def download_doc_markdown(doc_url: str) -> Path:
    paths = mcp_lark_download(doc_url)
    for path in paths:
        if path.suffix.lower() in {".md", ".lark.md"} or path.name.endswith(".lark.md"):
            return path.resolve()
    if paths:
        return paths[0].resolve()
    raise RuntimeError(f"Unable to download markdown for doc: {doc_url}")


# --- 说明文档正文版本标识（V5.20 加固） ---
# 两类版本标识都必须同步：
#   1) 带标签的：`version: 5.19` / `版本号：v5.19` / `- `version`: `5.19``
#   2) 标题内嵌的：`# 【技能说明】xxx · 技能锻造流水线 (Forge Pipeline V5.19)`
# 历史缺陷：旧实现只认第 1 类，且写替换串时错写成 r"\\1"（字面反斜杠+1，不是分组
# 反向引用），再加上 .lark.md 兜底下载没有 BLOCK 标记 → 循环永远走不进替换分支，
# 最终只打印一句 "skip SSOT doc sync" 就静默放行，正文版本长期停在旧版。
DOC_LABELED_VERSION_RE = re.compile(
    r"(?i)((?:version|版本号|版本)\s*`?\s*[:：]\s*\**\s*`?\s*)([vV]?)(\d+(?:\.\d+){1,2})"
)
DOC_TITLE_VERSION_RE = re.compile(r"(?i)(?<![A-Za-z0-9])([vV])(\d+(?:\.\d+){1,2})")
# 文档标题元数据行：`<title>【技能说明】... (Forge Pipeline V5.19)</title>`
# 真机踩坑（V5.20.1）：文档标题与正文 H1 是两个独立对象，只同步 H1 会留下
# 「正文 V5.20 / 标题仍 V5.19」的半同步幽灵，必须一并纳入改写与断言。
DOC_TITLE_TAG_RE = re.compile(r"^\s*<title>.*</title>\s*$")


def _is_heading_line(line: str) -> bool:
    return bool(re.match(r"^\s{0,3}#{1,6}\s", line)) or bool(DOC_TITLE_TAG_RE.match(line))


def collect_doc_version_lines(text: str) -> list[Tuple[str, list[str]]]:
    """Return [(line, [version_literals...])] for every version-bearing line.

    Only two shapes are trusted as「正文版本标识」: headings with an inline
    `Vx.y[.z]` marker, and lines carrying an explicit version label. Changelog
    history lines (e.g. `- 2026-04-27：v5.2.0`) are deliberately NOT matched, so
    历史记录不会被改写。
    """

    found: list[Tuple[str, list[str]]] = []
    for line in text.splitlines():
        versions = [m.group(3) for m in DOC_LABELED_VERSION_RE.finditer(line)]
        if _is_heading_line(line):
            versions += [m.group(2) for m in DOC_TITLE_VERSION_RE.finditer(line)]
        if versions:
            found.append((line, versions))
    return found


def _rewrite_doc_version_line(line: str, new_version: str) -> str:
    updated = DOC_LABELED_VERSION_RE.sub(
        lambda m: f"{m.group(1)}{m.group(2)}{new_version}", line
    )
    if _is_heading_line(updated):
        updated = DOC_TITLE_VERSION_RE.sub(
            lambda m: f"{m.group(1)}{new_version}", updated
        )
    return updated


def assert_doc_body_version_synced(doc_url: str, new_version: str) -> Dict[str, Any]:
    """L3 runtime gate: RAW read-back of the doc body version markers.

    Post-forge, the Feishu skill doc body/title MUST carry the freshly forged
    version. Anything else (stale version, or no marker at all) is a
    「文档版本未同步」defect and must circuit-break — 静默成功是明确禁止的。
    """

    expected = normalize_version_text(new_version)
    md_path = download_doc_markdown(doc_url)
    text = md_path.read_text(encoding="utf-8", errors="ignore")
    found = collect_doc_version_lines(text)

    if not found:
        raise GuardrailViolation(
            "【文档版本未同步】说明文档正文/标题未找到任何版本标识，无法断言版本已同步。"
            f"doc={doc_url}, expected={expected}. "
            "请在文档标题中加入 `(... Vx.y)` 或正文加入 `版本号：x.y` 标识后重跑。"
        )

    stale = [
        (line, v)
        for line, versions in found
        for v in versions
        if normalize_version_text(v) != expected
    ]
    if stale:
        detail = "; ".join(f"{v!r} @ {line.strip()[:80]!r}" for line, v in stale)
        raise GuardrailViolation(
            f"【文档版本未同步】说明文档正文版本回读断言 FAILED：expected={expected}，"
            f"仍存在旧版本标识 -> {detail}。doc={doc_url}"
        )

    print(
        f"✅ 说明文档正文版本回读断言 PASS：{len(found)} 处版本标识均为 {expected}。"
    )
    return {"expected": expected, "checked_lines": len(found)}


def sync_version_to_skill_doc_via_mcp(doc_url: str, new_version: str) -> Dict[str, Any]:
    """Sync point #3: rewrite EVERY version marker inside the Feishu skill doc.

    Covers both the doc title (`... Forge Pipeline V5.19`) and labeled markers
    (`版本号：5.19`), then hands over to `assert_doc_body_version_synced()` for a
    RAW read-back assertion. No marker found => raise (never a silent skip).
    """

    expected = normalize_version_text(new_version)
    md_path = download_doc_markdown(doc_url)
    text = md_path.read_text(encoding="utf-8", errors="ignore")
    found = collect_doc_version_lines(text)

    if not found:
        raise GuardrailViolation(
            "【文档版本未同步】说明文档正文/标题未找到任何版本标识，拒绝静默跳过。"
            f"doc={doc_url}, target={expected}"
        )

    updated_count = 0
    # 顺序契约（V5.20.1 真机踩坑）：文档标题的内层文字常与正文 H1 完全相同，
    # 若先改标题，str_replace 会命中两处并报 degrade_code=1014（ambiguous）。
    # 因此先改正文各行，最后才改 <title> —— 那时旧文案只剩标题一处，唯一可命中。
    ordered = sorted(found, key=lambda item: 1 if DOC_TITLE_TAG_RE.match(item[0]) else 0)
    for line, _versions in ordered:
        updated_line = _rewrite_doc_version_line(line, expected)
        if updated_line == line:
            continue

        pattern, replacement = line, updated_line
        title_m = DOC_TITLE_TAG_RE.match(line)
        if title_m:
            # `<title>` 的标签本身不是文档正文文本，str_replace 只能匹配到内层文字，
            # 整行下发会命中 degrade_code=1013（pattern not found）。故剥掉标签再替换。
            pattern = re.sub(r"^\s*<title>|</title>\s*$", "", line)
            replacement = re.sub(r"^\s*<title>|</title>\s*$", "", updated_line)

        print(f"📝 Syncing doc version marker: {pattern.strip()[:80]!r} -> {expected}")
        run_subprocess(
            [
                "lark-cli",
                "docs",
                "+update",
                "--api-version",
                "v2",
                "--as",
                "user",
                "--doc",
                doc_url,
                "--command",
                "str_replace",
                "--doc-format",
                "markdown",
                "--pattern",
                pattern,
                "--content",
                replacement,
            ],
            "sync doc version marker",
        )
        updated_count += 1

    if updated_count == 0:
        print(f"ℹ️ 说明文档版本标识已是 {expected}，无需改写；继续执行回读断言。")

    time.sleep(2)
    assertion = assert_doc_body_version_synced(doc_url, expected)
    return {"updated_lines": updated_count, **assertion}


def verify_file_block_attached(doc_url: str, zip_name: str) -> bool:
    markdown_path = download_doc_markdown(doc_url)
    content = markdown_path.read_text(encoding="utf-8")
    zip_stem = Path(zip_name).stem
    return f"file_{zip_stem}" in content or zip_name in content or zip_stem in content


def move_doc_to_wiki_via_mcp(doc_url: str, wiki_node_token: str) -> Dict[str, str]:
    doc_token = doc_url.rstrip("/").split("/")[-1].split("?")[0]
    node_info_output = run_subprocess(
        [
            "lark-cli",
            "wiki",
            "+node-get",
            "--as",
            "user",
            "--node-token",
            wiki_node_token,
            "--format",
            "json",
        ],
        "resolve target wiki node",
    )
    try:
        node_info = json.loads(node_info_output[node_info_output.find("{"):])
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Unable to parse target wiki node info: {node_info_output}") from exc
    data = node_info.get("data", {}) if isinstance(node_info, dict) else {}
    node = data.get("node", data) if isinstance(data, dict) else {}
    space_id = str(node.get("space_id") or data.get("space_id") or "").strip()
    if not space_id:
        raise RuntimeError(f"Unable to resolve target wiki space id: {node_info_output}")

    output = run_subprocess(
        [
            "lark-cli",
            "wiki",
            "+move",
            "--as",
            "user",
            "--obj-type",
            "docx",
            "--obj-token",
            doc_token,
            "--target-space-id",
            space_id,
            "--target-parent-token",
            wiki_node_token,
            "--format",
            "json",
        ],
        "lark move doc to wiki",
    )

    urls = re.findall(r'https?://[^\s"\'<>]+', output)
    moved_doc_url = next((url for url in urls if "/docx/" in url or "/docs/" in url), doc_url)
    moved_wiki_url = next((url for url in urls if "/wiki/" in url), "")
    if not moved_wiki_url:
        try:
            payload = json.loads(output[output.find("{"):])
            payload_text = json.dumps(payload, ensure_ascii=False)
            urls = re.findall(r'https?://[^\s"\'<>]+', payload_text)
            moved_wiki_url = next((url for url in urls if "/wiki/" in url), "")
            moved_doc_url = next((url for url in urls if "/docx/" in url or "/docs/" in url), moved_doc_url)
        except json.JSONDecodeError:
            pass

    return {
        "doc_link": moved_doc_url,
        "wiki_url": moved_wiki_url or f"https://bytedance.larkoffice.com/wiki/{wiki_node_token}",
        "wiki_node_token": wiki_node_token,
        "raw_output": output,
    }


def parse_doc_token(doc_url: str) -> str:
    path = urlparse(doc_url).path or ""
    parts = [part for part in path.split("/") if part]
    if len(parts) >= 2 and parts[0] in {"docx", "docs", "doc"}:
        return parts[1]
    raise ValueError(f"Unsupported doc url: {doc_url}")


def call_lark_proxy(method_name: str, *, path: Optional[Dict[str, Any]] = None, query: Optional[Dict[str, Any]] = None, body: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    api_host = (os.environ.get("IRIS_RUNTIME_AIME_API_HOST") or "aime.bytedance.net").strip()
    if api_host.startswith("http://") or api_host.startswith("https://"):
        proxy_url = f"{api_host.rstrip('/')}/api/agents/v2/internal/proxy/lark_api"
    else:
        proxy_url = f"https://{api_host}/api/agents/v2/internal/proxy/lark_api"

    response = requests.post(
        proxy_url,
        headers={"Authorization": f"Byte-Cloud-JWT {get_raw_cloud_jwt()}"},
        json={
            "method_name": method_name,
            "path": path or {},
            "query": query or {},
            "body": body or {},
            "use_user_token": True,
        },
        timeout=60,
    )
    data = response.json()
    if response.status_code != 200:
        raise RuntimeError(f"lark proxy failed: http={response.status_code}, resp={data}")
    return data


def list_doc_file_blocks(doc_url: str) -> list[Dict[str, str]]:
    document_id = parse_doc_token(doc_url)
    response = call_lark_proxy(
        "docx.v1.document_block.list",
        path={"document_id": document_id},
        query={"page_size": "500"},
    )

    if isinstance(response, dict) and response.get("code") not in (0, None):
        # Never degrade silently: an unsupported/failed proxy call used to return an
        # empty list, which made stale-block pruning a no-op without any signal.
        raise RuntimeError(f"docx.v1.document_block.list failed: {response}")

    items = response.get("data", {}).get("items") or response.get("items") or []
    file_blocks: list[Dict[str, str]] = []
    for item in items:
        file_info = item.get("file") or {}
        file_token = file_info.get("file_token") or file_info.get("token") or ""
        file_name = file_info.get("name") or file_info.get("file_name") or ""
        if file_token:
            file_blocks.append(
                {
                    "block_id": item.get("block_id", ""),
                    "file_token": file_token,
                    "file_name": file_name,
                }
            )
    return file_blocks


def parse_file_token_from_attach_output(attach_output: str) -> str:
    if not attach_output:
        return ""

    json_like_matches = re.findall(r'\{[^{}]*"file_token"\s*:\s*"([^"]+)"[^{}]*\}', attach_output)
    if json_like_matches:
        return json_like_matches[0]

    dict_like_matches = re.findall(r"['\"]file_token['\"]\s*[:=]\s*['\"]([^'\"\s]+)['\"]", attach_output)
    if dict_like_matches:
        return dict_like_matches[0]

    # Some MCP / download outputs only expose a raw token-like field or a
    # canonical /file/<token> URL instead of a JSON object.
    url_match = re.search(r"/file/([A-Za-z0-9_-]+)", attach_output)
    if url_match:
        return url_match.group(1)

    loose_token_match = re.search(
        r"(?:^|[\s'\"=:,])(file_[A-Za-z0-9_-]{10,}|boxcn[A-Za-z0-9_-]{10,}|[A-Za-z0-9_-]{20,})(?=$|[\s'\",}])",
        attach_output,
    )
    if loose_token_match:
        return loose_token_match.group(1)

    return ""


def resolve_attached_drive_file_token(
    attach_output: str,
    before_blocks: list[Dict[str, str]],
    doc_url: str,
    file_name: str,
    before_blocks_error: str = "",
) -> str:
    """Resolve the Drive file_token for the newly attached ZIP.

    Resolution order (fail-fast, but with an extra zero-trust fallback):
    1) Parse stdout/stderr from the MCP attach script (preferred, includes raw file_token).
    2) If that fails, call docx.v1.document_block.list via internal proxy and diff
       before/after file blocks to locate the new block.
    3) If the OpenAPI path fails (e.g. missing scope, doc type not supported),
       fall back to inspecting the latest downloaded .lark.md and reusing the
       same parsing heuristics we use for attach_output (searching for /file/<token>
       or JSON blobs containing "file_token").

    Only after all three paths fail do we hard-stop, to avoid producing an
    orphan ZIP without a recorded Drive token / permission grant.
    """

    # 1) Primary: parse attach_output directly (most structured / least fragile).
    parsed_token = parse_file_token_from_attach_output(attach_output)
    if parsed_token:
        return parsed_token

    attach_summary = "attach_output missing parseable file_token"
    list_failure_summary = before_blocks_error or "before attachment snapshot ok"

    # 2) Secondary: try listing file blocks via internal Lark proxy and diffing.
    try:
        after_blocks = list_doc_file_blocks(doc_url)
        return find_new_file_token(before_blocks, after_blocks, file_name)
    except Exception as exc:
        list_failure_summary = f"{list_failure_summary}; after attachment lookup failed: {exc}"

    # 3) Tertiary: zero-trust markdown introspection — download latest doc as
    # .lark.md and reuse the same token parsing heuristics on its content.
    try:
        md_path = download_doc_markdown(doc_url)
        md_text = md_path.read_text(encoding="utf-8", errors="ignore")
        markdown_token = parse_file_token_from_attach_output(md_text)
        if markdown_token:
            return markdown_token
    except Exception as exc:
        list_failure_summary = f"{list_failure_summary}; markdown introspection failed: {exc}"

    # All paths failed: refuse to continue in order to avoid delivering an
    # untracked, potentially permissionless orphan Drive file.
    raise RuntimeError(
        "Unable to resolve attached drive file token; refuse to continue because this would "
        "deliver a potentially permissionless orphan file. "
        f"attach_output path failed: {attach_summary}. "
        f"fallbacks failed: {list_failure_summary}."
    )


def find_new_file_token(before_blocks: list[Dict[str, str]], after_blocks: list[Dict[str, str]], file_name: str) -> str:
    before_tokens = {item.get("file_token", "") for item in before_blocks}
    named_candidates = [
        item for item in after_blocks if item.get("file_name") == file_name and item.get("file_token") not in before_tokens
    ]
    if named_candidates:
        return named_candidates[0]["file_token"]

    new_candidates = [item for item in after_blocks if item.get("file_token") not in before_tokens]
    if new_candidates:
        return new_candidates[0]["file_token"]

    raise RuntimeError(f"Unable to locate the newly inserted File Block token for {file_name}")


def extract_metadata(
    name: str,
    desc: str,
    doc_link: str,
    version: str = "",
    skill_id: Optional[str] = None,
    skill_dir: Optional[Path] = None,
    zip_path: Optional[Path] = None,
    drive_file_token: Optional[str] = None,
    drive_file_url: Optional[str] = None,
    wiki_url: str = "",
    wiki_node_token: str = "",
    doc_version_sync: Optional[Dict[str, Any]] = None,
    doc_zone_sync: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    now = datetime.datetime.now()
    return {
        "skill_id": skill_id or f"SKILL-{now.strftime('%m%d%H%M')}",
        "name": name,
        "description": desc,
        "version": version,
        "doc_link": doc_link,
        "wiki_url": wiki_url,
        "wiki_node_token": wiki_node_token,
        "date": now.strftime("%Y-%m-%d"),
        "updated_at": now.strftime("%Y-%m-%d %H:%M"),
        "skill_dir": str(skill_dir.resolve()) if skill_dir else "",
        "zip_path": str(zip_path.resolve()) if zip_path else "",
        "zip_name": zip_path.name if zip_path else "",
        "drive_file_token": drive_file_token or "",
        "drive_file_url": drive_file_url or "",
        "doc_version_synced": bool(doc_version_sync),
        "doc_version_sync": doc_version_sync or {},
        # 三分区（Zone）同步结果（V5.23）：含降级标记与回读断言证据
        "doc_zone_synced": bool(doc_zone_sync),
        "doc_zone_sync": doc_zone_sync or {},
        "doc_zone_degraded": (doc_zone_sync or {}).get("degraded", ""),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--name", required=True, help="目标技能名称")
    parser.add_argument("--desc", required=True, help="目标技能功能描述")
    parser.add_argument("--path", required=True, help="目标技能说明飞书文档 URL")
    parser.add_argument("--skill-dir", help="目标技能目录，例如 user_skills/xxx")
    parser.add_argument("--id", required=False, help="技能编号或包 ID")
    parser.add_argument("--user-email", default=DEFAULT_EMAIL, help="默认授予 Full Access 的邮箱")
    parser.add_argument("--output-zip", help="可选，自定义 zip 输出路径")
    parser.add_argument(
        "--wiki-node-token",
        default=DEFAULT_WIKI_NODE_TOKEN,
        help="技能说明文档默认迁入的 Wiki 节点 token（默认 Aime 技能库根节点）。",
    )
    parser.add_argument(
        "--skip-wiki-mount",
        action="store_true",
        help="调试用：跳过 Wiki Mount Phase（正式发布默认必须执行）。",
    )
    parser.add_argument(
        "--wiki-registry-url",
        default=DEFAULT_WIKI_REGISTRY_URL,
        help="Wiki「技能存量清单」表格所在页面 URL（默认 Aime 技能库首页）。",
    )
    parser.add_argument(
        "--skip-wiki-sync",
        action="store_true",
        help="调试用：跳过 Wiki 技能存量清单表格 Upsert（增强项，失败本身也只降级为 WARNING）。",
    )

    # SSOT version sync bus
    parser.add_argument(
        "--bump",
        choices=["major", "minor", "patch"],
        help="SSOT 版本升迁：major(+1.0) / minor(+0.1)。若未提供且为交互式终端，会提示选择；非交互式将直接报错。",
    )
    parser.add_argument(
        "--new-version",
        help="直接指定新版本号（标准 X.Y 两段式，如 5.2）。优先级高于 --bump。",
    )
    parser.add_argument(
        "--initial-version",
        default=DEFAULT_INITIAL_VERSION,
        help=(
            "首次发布时的起始版本号（默认 1.1）。"
            "当 SKILL.md 当前版本仍处于 0.x 脚手架阶段时，"
            "本流水线会忽略 --bump，直接将版本设为该值。"
        ),
    )
    parser.add_argument(
        "--skip-ssot",
        action="store_true",
        help="调试用：跳过 SSOT 版本同步（不会改动目标技能的 SKILL.md，也不会同步台账/文档版本标识）。",
    )
    parser.add_argument(
        "--ssot-spreadsheet-token",
        default=DEFAULT_SSOT_SPREADSHEET_TOKEN,
        help="飞书技能台账 Spreadsheet Token 或完整 URL（默认使用主台账）。",
    )
    parser.add_argument(
        "--ssot-sheet-name",
        default=DEFAULT_SSOT_SHEET_NAME,
        help="飞书技能台账工作表名称（默认：专属技能清单）。",
    )
    parser.add_argument(
        "--skip-ssot-sheet-sync",
        action="store_true",
        help="调试用：跳过技能台账整行 upsert（omni-asset-archiver）。",
    )
    parser.add_argument(
        "--skip-ssot-doc-sync",
        action="store_true",
        help="调试用：跳过说明文档版本标识同步。",
    )
    parser.add_argument(
        "--emit-new-doc-markdown",
        metavar="OUT_PATH",
        help=(
            "新建文档场景：把三分区骨架（Overwrite -> Preserve 占位 -> Append）"
            "渲染成 Markdown 落到指定路径，供 `lark-cli docs +create` 导入后即刻具备 Zone 锚点。"
        ),
    )
    parser.add_argument(
        "--skip-doc-zones",
        action="store_true",
        help=(
            "调试用：跳过飞书说明文档三分区（Overwrite/Preserve/Append）同步。"
            "正式发布默认必须执行 —— 它同时承担 Preserve Zone 人工沉淀的保护断言。"
        ),
    )

    parser.add_argument(
        "--skip-remote",
        action="store_true",
        help="仅执行本地打包与 metadata 生成，跳过飞书云盘上传、权限赋予、文档挂载与 SSOT 远端同步",
    )

    # --- Cloud Publish (V5.19) ---
    parser.add_argument(
        "--cloud-scope",
        choices=["user", "space"],
        default=DEFAULT_CLOUD_SCOPE,
        help=(
            "云端发布可见性：user（默认，个人可见）/ space（空间可见）。"
            "只有用户显式要求「空间可见/全员可用」时才允许传 space。"
        ),
    )
    parser.add_argument(
        "--enable-by-default",
        action="store_true",
        help=(
            "云端发布时对空间全员默认启用（仅 --cloud-scope space 有效，默认不传）。"
            "属于扩大可见范围的动作，必须由用户显式要求。"
        ),
    )
    args = parser.parse_args()

    workspace_root = get_workspace_root()
    skill_dir = (workspace_root / args.skill_dir).resolve() if args.skill_dir else None

    # 新建文档场景：先产出带三分区锚点的骨架，再交给 `docs +create` 导入。
    # 这样文档「出生即合规」，后续迭代不必再走老文档降级补锚点的路径。
    if args.emit_new_doc_markdown:
        if not skill_dir:
            raise RuntimeError("--emit-new-doc-markdown requires --skill-dir")
        skeleton_version = args.new_version or read_skill_version_from_skill_md(skill_dir)
        skeleton = build_new_doc_markdown(
            skill_dir,
            normalize_version_text(skeleton_version),
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
        out_path = Path(args.emit_new_doc_markdown).resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(skeleton, encoding="utf-8")
        print(f"✅ 三分区新建文档骨架已生成：{out_path}")
        print(f"   Zone 锚点：『{PRESERVE_ANCHOR_TITLE}』/『{APPEND_ANCHOR_TITLE}』")
        return 0
    zip_path: Optional[Path] = None
    drive_file_token = ""
    drive_file_url = ""
    final_doc_link = args.path
    wiki_url = ""
    wiki_node_token = ""
    doc_version_sync: Optional[Dict[str, Any]] = None
    doc_zone_sync: Optional[Dict[str, Any]] = None

    # SSOT version (X.Y)
    ssot_version = ""
    if skill_dir:
        if args.skip_ssot:
            ssot_version = read_skill_version_from_skill_md(skill_dir)
            print(f"🚌 SSOT disabled (--skip-ssot). Keep current version: {ssot_version}")
        else:
            current_version = read_skill_version_from_skill_md(skill_dir)
            new_version = ""

            if args.new_version:
                new_version = normalize_version_text(args.new_version)
            elif is_initial_version(current_version):
                # First publish: ignore --bump, jump straight to initial version (default 1.1).
                new_version = normalize_version_text(args.initial_version)
                print(
                    f"🚌 SSOT initial publish detected (current={current_version}). "
                    f"Forcing initial version: {new_version}"
                )
            else:
                bump_type = args.bump
                if not bump_type:
                    if sys.stdin.isatty():
                        print("\n=== SSOT Version Bump ===")
                        print(f"current: {current_version}")
                        print("choose bump type:")
                        print("  1) minor (+0.1)")
                        print("  2) major (+1.0)")
                        print("  3) patch (+0.0.1)")
                        choice = (input("Select (default=1): ") or "1").strip()
                        bump_type = {"2": "major", "3": "patch"}.get(choice, "minor")
                    else:
                        raise RuntimeError(
                            "SSOT version sync requires --bump {major|minor|patch} or --new-version X.Y[.Z] in non-interactive mode."
                        )
                new_version = bump_version(current_version, bump_type)

            ssot_version = new_version
            if ssot_version != current_version:
                print(f"🚌 SSOT bump: {current_version} -> {ssot_version}")
                write_skill_md_version(skill_dir, ssot_version)
            else:
                print(f"🚌 SSOT version unchanged: {ssot_version}")

        # --- CDA Guardrails Checkpoint (fail fast) ---
        print("🚧 Running CDA-Guardrails-Selfcheck before packaging...")
        print(run_cda_guardrails_selfcheck(skill_dir))

        output_zip = Path(args.output_zip).resolve() if args.output_zip else skill_dir.parent / f"{skill_dir.name}.zip"
        print(f"🚀 Packaging skill directory: {skill_dir}")
        zip_path = create_skill_zip(skill_dir, output_zip)
        print(f"✅ Skill package created: {zip_path}")

        if not args.skip_remote:
            # Per system guardrails: authenticate bytedcli before any Feishu MCP operations.
            ensure_bytedcli_auth()
            bearer_token = ensure_bearer_token()

            print("🚧 Validating skill doc template markers...")
            call_with_retry(
                "validate skill doc template markers",
                lambda: validate_doc_template_markers(args.path),
            )

            before_blocks: list[Dict[str, str]] = []
            before_blocks_error = ""
            try:
                before_blocks = list_doc_file_blocks(args.path)
            except Exception as exc:
                before_blocks_error = f"before attachment snapshot failed: {exc}"
                print(f"⚠️ Unable to list doc blocks before attachment: {exc}")

            print("🚀 Inserting native File Block into skill doc via Lark MCP...")
            attach_output = call_with_retry(
                "insert file block into skill doc",
                lambda: attach_zip_to_doc_via_mcp(args.path, zip_path, skill_dir.name),
            )
            verified = call_with_retry(
                "verify attached file block via lark MCP download",
                lambda: verify_file_block_attached(args.path, zip_path.name),
            )
            if not verified:
                raise RuntimeError(f"File Block verification failed for {zip_path.name}")

            print("✅ Skill doc attachment sync finished and verified.")
            if attach_output:
                print(attach_output)

            if ssot_version and (not args.skip_ssot) and (not args.skip_ssot_doc_sync):
                print("🚌 Syncing SSOT version marker to skill doc...")
                doc_version_sync = call_with_retry(
                    "sync ssot version marker",
                    lambda: sync_version_to_skill_doc_via_mcp(args.path, ssot_version),
                )

            # ---------- 三分区（Zone）同步（V5.23） ----------
            # 顺序契约：必须在版本标识同步之后、Wiki Mount 之前。
            #   * Overwrite Zone -> 从 SKILL.md 重新渲染覆盖（版本框 / 触发词 / 接口契约）
            #   * Preserve Zone  -> 一律不动（人工沉淀的使用案例与踩坑记录）
            #   * Append Zone    -> 末尾追加本版本 Changelog 条目
            # 写后 RAW 回读断言：两个锚点各出现恰好 1 次 + Preserve 正文存在性断言。
            if ssot_version and (not args.skip_doc_zones):
                print("🧭 三分区策略：同步飞书说明文档 Zone...")
                changelog_entry = build_changelog_entry_from_skill_md(skill_dir, ssot_version)
                doc_zone_sync = call_with_retry(
                    "sync skill doc zones (overwrite/preserve/append)",
                    lambda: sync_doc_zones(
                        args.path,
                        skill_dir,
                        ssot_version,
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                        changelog_entry,
                    ),
                )
                if doc_zone_sync.get("degraded"):
                    print(
                        "⚠️ [ZONE-DEGRADED] 本次三分区同步发生降级（老文档缺锚点或边界不可信）："
                        f"{doc_zone_sync['degraded']}"
                    )
            elif args.skip_doc_zones:
                print("⚠️ 三分区同步被 --skip-doc-zones 跳过（调试用）。")

            if args.skip_wiki_mount:
                print("⚠️ Wiki Mount Phase skipped by --skip-wiki-mount (debug only).")
            else:
                print("🚀 Wiki Mount Phase: moving skill doc into target wiki node...")
                wiki_mount_result = call_with_retry(
                    "move skill doc to wiki",
                    lambda: move_doc_to_wiki_via_mcp(args.path, args.wiki_node_token),
                )
                final_doc_link = wiki_mount_result.get("doc_link") or args.path
                wiki_url = wiki_mount_result.get("wiki_url") or DEFAULT_WIKI_URL
                wiki_node_token = wiki_mount_result.get("wiki_node_token") or args.wiki_node_token
                print("✅ Wiki Mount Phase finished.")
                if wiki_mount_result.get("raw_output"):
                    print(wiki_mount_result["raw_output"])

            # L3 final gate: 迁入 Wiki 后再回读一次正文版本，确保搬家过程没把旧版本带回来。
            if ssot_version and (not args.skip_ssot) and (not args.skip_ssot_doc_sync):
                print("🚧 Final assert: 说明文档正文版本回读（post wiki mount）...")
                doc_version_sync = call_with_retry(
                    "assert doc body version synced",
                    lambda: assert_doc_body_version_synced(final_doc_link, ssot_version),
                )

            drive_file_token = resolve_attached_drive_file_token(
                attach_output=attach_output,
                before_blocks=before_blocks,
                doc_url=final_doc_link,
                file_name=zip_path.name,
                before_blocks_error=before_blocks_error,
            )
            drive_file_url = build_drive_file_url(final_doc_link, drive_file_token)
            print(f"🚀 Granting drive asset access (full_access) to {args.user_email} via lark-cli drive +member-add...")
            repair_output = call_with_retry(
                "repair drive asset access via MCP personal flow",
                lambda: ensure_drive_asset_access_via_mcp(
                    drive_file_url,
                    email=args.user_email,
                ),
            )
            print("✅ Drive asset access granted & verified via +member-list RAW readback.")
            if repair_output:
                print(repair_output)

    metadata = extract_metadata(
        args.name,
        args.desc,
        final_doc_link,
        ssot_version,
        args.id,
        skill_dir,
        zip_path,
        drive_file_token,
        drive_file_url,
        wiki_url,
        wiki_node_token,
        doc_version_sync,
        doc_zone_sync,
    )

    if (
        skill_dir
        and ssot_version
        and (not args.skip_ssot)
        and (not args.skip_remote)
        and (not args.skip_ssot_sheet_sync)
    ):
        sync_result = call_with_retry(
            "sync skill inventory via omni-asset-archiver",
            lambda: sync_skill_inventory_via_omni(metadata),
        )
        row_number = int(sync_result.get("row_number") or 0)
        call_with_retry(
            "sync skill inventory updated_at formatter",
            lambda: ensure_skill_inventory_updated_at_formatter(row_number, metadata.get("updated_at", "")),
        )

    # ---------- Wiki 技能存量清单 Upsert（V5.24） ----------
    # 顺序契约：Sheet 台账写入完成之后、Celebrate（cyber-inspiration-generator）之前。
    # 失败降级为 WARNING，不阻断 forge 主流程（增强项）。
    if skill_dir and (not args.skip_remote) and (not args.skip_wiki_sync):
        print("🚌 Syncing Wiki skill registry table (技能存量清单)...")
        wiki_sync_report = run_wiki_skill_list_sync(metadata, args.wiki_registry_url)
        metadata["wiki_registry_sync_status"] = wiki_sync_report.get("status", "")
        metadata["wiki_registry_sync"] = wiki_sync_report
    elif args.skip_wiki_sync:
        print("⚠️ Wiki 技能存量清单同步被 --skip-wiki-sync 跳过（调试用）。")
        metadata["wiki_registry_sync_status"] = "skipped"

    # Forge receipt lands INSIDE the target skill dir (never Path.cwd()).
    # Historical bug: `Path.cwd() / "metadata.json"` scattered ghost receipts
    # into whatever directory the pipeline happened to run from, and the
    # "metadata" name made them look like authoritative skill metadata
    # (they are not — Skill ID must come from the Feishu inventory sheet).
    receipt_dir = skill_dir if skill_dir else Path.cwd()
    receipt_path = receipt_dir / ".forge_receipt.json"
    receipt_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=4), encoding="utf-8")

    print("🚀 Metadata packaged for omni-asset-archiver:")
    print(json.dumps(metadata, indent=4, ensure_ascii=False))
    print(f"\n✅ Forge receipt written to {receipt_path.resolve()}")

    if skill_dir:
        print("🚀 Running post-forge git push hook...")
        print(run_post_forge_git_push(workspace_root, args.name, ssot_version or "latest"))

        # ---------- 第四步：Cloud Publish（云端发布） ----------
        # 顺序契约：必须在 Git Push 成功之后执行。
        # run_post_forge_git_push 内部失败会 raise（hook 非 0 退出即熔断），
        # 因此能走到这里就说明远端 SHA 回读断言已 PASS。
        print("🚀 Running cloud publish (aime skill upload + cloud readback assert)...")
        cloud_result = run_cloud_publish(
            skill_dir,
            version=ssot_version or "",
            cloud_scope=args.cloud_scope,
            enable_by_default=args.enable_by_default,
            workspace_root=workspace_root,
        )
        print(cloud_result.get("log", ""))

        # 云端发布结果写入 metadata（成功/失败都要如实落盘，禁止静默）
        metadata["cloud_publish_status"] = cloud_result.get("cloud_publish_status", "")
        metadata["cloud_scope"] = cloud_result.get("cloud_scope", "")
        metadata["cloud_published_at"] = cloud_result.get("cloud_published_at", "")
        if cloud_result.get("cloud_skill_id"):
            metadata["cloud_skill_id"] = cloud_result["cloud_skill_id"]
        if cloud_result.get("dlq_path"):
            metadata["cloud_publish_dlq"] = cloud_result["dlq_path"]
        if cloud_result.get("error"):
            metadata["cloud_publish_error"] = cloud_result["error"]
        receipt_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=4), encoding="utf-8")
        print(f"✅ .forge_receipt.json updated with cloud_publish_status={metadata['cloud_publish_status']}")

        if str(metadata["cloud_publish_status"]).startswith("FAILED"):
            print(
                "❗ 云端发布未完成：技能仍是本地草稿，需要手动上传到云端才能全局生效。"
                "详情见上方 ERROR 与死信队列。"
            )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
