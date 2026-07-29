#!/usr/bin/env python3
"""根据候选时段与参会人忙闲信息生成忙闲矩阵。

升级点：
- 支持多天、多时段的候选输入，只要 `top_slots` 中的时间包含不同日期即可；
- 默认输出 `.xlsx` 文件，并在覆盖率列使用颜色标注：
  - 覆盖率 >= 80%：浅绿色
  - 50% <= 覆盖率 < 80%：浅黄色
  - 覆盖率 < 50%：浅红色
- 若运行环境中无法导入 openpyxl，则自动降级为 `.csv` 输出，避免中断主流程。

`top_slots` 约定结构：
    [(start_dt, end_dt, conflict_count, conflicts), ...]
其中 `conflict_count`、`conflicts` 仅作为备查信息，不直接影响矩阵构造。
"""

import os
from typing import Dict, Iterable, List, Optional, Tuple

try:
    from suggest_timeslots import build_timeslot_display_row
except ModuleNotFoundError:  # 兼容从 skill 根目录以 `scripts.xxx` 方式导入
    from scripts.suggest_timeslots import build_timeslot_display_row

try:  # 尝试加载 openpyxl，用于生成带颜色标注的 xlsx
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    OPENPYXL_AVAILABLE = True
except Exception:  # noqa: BLE001 - 对所有导入异常统一降级处理
    OPENPYXL_AVAILABLE = False


def build_availability_xlsx(
    top_slots: Iterable[Tuple[object, object, int, List[str]]],
    attendees: List[str],
    busy_map: Dict[str, List[Tuple[object, object]]],
    output_path: str = "availability.xlsx",
    primary_tz_label: str = "BJT",
    primary_tz_name: Optional[str] = "Asia/Shanghai",
    secondary_tz_label: str = "",
    secondary_tz_name: Optional[str] = None,
) -> str:
    """生成忙闲矩阵文件，并返回本地文件路径。

    参数：
        top_slots: [(start_dt, end_dt, conflict_count, conflicts), ...]
        attendees: 参会人列表（名称或邮箱）
        busy_map: 每个参会人的忙碌区间，用于重新计算覆盖率
        output_path: 输出文件路径，建议以 .xlsx 结尾

    返回：
        实际写入的文件路径（可能为 .xlsx 或降级后的 .csv）。
    """

    # 统一构造用于写入的行数据
    rows = []
    time_headers = [f"时间段（{primary_tz_label}）"]
    include_secondary_timezone = bool(secondary_tz_label and secondary_tz_name)
    if include_secondary_timezone:
        time_headers.append(f"时间段（{secondary_tz_label}）")
    header = ["日期"] + time_headers + ["覆盖率"] + attendees

    for start_dt, end_dt, _conflict_count, _conflicts in top_slots:
        display_row = build_timeslot_display_row(
            start_dt,
            end_dt,
            primary_tz_name=primary_tz_name,
            secondary_tz_name=secondary_tz_name if include_secondary_timezone else None,
        )

        row_states: List[str] = []
        free_count = 0
        for person in attendees:
            is_busy = False
            for b_start, b_end in busy_map.get(person, []):
                if not (end_dt <= b_start or start_dt >= b_end):
                    is_busy = True
                    break
            state = "忙碌" if is_busy else "空闲"
            if not is_busy:
                free_count += 1
            row_states.append(state)

        coverage = free_count / float(len(attendees) or 1)
        rows.append(
            {
                "date": display_row["date"],
                "primary_slot": display_row["primary_slot"],
                "secondary_slot": display_row["secondary_slot"],
                "coverage": coverage,
                "states": row_states,
            }
        )

    if OPENPYXL_AVAILABLE:
        # 首选 xlsx 输出，并在覆盖率列使用颜色标注
        wb = Workbook()
        ws = wb.active
        ws.title = "Availability"
        ws.append(header)

        green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")  # 高覆盖率
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")  # 中覆盖率
        red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")  # 低覆盖率

        for r in rows:
            coverage_pct = round(r["coverage"] * 100)
            row_values = [r["date"], r["primary_slot"]]
            if include_secondary_timezone:
                row_values.append(r["secondary_slot"])
            row_values.extend([f"{coverage_pct}%"] + r["states"])
            ws.append(row_values)
            row_idx = ws.max_row
            cov_column_idx = 3 + (1 if include_secondary_timezone else 0)
            cov_cell = ws.cell(row=row_idx, column=cov_column_idx)
            if r["coverage"] >= 0.8:
                cov_cell.fill = green_fill
            elif r["coverage"] >= 0.5:
                cov_cell.fill = yellow_fill
            else:
                cov_cell.fill = red_fill

        wb.save(output_path)
        print(f"Generated availability matrix (xlsx): {output_path}")
        return output_path

    # 降级：无 openpyxl 时生成 CSV，避免打断主流程
    csv_path = os.path.splitext(output_path)[0] + ".csv"
    lines: List[str] = [",".join(header)]
    for r in rows:
        coverage_pct = round(r["coverage"] * 100)
        line = [r["date"], r["primary_slot"]]
        if include_secondary_timezone:
            line.append(r["secondary_slot"])
        line.extend([f"{coverage_pct}%"] + r["states"])
        lines.append(",".join(line))

    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"Generated availability matrix (csv fallback): {csv_path}")
    return csv_path


if __name__ == "__main__":
    # 此脚本通常由 AIME 在业务逻辑中调用，命令行仅作提示。
    print(
        "This script builds an availability matrix as an .xlsx (with color-coded coverage when openpyxl is available).",
    )
