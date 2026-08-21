#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Daily_Logs 零信任安全插入（一键脚本）

目标：把"写日报→归档到 Daily_Logs 台账"的流程固化为可执行脚本，避免脏写/错列。

写入模式（V6.3 起）：
- **降序插入法（Descending Insert / 治本）**：在表头（第 1 行）下方物理插入空行（`lark-cli sheets +dim-insert
  --position 2 --count 1`），再把 [编号, 日期, 日报内容] 写入 A2:C2。最新日报永远置顶，天然维持降序。
- **禁止 append-to-tail**：历史 `+append` 路径会把最新日报埋到表尾（P1 事故：埋到第 108 行），已彻底移除。

强约束：
- 必须先通过 MCP / lark-cli 读取表头 Schema（第 1 行）
- 当存在【编号】列时必须自动生成主键（DL-YYYYMMDD / DL-YYYYMMDD-02 ...）
- 【治本封堵】必须写满 3 列：[[编号, 日期, 日报内容]]，任一字段为空即熔断。
- 写后即读（RAW 原子锁）：写入后等待 >=2s 回读，并强制执行三断言（治表）：
  * 断言 A `assert_top_row_is_today()`：B2 == 当日日期（最新行置顶）
  * 断言 B `assert_date_desc()`：B 列（排除表头）严格降序（无乱序）
  * 断言 C `assert_no_empty_cells()`：A2 / B2 / C2 均非空（无错列）
  统一入口 `assert_daily_logs_invariants()`，任一断言失败即 raise 熔断 + 落 DLQ，绝不静默通过。

注意：
- 飞书读写一律走 MCP / lark-cli，严禁裸调 OpenAPI。
- 本脚本不负责 bytedcli 登录；请在 Aime 执行时先挂载 bytedcli-auth 并 include_secrets=true。
"""

from __future__ import annotations

import argparse
import ast
import datetime as dt
import json
import os
import re
import shutil
import sys
import time
import traceback
from pathlib import Path

import openpyxl

DEFAULT_SHEET_URL = "https://bytedance.larkoffice.com/sheets/ECQ0sDwmbhDex9tcUSjlkU7Bgdh"
DEFAULT_SHEET_NAME = "Daily_Logs"
# 写入模式约定（V6.3）：降序插入法。永远在表头下方（第 2 行）物理插入空行后写入 A2:C2。
# 严禁 append-to-tail：那会把最新日报埋在表尾（P1 事故根因）。
TOP_ROW_INDEX = 2
DEFAULT_ROW_INDEX = TOP_ROW_INDEX  # 兼容历史入参；实际写入位置恒为第 2 行

REQUIRED_HEADERS = ["编号", "日期", "日报内容"]


def normalize_header_name(header: str) -> str:
    """Normalize Daily_Logs header names for schema comparison.

    The online sheet may wrap business headers with full-width Chinese brackets,
    e.g. 【日期】/【日报内容】. Strip the wrapper only at the schema-compat layer;
    keep raw headers unchanged for logging and DLQ evidence.
    """
    value = str(header).strip()
    if value.startswith("【") and value.endswith("】"):
        value = value[1:-1].strip()
    return value


def headers_match_required(headers: list[str]) -> bool:
    return [normalize_header_name(h) for h in headers[: len(REQUIRED_HEADERS)]] == REQUIRED_HEADERS


DEFAULT_TASK_STATS_WIKI_URL = "https://bytedance.larkoffice.com/sheets/TnNYsLq9phIJwutJGwBl730ygjd"
DEFAULT_TASK_STATS_SHEET_NAME = "任务库"
TASK_STATUS_COLUMN = "完成情况"
TASK_OPENED_STATUSES = {"进行中", "准备中", "已开启", "开启中"}
TASK_COMPLETED_STATUSES = {"已完成", "完成"}
TASK_PAUSED_STATUSES = {"暂停", "已暂停", "暂停中", "搁置"}
UNRESOLVED_SENTINEL = "⚠️[数据断链_待自愈]"
TASK_STATS_PLACEHOLDERS = {
    "{{TASK_STATS_TODAY_OPENED}}": "opened",
    "{{TASK_STATS_TODAY_COMPLETED}}": "completed",
    "{{TASK_STATS_TODAY_PAUSED}}": "paused",
    "{{TODAY_OPENED_TASKS}}": "opened",
    "{{TODAY_COMPLETED_TASKS}}": "completed",
    "{{TODAY_PAUSED_TASKS}}": "paused",
    "{{TASK_STATS_SUMMARY}}": "summary",
    "{{TASK_STATS_TABLE}}": "table",
}


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def _skill_root() -> Path:
    # scripts/xxx.py -> skill_root
    return Path(__file__).resolve().parents[1]


def _workspace_root() -> Path:
    # 向上回溯直到找到 inner_skills 目录（Aime VM 约定）
    p = Path(__file__).resolve()
    for parent in [p, *p.parents]:
        if (parent / "inner_skills").exists():
            return parent
    # 兜底：当前进程 cwd
    return Path.cwd().resolve()


def _parse_mcp_result_to_obj(res_text: str):
    """兼容 MCP 返回格式。

    call_aime_tool 可能返回：
    - JSON 字符串
    - Python 字面量字符串（如 ['a.xlsx']）
    - 纯文本
    """
    try:
        return json.loads(res_text)
    except Exception:
        pass

    try:
        return ast.literal_eval(res_text)
    except Exception:
        pass

    return res_text


def mcp_download_lark_sheet(document_url: str) -> list[str]:
    """通过当前可用的 lark-sheets CLI/MCP 导出 xlsx。

    历史实现依赖 `mcp:lark_lark_download`，当前环境该工具缺失；这里改为
    lark-sheets 原生导出能力（等价 MCP 簇能力）读取在线表格，避免 18:00 归档链路
    再被缺失工具卡死。
    """
    import subprocess
    import tempfile

    out_dir = Path.cwd().resolve() / ".tmp" / "aime_daily_logs_exports"
    _ensure_dir(out_dir)
    out_name = f"lark_sheet_export_{dt.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
    out_path = out_dir / out_name
    rel_out_path = str(out_path.relative_to(Path.cwd().resolve()))
    cmd = [
        _lark_sheets_cli_path(),
        "sheets",
        "+export",
        "--url",
        document_url,
        "--file-extension",
        "xlsx",
        "--output-path",
        rel_out_path,
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode != 0:
        raise RuntimeError(
            "lark-sheets +export 执行失败，无法读取在线表格："
            f"\ncmd={' '.join(cmd)}\nstderr={res.stderr}\nstdout={res.stdout}"
        )

    exported_paths: list[str] = []
    try:
        obj = json.loads(res.stdout)
        data = obj.get("data", {}) if isinstance(obj, dict) else {}
        for key in ("file_path", "path", "output_path"):
            if isinstance(data, dict) and data.get(key):
                exported_paths.append(str(data[key]))
        if isinstance(data, dict) and isinstance(data.get("file_paths"), list):
            exported_paths.extend(str(p) for p in data["file_paths"])
    except Exception:
        pass

    exported_paths.append(str(out_path))
    existing = [p for p in exported_paths if Path(p).exists()]
    if existing:
        return existing

    # 兼容 CLI 输出里只打印路径文本的情况
    matches = re.findall(r'(/[^"]+\.xlsx)', res.stdout + "\n" + res.stderr)
    existing = [p for p in matches if Path(p).exists()]
    if existing:
        return existing

    raise RuntimeError(
        "lark-sheets +export 未产生可用 xlsx 文件："
        f"\nexpected={out_path}\nstdout={res.stdout}\nstderr={res.stderr}"
    )


def pick_xlsx(paths: list[str]) -> str:
    for p in paths:
        if str(p).lower().endswith(".xlsx"):
            return p
    raise RuntimeError(f"下载结果中未找到 .xlsx 文件：{paths}")


def read_sheet_headers(xlsx_path: str, sheet_name: str) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"[Debug] Sheet names: {wb.sheetnames}")
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"xlsx 中未找到工作表 {sheet_name}。可用工作表：{wb.sheetnames}"
        )
    ws = wb[sheet_name]
    headers: list[str] = []
    for col in range(1, 4):
        v = ws.cell(row=1, column=col).value
        headers.append(str(v).strip() if v is not None else "")
    print(f"[Debug] Explicit headers: {headers}")
    return headers


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_header_row(ws, max_scan: int = 32) -> list[str]:
    headers = [_normalize_cell(ws.cell(row=1, column=col).value) for col in range(1, max_scan + 1)]
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def _iter_sheet_rows(ws, header_count: int, start_row: int = 2, max_scan_rows: int = 5000, blank_streak_limit: int = 50):
    blank_streak = 0
    for row_idx in range(start_row, max_scan_rows + 1):
        row_values = [
            _normalize_cell(ws.cell(row=row_idx, column=col).value)
            for col in range(1, header_count + 1)
        ]
        if any(row_values):
            blank_streak = 0
            yield row_idx, row_values
            continue
        blank_streak += 1
        if blank_streak >= blank_streak_limit:
            break


def compute_task_status_stats(xlsx_path: str, sheet_name: str) -> dict[str, int | str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"任务库 xlsx 中未找到工作表 {sheet_name}。可用工作表：{wb.sheetnames}"
        )
    ws = wb[sheet_name]

    headers = _read_header_row(ws)
    if TASK_STATUS_COLUMN not in headers:
        raise RuntimeError(
            f"任务库缺少状态列 {TASK_STATUS_COLUMN}。实际表头：{headers}"
        )

    status_idx = headers.index(TASK_STATUS_COLUMN)
    opened = 0
    completed = 0
    paused = 0
    recognized = 0
    unknown_statuses: dict[str, int] = {}

    for _row_idx, row_values in _iter_sheet_rows(ws, header_count=len(headers)):
        status = row_values[status_idx] if status_idx < len(row_values) else ""
        if not status:
            continue
        if status in TASK_OPENED_STATUSES:
            opened += 1
            recognized += 1
        elif status in TASK_COMPLETED_STATUSES:
            completed += 1
            recognized += 1
        elif status in TASK_PAUSED_STATUSES:
            paused += 1
            recognized += 1
        else:
            unknown_statuses[status] = unknown_statuses.get(status, 0) + 1

    summary = f"开启 {opened} 个，完成 {completed} 个，暂停 {paused} 个"
    table = (
        "| 状态 | 数量 |\n"
        "| --- | ---: |\n"
        f"| 开启 | {opened} |\n"
        f"| 完成 | {completed} |\n"
        f"| 暂停 | {paused} |"
    )
    return {
        "opened": opened,
        "completed": completed,
        "paused": paused,
        "recognized": recognized,
        "unknown_statuses": json.dumps(unknown_statuses, ensure_ascii=False),
        "summary": summary,
        "table": table,
    }


def replace_task_stats_placeholders(content: str, task_stats: dict[str, int | str]) -> str:
    resolved = content
    for placeholder, key in TASK_STATS_PLACEHOLDERS.items():
        if placeholder in resolved:
            resolved = resolved.replace(placeholder, str(task_stats[key]))

    patterns = {
        "opened": [r"(今日?开启\s*[：:：]\s*)⚠️\[数据断链_待自愈\]", r"(开启\s*[：:：]\s*)⚠️\[数据断链_待自愈\]"],
        "completed": [r"(今日?完成\s*[：:：]\s*)⚠️\[数据断链_待自愈\]", r"(完成\s*[：:：]\s*)⚠️\[数据断链_待自愈\]"],
        "paused": [r"(今日?暂停\s*[：:：]\s*)⚠️\[数据断链_待自愈\]", r"(暂停\s*[：:：]\s*)⚠️\[数据断链_待自愈\]"],
    }
    for key, regex_list in patterns.items():
        for pattern in regex_list:
            resolved = re.sub(pattern, lambda m: f"{m.group(1)}{task_stats[key]}", resolved)

    if UNRESOLVED_SENTINEL in resolved:
        if "任务状态汇总表" in resolved:
            resolved = resolved.replace(UNRESOLVED_SENTINEL, str(task_stats["table"]), 1)
        elif "任务状态汇总" in resolved:
            resolved = resolved.replace(UNRESOLVED_SENTINEL, str(task_stats["summary"]), 1)

    if "任务状态汇总" not in resolved and "| 状态 | 数量 |" not in resolved:
        resolved = f"任务状态汇总：{task_stats['summary']}\n\n{resolved}"

    return resolved


def resolve_task_stats_content(
    content: str,
    task_stats_wiki_url: str,
    task_stats_sheet_name: str,
    workspace_root: Path,
) -> tuple[str, dict[str, int | str], str]:
    file_paths = mcp_download_lark_sheet(task_stats_wiki_url)
    xlsx_path = pick_xlsx(file_paths)
    if not os.path.isabs(xlsx_path):
        xlsx_path = str((workspace_root / xlsx_path).resolve())
    task_stats = compute_task_status_stats(xlsx_path, task_stats_sheet_name)
    resolved_content = replace_task_stats_placeholders(content, task_stats)
    if UNRESOLVED_SENTINEL in resolved_content:
        raise RuntimeError(
            f"任务库统计已拉取，但日报内容仍残留未替换占位符 {UNRESOLVED_SENTINEL}。"
        )
    return resolved_content, task_stats, xlsx_path


def list_existing_ids(xlsx_path: str, sheet_name: str) -> set[str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    ids: set[str] = set()
    # 第一列：编号；从第 2 行开始
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            ids.add(s)
    return ids


# ---- lark-sheets CLI 直读：避免 MCP 下载缓存 ----

def _lark_sheets_cli_path() -> str:
    return str(
        _workspace_root() / "inner_skills" / "lark-sheets" / "bin" / "lark-sheets-cli"
    )


def _run_lark_sheets(args: list[str]) -> dict | None:
    import subprocess

    cmd = [_lark_sheets_cli_path()] + args
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode != 0:
        print(f"[CliError] {res.stderr}")
        return None
    try:
        return json.loads(res.stdout)
    except Exception:
        print(f"[CliParseError] stdout: {res.stdout[:400]}")
        return None


def _resolve_sheet_id(sheet_url: str, sheet_name: str) -> str | None:
    info = _run_lark_sheets(["sheets", "+info", "--url", sheet_url])
    if not info or not info.get("ok"):
        return None
    for s in info["data"]["sheets"]["sheets"]:
        if s.get("title") == sheet_name:
            return s.get("sheet_id")
    return None


def _get_row_count(sheet_url: str, sheet_name: str) -> int | None:
    info = _run_lark_sheets(["sheets", "+info", "--url", sheet_url])
    if not info or not info.get("ok"):
        return None
    for s in info["data"]["sheets"]["sheets"]:
        if s.get("title") == sheet_name:
            return s.get("grid_properties", {}).get("row_count")
    return None


def list_existing_ids_via_cli(sheet_url: str, sheet_name: str) -> set[str] | None:
    """通过 lark-sheets CLI 直读首列所有【编号】。

    避免 MCP 下载缓存导致的"刚写完读不到"问题。
    """
    sheet_id = _resolve_sheet_id(sheet_url, sheet_name)
    if not sheet_id:
        return None
    row_count = _get_row_count(sheet_url, sheet_name) or 200
    rng = f"{sheet_id}!A2:A{max(2, row_count)}"
    res = _run_lark_sheets(
        ["sheets", "+read", "--url", sheet_url, "--sheet-id", sheet_id, "--range", rng]
    )
    if not res or not res.get("ok"):
        return None
    rows = res.get("data", {}).get("valueRange", {}).get("values", []) or []
    ids: set[str] = set()
    for row in rows:
        v = row[0] if isinstance(row, list) and row else None
        if v is None:
            continue
        s = str(v).strip()
        if s:
            ids.add(s)
    return ids


def find_row_by_primary_key_via_cli(
    sheet_url: str, sheet_name: str, primary_key: str, col_count: int = 3
) -> tuple[int, list[str | None]] | None:
    """通过 lark-sheets CLI 直读，按主键回捞最末一条匹配行（绕开下载缓存）。"""
    sheet_id = _resolve_sheet_id(sheet_url, sheet_name)
    if not sheet_id:
        return None
    row_count = _get_row_count(sheet_url, sheet_name) or 200

    end_col_letter = chr(ord("A") + col_count - 1)
    rng = f"{sheet_id}!A1:{end_col_letter}{max(1, row_count)}"
    res = _run_lark_sheets(
        ["sheets", "+read", "--url", sheet_url, "--sheet-id", sheet_id, "--range", rng]
    )
    if not res or not res.get("ok"):
        return None
    rows = res.get("data", {}).get("valueRange", {}).get("values", []) or []

    last_idx: int | None = None
    last_row: list[str | None] | None = None
    # 第 1 行是表头；从第 2 行开始扫
    for idx, row in enumerate(rows[1:], start=2):
        first = row[0] if isinstance(row, list) and row else None
        if first is None:
            continue
        if str(first).strip() == primary_key:
            # 补齐到 col_count 长度
            full = list(row) + [None] * (col_count - len(row))
            last_idx = idx
            last_row = [None if v is None else str(v) for v in full[:col_count]]

    if last_idx is None or last_row is None:
        return None
    return last_idx, last_row


# ============================================================================
# V6.3 规则 1：降序插入法（Descending Insert）
# 走 lark-cli（MCP 同源通道）：+dim-insert 在第 2 行插空行 → +cells-set 写 A2:C2
# ============================================================================

LARK_CLI = "lark-cli"


def _run_lark_cli(args: list[str]) -> dict:
    """执行 lark-cli 并解析 JSON envelope；失败即 raise（禁止静默）。"""
    import subprocess

    cmd = [LARK_CLI] + args
    print(f"[lark-cli] {' '.join(cmd)}")
    res = subprocess.run(cmd, capture_output=True, text=True)
    if res.returncode != 0:
        raise RuntimeError(
            f"lark-cli 执行失败（exit={res.returncode}）：\ncmd={' '.join(cmd)}"
            f"\nstderr={res.stderr}\nstdout={res.stdout[:2000]}"
        )
    try:
        obj = json.loads(res.stdout)
    except Exception as exc:
        raise RuntimeError(f"lark-cli 输出无法解析为 JSON：{exc}\nstdout={res.stdout[:2000]}")
    if not obj.get("ok"):
        raise RuntimeError(f"lark-cli 返回 ok=false：{json.dumps(obj, ensure_ascii=False)[:2000]}")
    return obj


def insert_top_blank_row(sheet_url: str, sheet_name: str) -> None:
    """在表头（第 1 行）下方物理插入 1 个空行，为最新日报腾出置顶位置。"""
    _run_lark_cli(
        [
            "sheets",
            "+dim-insert",
            "--url",
            sheet_url,
            "--sheet-name",
            sheet_name,
            "--position",
            str(TOP_ROW_INDEX),
            "--count",
            "1",
            "--inherit-style",
            "after",
        ]
    )
    print(f"[DescendingInsert] 已在第 {TOP_ROW_INDEX} 行插入空行（表头下方置顶位）。")


def write_top_row(sheet_url: str, sheet_name: str, row_values: list[str]) -> None:
    """把 [编号, 日期, 日报内容] 写入 A2:C2（结构化 value 通道，非 csv-put）。"""
    if len(row_values) != 3:
        raise ValueError(f"write_top_row 需要 3 列数据，当前：{row_values}")
    cells = [[{"value": str(v)} for v in row_values]]
    _run_lark_cli(
        [
            "sheets",
            "+cells-set",
            "--url",
            sheet_url,
            "--sheet-name",
            sheet_name,
            "--range",
            f"{sheet_name}!A{TOP_ROW_INDEX}:C{TOP_ROW_INDEX}",
            "--cells",
            json.dumps(cells, ensure_ascii=False),
        ]
    )
    print(f"[DescendingInsert] 已写入 A{TOP_ROW_INDEX}:C{TOP_ROW_INDEX}。")


def descending_insert_daily_log(sheet_url: str, sheet_name: str, row_values: list[str]) -> None:
    """降序插入法统一入口：插空行 → 写置顶行。禁止 append-to-tail。"""
    insert_top_blank_row(sheet_url, sheet_name)
    write_top_row(sheet_url, sheet_name, row_values)


def read_daily_logs_rows(sheet_url: str, sheet_name: str, max_rows: int = 400) -> list[list[str]]:
    """RAW 回读 A:C 全量（含表头），返回 [[a,b,c], ...]，索引 0 即表头行。"""
    obj = _run_lark_cli(
        [
            "sheets",
            "+cells-get",
            "--url",
            sheet_url,
            "--sheet-name",
            sheet_name,
            "--range",
            f"{sheet_name}!A1:C{max_rows}",
            "--include",
            "value",
        ]
    )
    ranges = obj.get("data", {}).get("ranges", []) or []
    if not ranges:
        raise RuntimeError("RAW 回读失败：+cells-get 未返回任何 range。")
    raw_cells = ranges[0].get("cells", []) or []
    rows: list[list[str]] = []
    for raw_row in raw_cells:
        vals = []
        for i in range(3):
            cell = raw_row[i] if i < len(raw_row) and isinstance(raw_row[i], dict) else {}
            vals.append(_normalize_cell(cell.get("value")))
        rows.append(vals)
    # 去掉尾部全空行
    while rows and not any(rows[-1]):
        rows.pop()
    return rows


# ============================================================================
# V6.3 规则 2：写入后三断言熔断（治表）
# ============================================================================


def _parse_date(value: str) -> dt.date | None:
    s = _normalize_cell(value)
    if not s:
        return None
    s = s.split(" ")[0].replace(".", "-").replace("/", "-")
    parts = s.split("-")
    if len(parts) != 3:
        return None
    try:
        return dt.date(int(parts[0]), int(parts[1]), int(parts[2]))
    except Exception:
        return None


def assert_top_row_is_today(rows: list[list[str]], expected_date: str) -> dict:
    """断言 A：B2 == 当日日期（最新日报必须置顶）。"""
    if len(rows) < 2:
        raise AssertionError(f"[断言A-FAIL] Daily_Logs 缺少数据行（rows={len(rows)}），无法验证置顶。")
    b2 = rows[1][1]
    got = _parse_date(b2)
    want = _parse_date(expected_date)
    if want is None:
        raise AssertionError(f"[断言A-FAIL] 期望日期无法解析：{expected_date!r}")
    if got is None or got != want:
        raise AssertionError(
            f"[断言A-FAIL] 最新日报未置顶：B2={b2!r}，期望={expected_date!r}。"
            " 疑似仍在使用 append-to-tail 或插行位置错误。"
        )
    return {"assertion": "A_top_row_is_today", "result": "PASS", "b2": b2}


def assert_date_desc(rows: list[list[str]]) -> dict:
    """断言 B：B 列（排除表头）严格降序，不允许乱序。"""
    dates: list[tuple[int, dt.date]] = []
    unparsable: list[tuple[int, str]] = []
    for offset, row in enumerate(rows[1:], start=2):
        raw = row[1]
        if not raw:
            continue
        d = _parse_date(raw)
        if d is None:
            unparsable.append((offset, raw))
            continue
        dates.append((offset, d))
    if unparsable:
        raise AssertionError(f"[断言B-FAIL] 存在无法解析的日期单元格（疑似错列/脏写）：{unparsable[:10]}")
    violations = [
        {"row": dates[i + 1][0], "prev": str(dates[i][1]), "curr": str(dates[i + 1][1])}
        for i in range(len(dates) - 1)
        if dates[i][1] < dates[i + 1][1]
    ]
    if violations:
        raise AssertionError(f"[断言B-FAIL] B 列日期非降序，乱序点：{violations[:10]}")
    return {"assertion": "B_date_desc", "result": "PASS", "checked_rows": len(dates)}


def assert_no_empty_cells(rows: list[list[str]]) -> dict:
    """断言 C：A2 / B2 / C2 均非空（防错列 / 半行写入）。"""
    if len(rows) < 2:
        raise AssertionError("[断言C-FAIL] Daily_Logs 缺少数据行，无法验证非空。")
    a2, b2, c2 = rows[1][0], rows[1][1], rows[1][2]
    empties = [name for name, val in (("A2", a2), ("B2", b2), ("C2", c2)) if not val]
    if empties:
        raise AssertionError(f"[断言C-FAIL] 置顶行存在空单元格（错列/半行写入）：{empties}")
    return {"assertion": "C_no_empty_cells", "result": "PASS", "a2": a2, "b2": b2, "c2_len": len(c2)}


def assert_daily_logs_invariants(rows: list[list[str]], expected_date: str) -> list[dict]:
    """三断言统一入口：A/B/C 任一失败即 raise 熔断，绝不静默通过。"""
    evidence = [
        assert_top_row_is_today(rows, expected_date),
        assert_date_desc(rows),
        assert_no_empty_cells(rows),
    ]
    print("[Assertions] 三断言全部 PASS：")
    print(json.dumps(evidence, ensure_ascii=False, indent=2))
    return evidence


def gen_daily_log_id(date_str: str, existing: set[str]) -> str:
    yyyymmdd = date_str.replace("-", "")
    base = f"DL-{yyyymmdd}"
    if base not in existing:
        return base

    # 冲突时追加递增后缀：-02, -03 ...
    i = 2
    while True:
        cand = f"{base}-{i:02d}"
        if cand not in existing:
            return cand
        i += 1


def read_row_values(xlsx_path: str, sheet_name: str, row_index: int, col_count: int) -> list[str | None]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    row = []
    for col in range(1, col_count + 1):
        row.append(ws.cell(row=row_index, column=col).value)
    # 统一转 str 便于比对
    return [None if v is None else str(v) for v in row]


def find_row_by_primary_key(
    xlsx_path: str, sheet_name: str, primary_key: str, col_count: int = 3
) -> tuple[int, list[str | None]] | None:
    """按主键（第 1 列：编号）回捞最后一条匹配行。

    底部追加模式下，新写入的行位于表尾，因此需要从下往上扫描，命中即返回。
    返回 (row_index, row_values)；未命中返回 None。
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # 收集所有非空首列的行（read_only 模式不支持随机访问，整体迭代一次）
    last_match_idx: int | None = None
    last_match_row: list[str | None] | None = None

    for idx, row in enumerate(
        ws.iter_rows(min_row=2, max_col=col_count, values_only=True), start=2
    ):
        first = row[0]
        if first is None:
            continue
        if str(first).strip() == primary_key:
            last_match_idx = idx
            last_match_row = [None if v is None else str(v) for v in row]

    if last_match_idx is None or last_match_row is None:
        return None
    return last_match_idx, last_match_row


def write_dlq(skill_root: Path, payload: dict) -> Path:
    dlq_dir = skill_root / "assets" / "dlq"
    _ensure_dir(dlq_dir)

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    p = dlq_dir / f"daily_logs_dlq_{ts}.json"
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return p


def snapshot(skill_root: Path, xlsx_path: str, stage: str) -> Path:
    snap_dir = skill_root / "assets" / "snapshots"
    _ensure_dir(snap_dir)

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = snap_dir / f"daily_logs_{stage}_{ts}.xlsx"
    shutil.copyfile(xlsx_path, dst)
    return dst


def main() -> int:
    parser = argparse.ArgumentParser(description="Daily_Logs 零信任安全插入（一键脚本）")
    parser.add_argument("--sheet-url", default=DEFAULT_SHEET_URL, help="飞书表格链接")
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME, help="工作表名称")
    parser.add_argument("--date", default=dt.date.today().strftime("%Y-%m-%d"), help="日期 YYYY-MM-DD")
    parser.add_argument("--content", required=True, help="日报内容（建议约 100 字）")
    parser.add_argument(
        "--task-stats-wiki-url",
        default=DEFAULT_TASK_STATS_WIKI_URL,
        help="任务库 Wiki 链接，用于前置拉取任务状态统计",
    )
    parser.add_argument(
        "--task-stats-sheet-name",
        default=DEFAULT_TASK_STATS_SHEET_NAME,
        help="任务库工作表名称",
    )
    parser.add_argument(
        "--row-index",
        type=int,
        default=DEFAULT_ROW_INDEX,
        help="（已废弃）兼容入参；V6.3 起底层强制走「降序插入法」写入第 2 行，此参数不再生效",
    )
    parser.add_argument("--dry-run", action="store_true", help="只做 Schema 回捞与主键生成，不实际写入")

    args = parser.parse_args()

    skill_root = _skill_root()
    workspace_root = _workspace_root()

    resolved_content, task_stats, task_stats_xlsx = resolve_task_stats_content(
        content=args.content,
        task_stats_wiki_url=args.task_stats_wiki_url,
        task_stats_sheet_name=args.task_stats_sheet_name,
        workspace_root=workspace_root,
    )
    print("[TaskStats] source:")
    print(task_stats_xlsx)
    print("[TaskStats] values:")
    print(json.dumps(task_stats, ensure_ascii=False))

    # 1) MCP 下载（读 Schema）
    file_paths = mcp_download_lark_sheet(args.sheet_url)
    xlsx_path = pick_xlsx(file_paths)
    if not os.path.isabs(xlsx_path):
        xlsx_path = str((workspace_root / xlsx_path).resolve())

    pre_snap = snapshot(skill_root, xlsx_path, stage="pre")

    headers = read_sheet_headers(xlsx_path, args.sheet_name)
    if not headers_match_required(headers):
        dlq_file = write_dlq(
            skill_root,
            {
                "type": "Daily_Logs_SchemaMismatch",
                "timestamp": dt.datetime.now().isoformat(),
                "sheet_url": args.sheet_url,
                "sheet_name": args.sheet_name,
                "expected_headers_prefix": REQUIRED_HEADERS,
                "actual_headers": headers,
                "payload": {
                    "date": args.date,
                    "content": resolved_content,
                    "task_stats": task_stats,
                    "task_stats_wiki_url": args.task_stats_wiki_url,
                    "task_stats_sheet_name": args.task_stats_sheet_name,
                },
                "note": "⚠️[数据断链_待自愈] Schema 不匹配，已熔断写入。",
            },
        )
        raise RuntimeError(
            "Daily_Logs 表头 Schema 不符合预期，已熔断并落 DLQ："
            f"\n- 预期前缀：{REQUIRED_HEADERS}"
            f"\n- 实际表头：{headers}"
            f"\n- DLQ：{dlq_file}"
        )

    # 主键生成：优先通过 lark-sheets CLI 直读（避免 MCP 下载缓存），失败时回退 xlsx
    cli_existing = list_existing_ids_via_cli(args.sheet_url, args.sheet_name)
    if cli_existing is not None:
        print(f"[Schema] existing_ids via CLI: {len(cli_existing)} rows scanned")
        existing_ids = cli_existing
    else:
        print("[Schema] CLI 直读失败，回退使用 xlsx 缓存读取 existing_ids")
        existing_ids = list_existing_ids(xlsx_path, args.sheet_name)
    row_id = gen_daily_log_id(args.date, existing_ids)
    
    row_data = [[row_id, args.date, resolved_content]]

    # 【治本封堵】强契约校验：必须写满 3 列 [ID, Date, Content] 且均不能为空
    if not isinstance(row_data, list) or len(row_data) != 1 or len(row_data[0]) != 3:
        raise ValueError(f"CRITICAL [治本封堵]: row_data 必须为 1x3 数组，当前为: {row_data}")
    
    rid, rdate, rcont = row_data[0]
    if not str(rid).strip() or not str(rdate).strip() or not str(rcont).strip():
        raise ValueError(f"CRITICAL [治本封堵]: [ID, Date, Content] 均不能为空，当前行: {row_data[0]}")
    
    print("[Schema] headers:")
    print(json.dumps(headers, ensure_ascii=False))
    print("[Insert] row_data:")
    print(json.dumps(row_data, ensure_ascii=False))
    print("[Insert] resolved_content:")
    print(resolved_content)
    print(f"[Snapshot] pre: {pre_snap}")

    if args.dry_run:
        print(f"[DryRun] 写入模式=降序插入法（target_row={TOP_ROW_INDEX}，禁止 append-to-tail）")
        try:
            preflight_rows = read_daily_logs_rows(args.sheet_url, args.sheet_name)
            print(json.dumps(assert_date_desc(preflight_rows), ensure_ascii=False))
        except Exception as pre_err:
            print(f"[DryRun][Preflight-WARN] 现存表体断言预检未通过：{pre_err}")
        print("[DryRun] 已跳过实际写入。")
        return 0

    # 2) 降序插入写入（V6.3 规则 1）：第 2 行物理插入 + 写 A2:C2，禁止 append-to-tail
    try:
        descending_insert_daily_log(args.sheet_url, args.sheet_name, [row_id, args.date, resolved_content])
    except Exception as write_err:
        dlq_file = write_dlq(
            skill_root,
            {
                "type": "Daily_Logs_DescendingInsertFailed",
                "timestamp": dt.datetime.now().isoformat(),
                "sheet_url": args.sheet_url,
                "sheet_name": args.sheet_name,
                "mode": "descending_insert",
                "target_row_index": TOP_ROW_INDEX,
                "row_data": row_data,
                "error": str(write_err),
                "note": "⚠️[数据断链_待自愈] 降序插入写入失败，已落 DLQ。",
            },
        )
        raise RuntimeError(
            f"降序插入写入失败，已熔断并落 DLQ：\n- DLQ：{dlq_file}\n- error：{write_err}"
        )

    # 3) 写后即读（RAW 原子锁）+ 三断言熔断（V6.3 规则 2）
    time.sleep(2)

    expected_str = [str(row_id), str(args.date), str(resolved_content)]
    rows = read_daily_logs_rows(args.sheet_url, args.sheet_name)

    print("[ReadAfterWrite] top row (raw array):")
    print(json.dumps([rows[1]] if len(rows) > 1 else [], ensure_ascii=False))

    # 3.1 逐字段核对置顶行
    read_back = rows[1] if len(rows) > 1 else []
    if read_back != expected_str:
        dlq_file = write_dlq(
            skill_root,
            {
                "type": "Daily_Logs_ReadAfterWriteMismatch",
                "timestamp": dt.datetime.now().isoformat(),
                "sheet_url": args.sheet_url,
                "sheet_name": args.sheet_name,
                "mode": "descending_insert",
                "located_row_index": TOP_ROW_INDEX,
                "primary_key": row_id,
                "expected": expected_str,
                "read_back": read_back,
                "snapshots": {"pre": str(pre_snap)},
                "note": "⚠️[数据断链_待自愈] 写后即读不一致，已熔断并落 DLQ。",
            },
        )
        raise RuntimeError(
            "写后即读 RAW 校验失败（置顶行不一致），已熔断并落 DLQ："
            f"\n- expected：{expected_str}\n- read_back：{read_back}\n- DLQ：{dlq_file}"
        )

    # 3.2 三断言（A 置顶 / B 降序 / C 非空），任一失败即 raise
    try:
        evidence = assert_daily_logs_invariants(rows, args.date)
    except AssertionError as assert_err:
        dlq_file = write_dlq(
            skill_root,
            {
                "type": "Daily_Logs_InvariantAssertionFailed",
                "timestamp": dt.datetime.now().isoformat(),
                "sheet_url": args.sheet_url,
                "sheet_name": args.sheet_name,
                "mode": "descending_insert",
                "primary_key": row_id,
                "expected_date": args.date,
                "top_rows_preview": rows[:5],
                "error": str(assert_err),
                "note": "⚠️[数据断链_待自愈] 三断言熔断，已落 DLQ。",
            },
        )
        raise RuntimeError(
            f"Daily_Logs 三断言熔断，已落 DLQ：\n- DLQ：{dlq_file}\n- {assert_err}"
        )

    print("[OK] 写入并核对一致（降序插入置顶 + RAW 原子锁 + 三断言 PASS）。")
    print("[Evidence] " + json.dumps(evidence, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as e:
        print("[FATAL]", str(e))
        traceback.print_exc()
        raise
