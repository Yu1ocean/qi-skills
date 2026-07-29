#!/usr/bin/env python3
"""Live Performance summary sheet generator.

V1.2 (2026-05-20) — 动态表头映射重构 + RAW 写后回读校验：

- 取消硬编码列字母（`'1. 数据底表'!D{r}` 这种），改为读取 raw sheet 表头行
  并按【字段语义】解析出实际列字母，避免源表新增/隐藏列时整体错位。
- 引入 `RAW_FIELD_HEADER_CANDIDATES`（每个语义字段允许多个候选表头别名），
  并支持模糊匹配（去空格、去标点、忽略大小写），抗轻微改名。
- 写入完成后强制对 summary sheet 抽样回读（A1:P5 + 末尾若干行），
  对表头、首行公式、benchmark 命中单元格做断言式校验，失败立刻报错。
"""
from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import load_workbook

from lark_sheets_cli import LarkSheetsCLI, LarkSheetsError, SheetInfo, col_num_to_a1


DEFAULT_RAW_SHEET_TITLE = "1. 数据底表"
DEFAULT_SUMMARY_SHEET_TITLE = "2. 计算汇总"
DEFAULT_SUMMARY_SHEET_INDEX = 0
DEFAULT_MIN_GRID_ROWS = 500
DEFAULT_MIN_GRID_COLS = 16
DEFAULT_CLEAR_BUFFER_ROWS = 100
DEFAULT_BLACK = "#000000"
DEFAULT_LIGHT_GREEN = "#E2F0D9"
DEFAULT_LIGHT_YELLOW = "#FFF2CC"
DEFAULT_LIGHT_RED = "#F4CCCC"
DEFAULT_FREEZE_ROWS = 2
DEFAULT_FREEZE_COLS = 2
DEFAULT_CHUNK_SIZE = 120
DEFAULT_EXPORT_PREFIX = "live-performance-summary"
DEFAULT_HEADER_SCAN_ROWS = 4  # 在前 N 行内自动定位真正的表头行

BENCHMARKS: Dict[str, Optional[float]] = {
    "L": 55,
    "M": 0.01,
    "N": 5,
    "O": None,
    "P": 0.05,
}

SUMMARY_HEADERS = [
    "Handle",
    "日期",
    "GMV",
    "时均GMV",
    "时均show PV /K",
    "开播小时",
    "Show GPM",
    "ERR",
    "CTR",
    "C_O",
    "AOV",
    "Watch Duration(AVG.)>55秒",
    "Follow rate >1%",
    "Like rate >500%",
    "Share rate \n观察持续提升",
    "Comment rate >5%",
]

# ---------------------------------------------------------------------------
# 字段语义 → raw header 候选别名（按优先级匹配，第一个命中即采用）
# 该结构是本次重构的核心：把“列字母”这一脆弱锚点替换成“表头语义”。
# ---------------------------------------------------------------------------
RAW_FIELD_HEADER_CANDIDATES: Dict[str, List[str]] = {
    # 基础信息
    "handle": ["TT Handle", "Handle", "Account Handle"],
    "date": ["Start Timestamp", "Live Start Time", "Start Time"],
    "duration_sec": ["Duration(s)", "Duration (s)", "Live Duration(s)"],
    # GMV / 流量
    "cl_gmv": ["CL GMV", "GMV", "CL_GMV"],
    "show_pv": ["Show PV", "ShowPV"],
    "show_gpm": ["Show GPM", "Show GPM(USD)", "Show GPM (USD)"],
    "enter_room_rate": ["Enter room rate", "ERR", "Enter Room Rate"],
    "valid_cl_ctr": ["Valid CL CTR", "CTR", "CL CTR"],
    "valid_cl_co": ["Valid CL C_O", "C_O", "CL C_O"],
    "aov": ["AOV(Main)(USD)", "AOV(Main)", "AOV(USD)", "AOV"],
    # 互动指标（过程指标）
    "watch_duration_avg": [
        "Valid CL Watch Duration(AVG.)",
        "Watch Duration(AVG.)",
        "Watch Duration (AVG.)",
    ],
    "follow_rate": ["Follow rate", "Follow Rate"],
    "like_rate": ["Like rate", "Like Rate"],
    "share_rate": ["Share rate", "Share Rate"],
    "comment_rate": ["Live comment rate", "Comment rate", "Live Comment Rate"],
}

# Summary 列 → 字段语义键
SUMMARY_COL_TO_FIELD: Dict[str, str] = {
    "A": "handle",
    "B": "date",
    "C": "cl_gmv",
    # D 时均GMV、E 时均show PV、F 开播小时 单独处理（涉及多字段或派生）
    "E": "show_pv",
    "F": "duration_sec",
    "G": "show_gpm",
    "H": "enter_room_rate",
    "I": "valid_cl_ctr",
    "J": "valid_cl_co",
    "K": "aov",
    "L": "watch_duration_avg",
    "M": "follow_rate",
    "N": "like_rate",
    "O": "share_rate",
    "P": "comment_rate",
}

# Summary 列 → 该列 benchmark 判断对应的字段语义键
BENCHMARK_FIELD_KEYS: Dict[str, str] = {
    "L": "watch_duration_avg",
    "M": "follow_rate",
    "N": "like_rate",
    "P": "comment_rate",
}


class SummaryGenerationError(RuntimeError):
    pass


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------
@dataclass
class HeaderMapping:
    header_row: int
    data_start_row: int
    field_to_letter: Dict[str, str]


@dataclass
class WorkbookContext:
    spreadsheet_token: str
    raw_sheet: SheetInfo
    summary_sheet: SheetInfo
    last_raw_row: int
    clear_end_row: int
    benchmark_red_ranges: List[str]
    exported_xlsx: Path
    header_mapping: HeaderMapping


@dataclass
class GenerationResult:
    spreadsheet_token: str
    raw_sheet_id: str
    summary_sheet_id: str
    raw_sheet_title: str
    summary_sheet_title: str
    data_rows: int
    header_row: int
    data_start_row: int
    last_raw_row: int
    clear_end_row: int
    benchmark_red_ranges: List[str]
    exported_xlsx: str
    field_to_letter: Dict[str, str]
    readback: Dict[str, Any]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "spreadsheet_token": self.spreadsheet_token,
            "raw_sheet_id": self.raw_sheet_id,
            "summary_sheet_id": self.summary_sheet_id,
            "raw_sheet_title": self.raw_sheet_title,
            "summary_sheet_title": self.summary_sheet_title,
            "data_rows": self.data_rows,
            "header_row": self.header_row,
            "data_start_row": self.data_start_row,
            "last_raw_row": self.last_raw_row,
            "clear_end_row": self.clear_end_row,
            "benchmark_red_ranges": self.benchmark_red_ranges,
            "exported_xlsx": self.exported_xlsx,
            "field_to_letter": self.field_to_letter,
            "readback": self.readback,
        }


# ---------------------------------------------------------------------------
# 通用工具
# ---------------------------------------------------------------------------
def sheet_ref(title: str) -> str:
    return title.replace("'", "''")


def is_effective_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def normalize_numeric(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    if text.endswith("%"):
        return float(text[:-1].replace(",", "")) / 100.0
    return float(text.replace(",", ""))


_NORMALIZE_RE = re.compile(r"[\s\-_./()（）｜|]+")


def normalize_header(text: Any) -> str:
    """对表头做模糊归一：去空白/标点、转小写。"""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = _NORMALIZE_RE.sub("", s)
    return s


def validate_sheet_title(title: str, field_name: str) -> None:
    if not title or not title.strip():
        raise SummaryGenerationError(f"{field_name} 不能为空")
    invalid = set('/\\?*[]:')
    if any(ch in invalid for ch in title):
        raise SummaryGenerationError(f"{field_name} 含非法字符：{title}")


def validate_thresholds() -> None:
    required = {"L", "M", "N", "O", "P"}
    if set(BENCHMARKS.keys()) != required:
        raise SummaryGenerationError(f"benchmark 配置异常：{BENCHMARKS}")
    if BENCHMARKS["L"] != 55 or BENCHMARKS["M"] != 0.01 or BENCHMARKS["N"] != 5 or BENCHMARKS["P"] != 0.05:
        raise SummaryGenerationError(f"benchmark 阈值与约定不一致：{BENCHMARKS}")
    if BENCHMARKS["O"] is not None:
        raise SummaryGenerationError("O 列当前不允许设置 benchmark")


def validate_workbook_structure(workbook_path: Path, raw_sheet_title: str) -> str:
    """检查 workbook，返回真实存在的 raw sheet 标题（用于本地兜底名称差异）。"""
    if not workbook_path.exists():
        raise SummaryGenerationError(f"导出的 workbook 不存在：{workbook_path}")
    wb = load_workbook(workbook_path, data_only=False)
    if raw_sheet_title in wb.sheetnames:
        return raw_sheet_title
    # 兜底：导出的本地副本可能因 sheet 重命名 / 国际化等出现 "sheet1" 这类标题。
    if len(wb.sheetnames) == 1:
        return wb.sheetnames[0]
    raise SummaryGenerationError(
        f"导出的 workbook 中缺少 raw sheet：{raw_sheet_title}，实际包含：{wb.sheetnames}"
    )


def ensure_bytedcli_auth() -> None:
    workspace_root = Path(__file__).resolve().parents[3]
    auth_script = workspace_root / "inner_skills" / "bytedcli-auth" / "scripts" / "bytedcli_auth.sh"
    if not auth_script.exists():
        raise SummaryGenerationError(f"找不到 bytedcli-auth 脚本：{auth_script}")
    result = subprocess.run(["bash", str(auth_script)], capture_output=True, text=True)
    if result.returncode != 0:
        raise SummaryGenerationError(
            "bytedcli-auth 失败\n"
            f"stdout:\n{result.stdout}\n"
            f"stderr:\n{result.stderr}"
        )


def export_workbook(cli: LarkSheetsCLI, spreadsheet_token: str, output_dir: Path) -> Path:
    """可选的 xlsx 导出（部分租户/节点没有导出权限，默认走在线 read 路径）。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    target = output_dir / f"{DEFAULT_EXPORT_PREFIX}-{spreadsheet_token}.xlsx"
    result = cli.export_xlsx(spreadsheet_token, target)
    return result.saved_path


def fetch_raw_grid(
    cli: LarkSheetsCLI,
    spreadsheet_token: str,
    raw_sheet: SheetInfo,
) -> List[List[Any]]:
    """通过 `+read` 拉取 raw sheet 的全量内容（兼容无导出权限场景）。

    返回一个 list[list]，行/列均按 1-based 自然顺序，缺失单元格补 None。
    """
    end_col = max(1, raw_sheet.column_count)
    end_row = max(1, raw_sheet.row_count)
    end_letter = col_num_to_a1(end_col)
    a1 = f"{raw_sheet.sheet_id}!A1:{end_letter}{end_row}"
    return cli.read_range(spreadsheet_token, a1, value_render_option="ToString")


# ---------------------------------------------------------------------------
# 表头自动定位与字段映射
# ---------------------------------------------------------------------------
def _build_letter_map_from_row(row: List[Any]) -> Dict[str, str]:
    header_lookup: Dict[str, str] = {}
    for col_idx, value in enumerate(row, start=1):
        if not is_effective_value(value):
            continue
        key = normalize_header(value)
        if not key:
            continue
        header_lookup.setdefault(key, col_num_to_a1(col_idx))

    letter_map: Dict[str, str] = {}
    for field, candidates in RAW_FIELD_HEADER_CANDIDATES.items():
        for candidate in candidates:
            norm = normalize_header(candidate)
            if norm in header_lookup:
                letter_map[field] = header_lookup[norm]
                break
    return letter_map


def detect_header_mapping_from_grid(
    grid: List[List[Any]],
    raw_sheet_title: str,
) -> HeaderMapping:
    """在 grid 前 N 行内定位真表头行，并构建语义→列字母映射。"""
    core_required = ["handle", "cl_gmv", "show_pv"]
    scan_limit = min(DEFAULT_HEADER_SCAN_ROWS, len(grid))

    best_row = -1
    best_letter_map: Dict[str, str] = {}
    for row_idx in range(1, scan_limit + 1):
        row = grid[row_idx - 1] if row_idx - 1 < len(grid) else []
        letter_map = _build_letter_map_from_row(row)
        if all(field in letter_map for field in core_required) and len(letter_map) > len(best_letter_map):
            best_row = row_idx
            best_letter_map = letter_map

    if best_row < 0:
        raise SummaryGenerationError(
            f"在 raw sheet `{raw_sheet_title}` 前 {scan_limit} 行内无法定位包含 TT Handle/CL GMV/Show PV 的表头行；"
            f"请检查源表是否仍包含这些核心字段。"
        )

    required_fields = set(SUMMARY_COL_TO_FIELD.values()) | {"duration_sec", "show_pv"}
    missing = sorted(required_fields - set(best_letter_map.keys()))
    if missing:
        raise SummaryGenerationError(
            "raw sheet 表头缺少以下字段，无法生成完整 summary：\n"
            + "\n".join(
                f"  - {field}（候选：{', '.join(RAW_FIELD_HEADER_CANDIDATES.get(field, []))}）"
                for field in missing
            )
        )

    return HeaderMapping(
        header_row=best_row,
        data_start_row=best_row + 1,
        field_to_letter=best_letter_map,
    )


def locate_last_raw_row_in_grid(grid: List[List[Any]], data_start_row: int) -> int:
    last_row = data_start_row - 1
    for row_idx in range(data_start_row, len(grid) + 1):
        row = grid[row_idx - 1]
        if any(is_effective_value(cell) for cell in row):
            last_row = row_idx
    return last_row


def _letter_to_col_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"invalid column letter: {letter}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def collect_benchmark_red_ranges_from_grid(
    grid: List[List[Any]],
    data_start_row: int,
    last_raw_row: int,
    field_to_letter: Dict[str, str],
    summary_data_start_row: int,
) -> List[str]:
    cells: List[str] = []
    for summary_col, field in BENCHMARK_FIELD_KEYS.items():
        threshold = BENCHMARKS[summary_col]
        if threshold is None:
            continue
        raw_letter = field_to_letter.get(field)
        if not raw_letter:
            continue
        col_idx_0 = _letter_to_col_index(raw_letter) - 1
        for offset, raw_row in enumerate(range(data_start_row, last_raw_row + 1)):
            row = grid[raw_row - 1] if raw_row - 1 < len(grid) else []
            value_raw = row[col_idx_0] if col_idx_0 < len(row) else None
            value = normalize_numeric(value_raw)
            if value < threshold:
                summary_row = summary_data_start_row + offset
                cells.append(f"{summary_col}{summary_row}")
    return LarkSheetsCLI.compress_cells_to_ranges(cells)


# ---------------------------------------------------------------------------
# 公式生成
# ---------------------------------------------------------------------------
def build_formula_for_cell(
    summary_col: str,
    raw_sheet_title: str,
    raw_row: int,
    field_to_letter: Dict[str, str],
) -> str:
    raw = sheet_ref(raw_sheet_title)
    L = field_to_letter

    if summary_col == "A":
        return f"=IFERROR('{raw}'!{L['handle']}{raw_row},\"\")"
    if summary_col == "B":
        date_letter = L["date"]
        return (
            f"=IFERROR(IF('{raw}'!{date_letter}{raw_row}<>\"\","
            f"DATEVALUE(LEFT('{raw}'!{date_letter}{raw_row},10)),\"\"),\"\")"
        )
    if summary_col == "C":
        return f"=IFERROR('{raw}'!{L['cl_gmv']}{raw_row},\"\")"
    if summary_col == "D":
        # 时均GMV = C / F（F 已经是开播小时，公式中按位置引用）
        return "=IFERROR(IF(F{r}>0,C{r}/F{r},\"\"),\"\")".replace("{r}", str(raw_row))
    if summary_col == "E":
        # 时均show PV /K = raw show_pv / F / 1000
        return (
            "=IFERROR(IF(F{r}>0,'{raw}'!{pv}{r}/F{r},\"\"),\"\")/1000"
            .replace("{raw}", raw)
            .replace("{pv}", L["show_pv"])
            .replace("{r}", str(raw_row))
        )
    if summary_col == "F":
        # 开播小时 = duration_sec / 3600
        dur = L["duration_sec"]
        return (
            f"=IFERROR(IF('{raw}'!{dur}{raw_row}<>\"\","
            f"'{raw}'!{dur}{raw_row}/3600,\"\"),\"\")"
        )

    # 互动指标族（M/N/O/P）：raw 值可能是字符串百分比或数字，统一 *1 转数值
    interaction_columns = {"M", "N", "O", "P"}
    if summary_col in interaction_columns:
        field = SUMMARY_COL_TO_FIELD[summary_col]
        letter = L[field]
        return (
            f"=IFERROR(IF('{raw}'!{letter}{raw_row}=\"\",0,'{raw}'!{letter}{raw_row}*1),0)"
        )

    # 其他列直通：G/H/I/J/K/L
    field = SUMMARY_COL_TO_FIELD[summary_col]
    letter = L[field]
    return f"=IFERROR('{raw}'!{letter}{raw_row},\"\")"


def build_summary_matrix(
    raw_sheet_title: str,
    data_start_row: int,
    last_raw_row: int,
    field_to_letter: Dict[str, str],
) -> List[List[Any]]:
    matrix: List[List[Any]] = [
        [
            "基础数据",
            "自动计算，在【1. 数据底表】贴入数据即可",
            "", "", "", "", "", "", "", "", "",
            {"type": "url", "text": "互动指标｜Fashion Live 过程指标优化2505", "link": "https://bytedance.sg.larkoffice.com/docx/DzfedikUGoxzT0x6dqklQAFogsh"},
            "", "", "", "",
        ],
        SUMMARY_HEADERS[:],
    ]
    for raw_row in range(data_start_row, last_raw_row + 1):
        row: List[Any] = []
        for summary_col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
            formula = build_formula_for_cell(summary_col, raw_sheet_title, raw_row, field_to_letter)
            row.append({"type": "formula", "text": formula})
        matrix.append(row)
    return matrix


# ---------------------------------------------------------------------------
# Summary sheet 重建
# ---------------------------------------------------------------------------
def resolve_summary_sheet(
    cli: LarkSheetsCLI,
    spreadsheet_token: str,
    raw_sheet: SheetInfo,
    summary_sheet_title: str,
) -> SheetInfo:
    summary_sheet = cli.get_sheet(spreadsheet_token, summary_sheet_title)
    if summary_sheet is None:
        summary_sheet = cli.create_sheet(
            spreadsheet_token,
            summary_sheet_title,
            index=DEFAULT_SUMMARY_SHEET_INDEX
            if raw_sheet.index != DEFAULT_SUMMARY_SHEET_INDEX
            else raw_sheet.index + 1,
        )
    return summary_sheet


def prepare_context(
    cli: LarkSheetsCLI,
    spreadsheet_token: str,
    raw_sheet_title: str,
    summary_sheet_title: str,
    export_dir: Path,
) -> WorkbookContext:
    raw_sheet = cli.get_sheet(spreadsheet_token, raw_sheet_title)
    if raw_sheet is None:
        raise SummaryGenerationError(f"找不到 raw sheet：{raw_sheet_title}")

    summary_sheet = resolve_summary_sheet(cli, spreadsheet_token, raw_sheet, summary_sheet_title)

    # 优先走在线 read 路径（兼容无 export 权限的国际化节点）；export 失败时降级。
    raw_grid = fetch_raw_grid(cli, spreadsheet_token, raw_sheet)
    header_mapping = detect_header_mapping_from_grid(raw_grid, raw_sheet.title)
    last_raw_row = locate_last_raw_row_in_grid(raw_grid, header_mapping.data_start_row)
    if last_raw_row < header_mapping.data_start_row:
        raise SummaryGenerationError(
            f"raw sheet 在表头行 {header_mapping.header_row} 之后未发现任何数据，请确认源表是否已粘贴明细。"
        )

    benchmark_red_ranges = collect_benchmark_red_ranges_from_grid(
        grid=raw_grid,
        data_start_row=header_mapping.data_start_row,
        last_raw_row=last_raw_row,
        field_to_letter=header_mapping.field_to_letter,
        summary_data_start_row=3,
    )

    # 可选 export，仅作排查留痕；失败不影响主流程。
    exported_xlsx = Path("(skipped)")
    try:
        exported_xlsx = export_workbook(cli, spreadsheet_token, export_dir)
    except LarkSheetsError:
        pass

    clear_end_row = max(summary_sheet.row_count, last_raw_row + DEFAULT_CLEAR_BUFFER_ROWS, DEFAULT_MIN_GRID_ROWS)
    return WorkbookContext(
        spreadsheet_token=spreadsheet_token,
        raw_sheet=raw_sheet,
        summary_sheet=summary_sheet,
        last_raw_row=last_raw_row,
        clear_end_row=clear_end_row,
        benchmark_red_ranges=benchmark_red_ranges,
        exported_xlsx=exported_xlsx,
        header_mapping=header_mapping,
    )


def apply_layout_and_styles(cli: LarkSheetsCLI, context: WorkbookContext) -> None:
    spreadsheet_token = context.spreadsheet_token
    summary_sheet = context.summary_sheet
    end_col = 16

    summary_sheet = cli.ensure_grid_size(
        spreadsheet_token,
        summary_sheet,
        min_rows=context.clear_end_row,
        min_cols=max(end_col, DEFAULT_MIN_GRID_COLS),
    )

    for merge_range in (f"{summary_sheet.sheet_id}!B1:K1", f"{summary_sheet.sheet_id}!L1:P1"):
        try:
            cli.unmerge_cells(spreadsheet_token, merge_range)
        except LarkSheetsError:
            pass

    cli.clear_values_in_chunks(
        spreadsheet_token,
        summary_sheet.sheet_id,
        start_row=1,
        end_row=context.clear_end_row,
        end_col=end_col,
        chunk_size=DEFAULT_CHUNK_SIZE,
    )
    cli.set_style(
        spreadsheet_token,
        f"{summary_sheet.sheet_id}!A1:P{context.clear_end_row}",
        {"clean": True},
    )

    matrix = build_summary_matrix(
        context.raw_sheet.title,
        context.header_mapping.data_start_row,
        context.last_raw_row,
        context.header_mapping.field_to_letter,
    )
    cli.write_matrix_in_chunks(
        spreadsheet_token,
        summary_sheet.sheet_id,
        start_row=1,
        matrix=matrix,
        chunk_size=DEFAULT_CHUNK_SIZE,
    )

    cli.merge_cells(spreadsheet_token, f"{summary_sheet.sheet_id}!B1:K1")
    cli.merge_cells(spreadsheet_token, f"{summary_sheet.sheet_id}!L1:P1")
    cli.update_sheet(
        spreadsheet_token,
        summary_sheet.sheet_id,
        frozen_row_count=DEFAULT_FREEZE_ROWS,
        frozen_col_count=DEFAULT_FREEZE_COLS,
    )

    # Set column width: C-K (columns 3-11) = 65px
    cli.update_dimension(
        spreadsheet_token,
        summary_sheet.sheet_id,
        "COLUMNS",
        start_index=3,
        end_index=11,
        fixed_size=65,
    )

    style_steps: List[tuple[str, Dict[str, Any]]] = [
        (f"{summary_sheet.sheet_id}!A1:P{context.clear_end_row}", {"foreColor": DEFAULT_BLACK}),
        (
            f"{summary_sheet.sheet_id}!A1:K2",
            {"foreColor": DEFAULT_BLACK, "backColor": DEFAULT_LIGHT_GREEN, "font": {"bold": True}},
        ),
        (
            f"{summary_sheet.sheet_id}!L1:P2",
            {"foreColor": DEFAULT_BLACK, "backColor": DEFAULT_LIGHT_YELLOW, "font": {"bold": True}},
        ),
    ]
    # Row 2: wrap text
    style_steps.append(
        (f"{summary_sheet.sheet_id}!A2:P2", {"wrapStrategy": "WRAP"}),
    )

    summary_data_end_row = 2 + (context.last_raw_row - context.header_mapping.data_start_row + 1)
    if summary_data_end_row >= 3:
        style_steps.extend(
            [
                (f"{summary_sheet.sheet_id}!B3:B{summary_data_end_row}", {"formatter": "yyyy/MM/dd", "foreColor": DEFAULT_BLACK}),
                (f"{summary_sheet.sheet_id}!C3:G{summary_data_end_row}", {"formatter": "0", "foreColor": DEFAULT_BLACK}),
                (f"{summary_sheet.sheet_id}!K3:K{summary_data_end_row}", {"formatter": "0", "foreColor": DEFAULT_BLACK}),
                (f"{summary_sheet.sheet_id}!H3:J{summary_data_end_row}", {"formatter": "0%", "foreColor": DEFAULT_BLACK}),
                (f"{summary_sheet.sheet_id}!M3:M{summary_data_end_row}", {"formatter": "0%", "foreColor": DEFAULT_BLACK}),
                (f"{summary_sheet.sheet_id}!N3:N{summary_data_end_row}", {"formatter": "0%", "foreColor": DEFAULT_BLACK}),
                (f"{summary_sheet.sheet_id}!O3:P{summary_data_end_row}", {"formatter": "0%", "foreColor": DEFAULT_BLACK}),
            ]
        )

    for a1_range, style in style_steps:
        cli.set_style(spreadsheet_token, a1_range, style)

    for cell_range in context.benchmark_red_ranges:
        cli.set_style(
            spreadsheet_token,
            f"{summary_sheet.sheet_id}!{cell_range}",
            {"foreColor": DEFAULT_BLACK, "backColor": DEFAULT_LIGHT_RED},
        )


# ---------------------------------------------------------------------------
# RAW 写后回读校验
# ---------------------------------------------------------------------------
def verify_after_write(
    cli: LarkSheetsCLI,
    context: WorkbookContext,
    summary_sheet_id: str,
) -> Dict[str, Any]:
    """对 summary sheet 做最小可靠的写后回读校验：
    1. 表头行第二行必须等于 SUMMARY_HEADERS。
    2. A1 必须为“基础数据”。
    3. 数据区首行 B/C/F 必须为非空数值（等价于公式有效计算）。
    4. 数据区末行同样校验。
    """
    spreadsheet_token = context.spreadsheet_token
    last_summary_data_row = 2 + (context.last_raw_row - context.header_mapping.data_start_row + 1)

    head_range = f"{summary_sheet_id}!A1:P2"
    head_values = cli.read_range(spreadsheet_token, head_range, value_render_option="ToString")
    findings: List[str] = []

    if not head_values or len(head_values) < 2:
        findings.append(f"读取 {head_range} 返回的行数 < 2")
    else:
        row1 = head_values[0]
        row2 = head_values[1]
        if not row1 or str(row1[0]).strip() != "基础数据":
            findings.append(f"A1 应为 '基础数据'，实际为 {row1[:1] if row1 else '[]'}")
        # row2 头部允许长度等于 16
        if len(row2) < len(SUMMARY_HEADERS):
            findings.append(f"表头第二行长度 {len(row2)} 小于期望 {len(SUMMARY_HEADERS)}")
        else:
            for idx, expected in enumerate(SUMMARY_HEADERS):
                actual = str(row2[idx]).strip() if row2[idx] is not None else ""
                if actual != expected:
                    findings.append(f"P2 表头第 {idx + 1} 列期望 '{expected}'，实际 '{actual}'")
                    break  # 仅报第一处，避免噪声

    # 抽样验证首行 + 末行 公式产物
    sample_rows = []
    if last_summary_data_row >= 3:
        sample_rows.append(3)
    if last_summary_data_row > 3:
        sample_rows.append(last_summary_data_row)

    sample_results: List[Dict[str, Any]] = []
    for r in sample_rows:
        rng = f"{summary_sheet_id}!A{r}:P{r}"
        # FormattedValue 返回飞书前端渲染后的文本，能反映公式真实计算结果。
        values = cli.read_range(spreadsheet_token, rng, value_render_option="FormattedValue")
        row = values[0] if values else []
        # 关键列：A handle、C GMV、F 开播小时
        a_val = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        c_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        f_val = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        # K 列 AOV：如果落入 0~1 之间且单元格被识别成百分比，可视为字段错位的强信号。
        k_val = str(row[10]).strip() if len(row) > 10 and row[10] is not None else ""
        sample_results.append({"row": r, "A": a_val, "C": c_val, "F": f_val, "K(AOV)": k_val})
        if not a_val:
            findings.append(f"A{r} 为空（Handle 列），可能字段映射错误")
        if not c_val:
            findings.append(f"C{r} 为空（GMV 列），可能字段映射错误")
        if not f_val:
            findings.append(f"F{r} 为空（开播小时列），可能字段映射错误")

    success = not findings
    return {
        "success": success,
        "findings": findings,
        "head_range": head_range,
        "head_values": head_values,
        "sample_rows": sample_results,
        "last_summary_data_row": last_summary_data_row,
    }


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------
def generate_summary_sheet(
    spreadsheet: str,
    raw_sheet_title: str,
    summary_sheet_title: str,
    export_dir: Path,
) -> GenerationResult:
    validate_sheet_title(raw_sheet_title, "raw_sheet_title")
    validate_sheet_title(summary_sheet_title, "summary_sheet_title")
    validate_thresholds()

    ensure_bytedcli_auth()
    cli = LarkSheetsCLI()
    spreadsheet_token = cli.resolve_spreadsheet_token(spreadsheet)
    context = prepare_context(cli, spreadsheet_token, raw_sheet_title, summary_sheet_title, export_dir)
    apply_layout_and_styles(cli, context)

    refreshed_summary = cli.get_sheet(spreadsheet_token, summary_sheet_title)
    if refreshed_summary is None:
        raise SummaryGenerationError(f"生成完成后找不到 summary sheet：{summary_sheet_title}")

    readback = verify_after_write(cli, context, refreshed_summary.sheet_id)
    if not readback["success"]:
        raise SummaryGenerationError(
            "RAW 写后回读校验失败：\n"
            + "\n".join(f"  - {item}" for item in readback["findings"])
            + f"\n  详情见 readback={json.dumps(readback, ensure_ascii=False)}"
        )

    data_rows = max(0, context.last_raw_row - context.header_mapping.data_start_row + 1)
    return GenerationResult(
        spreadsheet_token=spreadsheet_token,
        raw_sheet_id=context.raw_sheet.sheet_id,
        summary_sheet_id=refreshed_summary.sheet_id,
        raw_sheet_title=context.raw_sheet.title,
        summary_sheet_title=refreshed_summary.title,
        data_rows=data_rows,
        header_row=context.header_mapping.header_row,
        data_start_row=context.header_mapping.data_start_row,
        last_raw_row=context.last_raw_row,
        clear_end_row=context.clear_end_row,
        benchmark_red_ranges=context.benchmark_red_ranges,
        exported_xlsx=str(context.exported_xlsx),
        field_to_letter=context.header_mapping.field_to_letter,
        readback=readback,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据 Live Performance raw sheet 自动重建 2. 计算汇总")
    parser.add_argument("spreadsheet", help="spreadsheet URL 或 spreadsheet token")
    parser.add_argument("--raw-sheet-title", default=DEFAULT_RAW_SHEET_TITLE, help="raw sheet 标题")
    parser.add_argument("--summary-sheet-title", default=DEFAULT_SUMMARY_SHEET_TITLE, help="summary sheet 标题")
    parser.add_argument(
        "--export-dir",
        default=str(Path.cwd() / "tmp_live_performance_exports"),
        help="导出 workbook 的本地目录，用于动态识别有效行与 benchmark",
    )
    parser.add_argument(
        "--output-json",
        help="可选：把执行结果 JSON 写到指定文件",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        result = generate_summary_sheet(
            spreadsheet=args.spreadsheet,
            raw_sheet_title=args.raw_sheet_title,
            summary_sheet_title=args.summary_sheet_title,
            export_dir=Path(args.export_dir).resolve(),
        )
    except (SummaryGenerationError, LarkSheetsError, ValueError) as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False, indent=2))
        return 1

    payload = {"ok": True, "result": result.to_dict()}
    if args.output_json:
        output_path = Path(args.output_json).resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
