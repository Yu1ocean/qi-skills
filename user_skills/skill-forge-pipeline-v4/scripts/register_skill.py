#!/usr/bin/env python3
import argparse
import datetime
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
import zipfile
from pathlib import Path
from typing import Any, Callable, Dict, Optional, Tuple
from urllib.parse import quote, urlparse

import requests

DEFAULT_EMAIL = "yuqinan@bytedance.com"
DEFAULT_OPEN_API_BASE = "https://fsopen.bytedance.net"
DEFAULT_UPLOAD_API_BASE = "https://open.feishu.cn"
DEFAULT_SKILL_INVENTORY_URL = "https://bytedance.sg.larkoffice.com/sheets/ECQ0sDwmbhDex9tcUSjlkU7Bgdh"
DEFAULT_SKILL_INVENTORY_SHEET_NAME = "专属技能清单"
DEFAULT_SKILL_INVENTORY_UPDATED_AT_FORMATTER = "yyyy/MM/dd"

# --- SSOT (Single Source of Truth) Version Sync Bus ---
# We treat `version:` in the target skill's SKILL.md frontmatter as SSOT.
# During Archive, we bump version (Major: +1.0 / Minor: +0.1) and sync it to:
# 1) local SKILL.md
# 2) Feishu skill inventory sheet (full-row upsert via omni-asset-archiver)
# 3) (optional) Feishu skill doc version marker
#
# 【Initial Version Policy】
# Per user (奇楠) directive (2026-05-20):
# Brand-new skills MUST debut at version 1.1, not 0.x.
# Scaffolding tools (e.g. aime-skill-creator) typically seed SKILL.md with
# `version: 0.x` placeholders. When register_skill.py detects this
# "fresh skill" state during its first run, it will force-set the
# published version to DEFAULT_INITIAL_VERSION (1.1) instead of
# performing the usual +0.1 bump within the 0.x range.
DEFAULT_INITIAL_VERSION = "1.1"
DEFAULT_SSOT_SPREADSHEET_TOKEN = "ECQ0sDwmbhDex9tcUSjlkU7Bgdh"
DEFAULT_SSOT_SHEET_NAME = "专属技能清单"
DEFAULT_SSOT_SHEET_NAME_FALLBACKS = ["专属技能清单_Sheet"]
DEFAULT_WIKI_NODE_TOKEN = "GU0ewkyaGi4i5nkwBtNcM3aPn9g"
DEFAULT_WIKI_URL = f"https://bytedance.larkoffice.com/wiki/{DEFAULT_WIKI_NODE_TOKEN}"

DEFAULT_SSOT_VERSION_HEADERS = ["版本号", "版本", "version", "Version"]
DEFAULT_SSOT_ID_HEADERS = ["技能编号", "Skill ID", "skill_id", "SkillID", "ID", "编号", "包 ID"]
DEFAULT_SSOT_NAME_HEADERS = ["技能名称", "技能名", "name", "Name"]

REQUIRED_DOC_TEMPLATE_MARKERS = [
    "🔑 触发词",
    "📖 案例实录",
]


class GuardrailViolation(RuntimeError):
    """Hard-stop exception for guardrail violations."""


def validate_doc_template_markers(doc_url: str) -> None:
    """L1/L2 doc-structure validator for skill docs.

    We validate markers AFTER downloading the doc via Lark MCP.
    """

    markdown_path = download_doc_markdown(doc_url)
    content = markdown_path.read_text(encoding="utf-8", errors="ignore")

    missing = [m for m in REQUIRED_DOC_TEMPLATE_MARKERS if m not in content]
    if missing:
        raise GuardrailViolation(
            "Skill doc template markers missing: " + ", ".join(missing)
        )


def run_cda_guardrails_selfcheck(skill_dir: Path, risk: str = "auto") -> str:
    """CDA Guardrails checkpoint: fail-fast before side effects."""

    selfcheck_script = Path(__file__).resolve().parent / "cda_guardrails_selfcheck.py"
    if not selfcheck_script.exists():
        raise FileNotFoundError(f"CDA selfcheck script not found: {selfcheck_script}")

    return run_subprocess(
        [
            "python3",
            str(selfcheck_script),
            "--skill-dir",
            str(skill_dir),
            "--risk",
            risk,
        ],
        "CDA-Guardrails-Selfcheck",
    )


def run_post_forge_git_push(workspace_root: Path, skill_name: str, skill_version: str) -> str:
    """Run the qi-skills post-forge git sync hook after a successful archive."""

    if os.environ.get("SKIP_POST_FORGE_GIT_PUSH") == "1":
        return "⏭️ post-forge git push skipped by SKIP_POST_FORGE_GIT_PUSH=1"

    hook_path = workspace_root / "user_skills" / "scripts" / "post_forge_git_push.sh"
    if not hook_path.exists():
        raise FileNotFoundError(f"Post-forge git push hook not found: {hook_path}")

    return run_subprocess(
        [
            "bash",
            str(hook_path),
            skill_name or "unknown-skill",
            skill_version or "latest",
        ],
        "post-forge git push hook",
    )


def get_workspace_root() -> Path:
    env_path = os.environ.get("IRIS_WORKSPACE_PATH")
    if env_path:
        return Path(env_path).resolve()
    return Path(__file__).resolve().parents[3]


FEISHU_DATE_EPOCH = datetime.datetime(1899, 12, 30)
FEISHU_DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%d",
    "%Y/%m/%d",
)


def maybe_to_feishu_datetime_serial(value: Any) -> Any:
    if isinstance(value, (int, float)):
        return value
    text = str(value or "").strip()
    if not text:
        return value
    for fmt in FEISHU_DATETIME_FORMATS:
        try:
            parsed = datetime.datetime.strptime(text, fmt)
            return round((parsed - FEISHU_DATE_EPOCH).total_seconds() / 86400, 10)
        except ValueError:
            continue
    return value


def get_raw_cloud_jwt() -> str:
    raw = (os.environ.get("AIME_USER_CLOUD_JWT") or "").strip()
    if not raw:
        raise RuntimeError(
            "Missing env var AIME_USER_CLOUD_JWT. Run this script with include_secrets=true."
        )
    return raw


def ensure_bearer_token() -> str:
    raw = get_raw_cloud_jwt()
    if raw.startswith("Bearer "):
        return raw
    return f"Bearer {raw}"


def build_headers(bearer_token: str, content_type: Optional[str] = None) -> Dict[str, str]:
    headers = {"Authorization": bearer_token}
    if content_type:
        headers["Content-Type"] = content_type
    if bearer_token.startswith("Bearer "):
        headers["Cookie"] = f"session={bearer_token.split(' ', 1)[1]}"
    return headers


def call_with_retry(action: str, callback: Callable[[], Any], attempts: int = 3, delay: float = 2.0) -> Any:
    last_error: Optional[Exception] = None
    for attempt in range(1, attempts + 1):
        try:
            return callback()
        except Exception as exc:  # noqa: PERF203
            last_error = exc
            if attempt == attempts:
                break
            print(f"⚠️ {action} failed on attempt {attempt}/{attempts}: {exc}")
            time.sleep(delay)
    raise RuntimeError(f"{action} failed after {attempts} attempts: {last_error}")


def create_skill_zip(skill_dir: Path, output_zip: Path) -> Path:
    if not skill_dir.is_dir():
        raise FileNotFoundError(f"Skill directory not found: {skill_dir}")

    skip_names = {".git", "__pycache__", ".DS_Store"}
    output_zip.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(output_zip, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(skill_dir.rglob("*")):
            if any(part in skip_names for part in path.parts):
                continue
            if path == output_zip:
                continue
            arcname = Path(skill_dir.name) / path.relative_to(skill_dir)
            archive.write(path, arcname)

    return output_zip.resolve()


def get_root_folder_token(bearer_token: str, open_api_base: str) -> str:
    url = f"{open_api_base}/open-apis/drive/explorer/v2/root_folder/meta"
    response = requests.get(url, headers=build_headers(bearer_token), timeout=30)
    data = response.json()
    if response.status_code != 200 or data.get("code") != 0:
        raise RuntimeError(f"get root folder failed: http={response.status_code}, resp={data}")
    token = data.get("data", {}).get("token")
    if not token:
        raise RuntimeError(f"root folder token missing in response: {data}")
    return token


def upload_zip_to_drive(
    bearer_token: str,
    open_api_base: str,
    file_path: Path,
    parent_node: str = "",
) -> str:
    url = f"{open_api_base}/open-apis/drive/v1/files/upload_all"
    with file_path.open("rb") as file_obj:
        response = requests.post(
            url,
            headers=build_headers(bearer_token),
            data={
                "file_name": file_path.name,
                "parent_type": "explorer",
                "parent_node": parent_node,
                "size": str(file_path.stat().st_size),
            },
            files={"file": (file_path.name, file_obj)},
            timeout=300,
        )
    data = response.json()
    if response.status_code != 200 or data.get("code") != 0:
        raise RuntimeError(f"drive upload failed: http={response.status_code}, resp={data}")
    file_token = data.get("data", {}).get("file_token")
    if not file_token:
        raise RuntimeError(f"file_token missing after drive upload: {data}")
    return file_token


def add_permission_member(
    bearer_token: str,
    open_api_base: str,
    file_token: str,
    email: str,
    perm: str = "full_access",
) -> Dict[str, Any]:
    url = f"{open_api_base}/open-apis/drive/v1/permissions/{file_token}/members"
    response = requests.post(
        url,
        params={"type": "file", "need_notification": "false"},
        headers=build_headers(bearer_token, "application/json; charset=utf-8"),
        json={
            "member_type": "email",
            "member_id": email,
            "perm": perm,
            "type": "user",
        },
        timeout=30,
    )
    return {"http_status": response.status_code, "data": response.json()}


def update_permission_member(
    bearer_token: str,
    open_api_base: str,
    file_token: str,
    email: str,
    perm: str = "full_access",
) -> Dict[str, Any]:
    member_id = quote(email, safe="")
    url = f"{open_api_base}/open-apis/drive/v1/permissions/{file_token}/members/{member_id}"
    response = requests.put(
        url,
        params={"type": "file", "need_notification": "false"},
        headers=build_headers(bearer_token, "application/json; charset=utf-8"),
        json={"member_type": "email", "perm": perm, "type": "user"},
        timeout=30,
    )
    return {"http_status": response.status_code, "data": response.json()}


def ensure_drive_permission(
    bearer_token: str,
    open_api_base: str,
    file_token: str,
    email: str,
    perm: str = "full_access",
) -> None:
    """Legacy compatibility shim for older callers.

    Deprecated: do not use JWT/OpenAPI permission grants. Prefer
    `ensure_drive_asset_access_via_mcp()` with a concrete /file/ URL.
    """

    del bearer_token, open_api_base
    drive_file_url = build_drive_file_url("https://bytedance.larkoffice.com/docx/placeholder", file_token)
    ensure_drive_asset_access_via_mcp(drive_file_url, email=email, perm=perm)


def build_drive_file_url(doc_url: str, file_token: str) -> str:
    parsed = urlparse(doc_url)
    scheme = parsed.scheme or "https"
    netloc = parsed.netloc or "bytedance.larkoffice.com"
    return f"{scheme}://{netloc}/file/{file_token}"


def ensure_drive_asset_access_via_mcp(drive_file_url: str, email: str, perm: str = "full_access") -> str:
    workspace_root = get_workspace_root()
    repair_script = workspace_root / "user_skills/feishu-doc-writing-guide/scripts/grant_doc_permissions.py"
    if not repair_script.exists():
        raise FileNotFoundError(f"grant_doc_permissions wrapper not found: {repair_script}")

    try:
        return run_subprocess(
            [
                "python3",
                str(repair_script),
                drive_file_url,
                "--email",
                email,
                "--perm",
                perm,
            ],
            "repair drive asset access via MCP personal flow",
        )
    except RuntimeError as exc:
        if "move_lark_doc" not in str(exc):
            raise
        return (
            "⚠️ drive asset access repair skipped: legacy move_lark_doc path is unavailable; "
            "file block was inserted via lark-cli docs +media-insert as user identity."
        )


def run_subprocess(command: list[str], action: str) -> str:
    result = subprocess.run(command, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(
            f"{action} failed with exit code {result.returncode}\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )
    return result.stdout.strip()


# -----------------------------
# SSOT: Version Sync Bus
# -----------------------------

def ensure_bytedcli_auth() -> None:
    """Authenticate via bytedcli (required before MCP-based Feishu ops).

    NOTE: this expects the script to be executed with `include_secrets=true`
    so that IRIS_USER_CLOUD_JWT is available.
    """

    workspace_root = get_workspace_root()
    auth_script = workspace_root / "inner_skills/bytedcli-auth/scripts/bytedcli_auth.sh"
    if not auth_script.exists():
        raise FileNotFoundError(f"bytedcli auth script not found: {auth_script}")

    print("🔐 Ensuring bytedcli login (bytedcli-auth)...")
    run_subprocess(["bash", str(auth_script)], "bytedcli-auth")


def _normalize_version_to_int_pair(raw: str) -> Tuple[int, int]:
    raw = (raw or "").strip()
    raw = raw.lstrip("vV")

    m = re.match(r"^(\d+)\.(\d+)(?:\.(\d+))?$", raw)
    if not m:
        raise ValueError(f"Unsupported version format: {raw!r} (expected X.Y or X.Y.Z)")
    return int(m.group(1)), int(m.group(2))


def _format_version_pair(major: int, minor: int) -> str:
    return f"{major}.{minor}"


def read_skill_version_from_skill_md(skill_dir: Path) -> str:
    """Read SSOT version from target skill's SKILL.md frontmatter."""

    skill_md = skill_dir / "SKILL.md"
    if not skill_md.exists():
        raise FileNotFoundError(f"SKILL.md not found: {skill_md}")

    content = skill_md.read_text(encoding="utf-8", errors="ignore")
    m = re.search(r"(?m)^version:\s*([^\n]+)\s*$", content)
    if not m:
        raise ValueError(f"version field not found in SKILL.md: {skill_md}")

    major, minor = _normalize_version_to_int_pair(m.group(1).strip())
    return _format_version_pair(major, minor)


def bump_version(current_version: str, bump_type: str) -> str:
    major, minor = _normalize_version_to_int_pair(current_version)
    bump_type = (bump_type or "").strip().lower()

    if bump_type == "major":
        return _format_version_pair(major + 1, 0)
    if bump_type == "minor":
        return _format_version_pair(major, minor + 1)

    raise ValueError(f"Unsupported bump type: {bump_type!r} (expected major/minor)")


def is_initial_version(current_version: str) -> bool:
    """Return True if the current SKILL.md version is still in the
    0.x scaffold range, meaning this is the very first publish of the skill.
    """

    try:
        major, _minor = _normalize_version_to_int_pair(current_version)
    except ValueError:
        return False
    return major == 0


def write_skill_md_version(skill_dir: Path, new_version: str) -> None:
    """Sync point #1: overwrite SSOT version in local SKILL.md."""

    skill_md = skill_dir / "SKILL.md"
    content = skill_md.read_text(encoding="utf-8", errors="ignore")

    # only replace the frontmatter version key; keep other content intact.
    replaced, n = re.subn(
        r"(?m)^version:\s*([^\n]+)\s*$",
        f"version: {new_version}",
        content,
        count=1,
    )
    if n != 1:
        raise RuntimeError(f"Unable to update version field in SKILL.md: {skill_md}")

    skill_md.write_text(replaced, encoding="utf-8")


def build_ssot_spreadsheet_url(spreadsheet_token: str) -> str:
    token = (spreadsheet_token or "").strip()
    if token.startswith("http://") or token.startswith("https://"):
        return token
    return f"https://bytedance.larkoffice.com/sheets/{token}"


def mcp_lark_download(document_url: str) -> list[Path]:
    workspace_root = get_workspace_root()
    download_script = workspace_root / "inner_skills/lark_download/lark_download.py"
    if not download_script.exists():
        raise FileNotFoundError(f"Lark download script not found: {download_script}")

    try:
        output = run_subprocess(
            [
                "python3",
                str(download_script),
                json.dumps({"document_url": document_url}, ensure_ascii=False),
            ],
            "lark download",
        )
    except RuntimeError as exc:
        if "toolset lark_download not found" not in str(exc):
            raise
        token = document_url.rstrip("/").split("/")[-1].split("?")[0]
        fallback_path = workspace_root / "user_skills" / "skill-forge-pipeline-v4" / f"{token}.lark.md"
        fetch_output = run_subprocess(
            [
                "lark-cli",
                "docs",
                "+fetch",
                "--api-version",
                "v2",
                "--as",
                "user",
                "--doc",
                document_url,
                "--doc-format",
                "markdown",
                "--format",
                "json",
            ],
            "lark-cli docs fetch fallback",
        )
        try:
            payload = json.loads(fetch_output[fetch_output.find("{"):])
            content = payload["data"]["document"]["content"]
        except (json.JSONDecodeError, KeyError, TypeError) as parse_exc:
            raise RuntimeError(f"Unable to parse lark-cli docs +fetch output: {fetch_output}") from parse_exc
        fallback_path.write_text(content, encoding="utf-8")
        return [fallback_path.resolve()]

    paths = [Path(p).resolve() for p in re.findall(r'file_path:\s*"([^"]+)"', output)]
    if not paths:
        paths = [Path(p).resolve() for p in re.findall(r'(/[^\s\"]+\.(?:lark\.md|xlsx|xls|md))', output)]

    if not paths:
        try:
            parsed = json.loads(output)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, list):
            paths = [Path(str(p)).resolve() for p in parsed if str(p).strip()]
        elif isinstance(parsed, str) and parsed.strip():
            paths = [Path(parsed).resolve()]

    if not paths:
        raise RuntimeError(f"Unable to parse downloaded file paths from output: {output}")

    return paths


def _pick_first_xlsx(paths: list[Path]) -> Path:
    for p in paths:
        if p.suffix.lower() in {".xlsx", ".xls"}:
            return p
    raise RuntimeError(f"No xlsx file found in downloaded paths: {paths}")


def _normalize_header_value(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_skill_name_cell(v: Any) -> str:
    """Normalize skill name cell.

    In the inventory sheet, "技能名称" is often a HYPERLINK formula like:
    =HYPERLINK("url", "name")
    We extract the visible name part for matching.
    """

    raw = _normalize_header_value(v)
    if not raw:
        return ""

    if raw.lstrip().upper().startswith("=HYPERLINK"):
        parts = re.findall(r'"([^"]+)"', raw)
        if len(parts) >= 2:
            return parts[-1].strip()
    return raw


def _find_header_col_index(headers: list[str], candidates: list[str]) -> Optional[int]:
    lowered = [h.lower() for h in headers]
    for c in candidates:
        if c.lower() in lowered:
            return lowered.index(c.lower())
    return None


def _choose_sheet_name(workbook_sheetnames: list[str], preferred: str) -> str:
    if preferred in workbook_sheetnames:
        return preferred
    for alt in DEFAULT_SSOT_SHEET_NAME_FALLBACKS:
        if alt in workbook_sheetnames:
            return alt
    raise RuntimeError(
        "SSOT sheet not found. "
        f"preferred={preferred!r}, available={workbook_sheetnames!r}, fallbacks={DEFAULT_SSOT_SHEET_NAME_FALLBACKS!r}"
    )


def update_xlsx_version_cell(
    *,
    source_xlsx: Path,
    sheet_name: str,
    skill_id: str,
    skill_name: str,
    new_version: str,
) -> Tuple[Path, str]:
    """Update the Version cell in a local xlsx copy.

    Returns: (updated_xlsx_path, debug_location)
    """

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: PERF203
        raise RuntimeError(
            "Missing dependency openpyxl. Please install it (pip install openpyxl) before running this script."
        ) from exc

    wb = load_workbook(source_xlsx)
    real_sheet_name = _choose_sheet_name(wb.sheetnames, sheet_name)
    ws = wb[real_sheet_name]

    # assume the first row is header
    header_row = [
        _normalize_header_value(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)
    ]

    id_col = _find_header_col_index(header_row, DEFAULT_SSOT_ID_HEADERS)
    name_col = _find_header_col_index(header_row, DEFAULT_SSOT_NAME_HEADERS)
    version_col = _find_header_col_index(header_row, DEFAULT_SSOT_VERSION_HEADERS)

    if version_col is None:
        raise RuntimeError(
            f"Unable to locate version column by headers {DEFAULT_SSOT_VERSION_HEADERS!r}. "
            f"sheet={real_sheet_name!r}, header_row={header_row!r}"
        )

    # openpyxl column index is 1-based
    version_col_1b = version_col + 1
    id_col_1b = id_col + 1 if id_col is not None else None
    name_col_1b = name_col + 1 if name_col is not None else None

    matched_row: Optional[int] = None
    matched_by = ""
    for r in range(2, ws.max_row + 1):
        rid = _normalize_header_value(ws.cell(row=r, column=id_col_1b).value) if id_col_1b else ""
        rname = _normalize_skill_name_cell(ws.cell(row=r, column=name_col_1b).value) if name_col_1b else ""

        if skill_id and rid and rid.strip() == skill_id.strip():
            matched_row = r
            matched_by = "skill_id"
            break
        if skill_name and rname and rname.strip() == skill_name.strip():
            matched_row = r
            matched_by = "skill_name"
            break

    if not matched_row:
        raise RuntimeError(
            "Unable to locate target skill row in SSOT sheet. "
            f"skill_id={skill_id!r}, skill_name={skill_name!r}, sheet={real_sheet_name!r}"
        )

    ws.cell(row=matched_row, column=version_col_1b).value = new_version

    updated_path = source_xlsx.parent / f"ssot_updated_{uuid.uuid4().hex}.xlsx"
    wb.save(updated_path)

    debug_loc = f"{real_sheet_name}!R{matched_row}C{version_col_1b} (matched_by={matched_by})"
    return updated_path, debug_loc


def sync_skill_inventory_via_omni(metadata: Dict[str, Any]) -> Dict[str, Any]:
    """Sync point #2: upsert the full skill row into Feishu inventory via omni-asset-archiver."""

    workspace_root = get_workspace_root()
    archiver_script = workspace_root / "user_skills/omni-asset-archiver/scripts/archiver_driver.py"
    if not archiver_script.exists():
        raise FileNotFoundError(f"Omni asset archiver script not found: {archiver_script}")

    payload = {
        "asset_type": "skill_inventory",
        "target_route_key": "skill_inventory",
        "skill_id": metadata.get("skill_id", ""),
        "title": metadata.get("name", ""),
        "doc_url": metadata.get("doc_link", ""),
        "description": metadata.get("description", ""),
        "version": metadata.get("version", ""),
        "updated_at": maybe_to_feishu_datetime_serial(metadata.get("updated_at", "")),
    }

    required_fields = ["skill_id", "title", "doc_url", "description", "version", "updated_at"]
    missing_fields = [field for field in required_fields if not str(payload.get(field, "")).strip()]
    if missing_fields:
        raise RuntimeError(
            "Omni skill inventory sync payload missing required fields: " + ", ".join(missing_fields)
        )

    print("🚌 Syncing skill inventory row via omni-asset-archiver...")
    output = run_subprocess(
        [
            "python3",
            str(archiver_script),
            "--payload-json",
            json.dumps(payload, ensure_ascii=False),
        ],
        "omni-asset-archiver skill inventory upsert",
    )

    json_match = re.search(r"(\{[\s\S]*\})\s*$", output.strip())
    if not json_match:
        raise RuntimeError(f"Unable to parse omni-asset-archiver output: {output}")

    result = json.loads(json_match.group(1))
    status = str(result.get("status", "")).strip().lower()
    if status not in {"appended", "updated", "dry_run"}:
        raise RuntimeError(f"Unexpected omni-asset-archiver status: {result}")

    print(
        "✅ Skill inventory sync finished via omni-asset-archiver: "
        f"status={result.get('status')}, sheet={result.get('sheet_name')}, row={result.get('row_number')}"
    )
    return result


def _run_lark_sheets_json(args: list[str], step_name: str) -> dict:
    workspace_root = get_workspace_root()
    cli_path = workspace_root / "inner_skills/lark-sheets/bin/lark-sheets-cli"
    if not cli_path.exists():
        raise FileNotFoundError(f"lark-sheets cli not found: {cli_path}")
    output = run_subprocess([str(cli_path), *args], step_name)
    json_match = re.search(r"(\{[\s\S]*\})", output)
    if not json_match:
        raise RuntimeError(f"Unable to parse lark-sheets output for {step_name}: {output}")
    try:
        return json.loads(json_match.group(1))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Unable to decode lark-sheets json for {step_name}: {output}") from exc


def _resolve_skill_inventory_sheet_id() -> str:
    info = _run_lark_sheets_json(
        ["sheets", "+info", "--url", DEFAULT_SKILL_INVENTORY_URL],
        "resolve skill inventory sheet id",
    )
    sheets = info.get("data", {}).get("sheets", {}).get("sheets", [])
    for sheet in sheets:
        if sheet.get("title") == DEFAULT_SKILL_INVENTORY_SHEET_NAME:
            return str(sheet.get("sheet_id") or "").strip()
    raise RuntimeError(f"Unable to resolve sheet id for {DEFAULT_SKILL_INVENTORY_SHEET_NAME}")


def ensure_skill_inventory_updated_at_formatter(row_number: int, expected_text: str) -> None:
    if row_number <= 0:
        raise ValueError(f"Invalid row_number: {row_number}")
    expected_display = str(expected_text or "").strip()[:10].replace("-", "/")
    if not expected_display:
        raise ValueError("expected_text is empty for skill inventory updated_at formatter verification")
    sheet_id = _resolve_skill_inventory_sheet_id()
    range_str = f"{sheet_id}!E{row_number}:E{row_number}"
    _run_lark_sheets_json(
        [
            "sheets",
            "+set-style",
            "--url",
            DEFAULT_SKILL_INVENTORY_URL,
            "--range",
            range_str,
            "--style",
            json.dumps({"formatter": DEFAULT_SKILL_INVENTORY_UPDATED_AT_FORMATTER}, ensure_ascii=False),
        ],
        "set updated_at formatter on skill inventory row",
    )
    time.sleep(2)
    readback = _run_lark_sheets_json(
        [
            "sheets",
            "+read",
            "--url",
            DEFAULT_SKILL_INVENTORY_URL,
            "--sheet-id",
            sheet_id,
            "--range",
            f"E{row_number}:E{row_number}",
            "--value-render-option",
            "FormattedValue",
        ],
        "verify updated_at formatter on skill inventory row",
    )
    values = readback.get("data", {}).get("valueRange", {}).get("values", [])
    actual = ""
    if values and values[0]:
        actual = str(values[0][0] or "").strip()
    if actual != expected_display:
        raise RuntimeError(
            "Skill inventory updated_at formatter verification failed: "
            f"expected={expected_display!r}, actual={actual!r}, row={row_number}"
        )


def _parse_block_id_from_attach_output(attach_output: str) -> str:
    """Extract the freshly created file block id from `docs +media-insert` output."""

    for match in re.finditer(r'"block_id"\s*:\s*"([^"]+)"', attach_output or ""):
        return match.group(1)
    return ""


def move_block_to_doc_begin(doc_url: str, block_id: str) -> None:
    """Relocate a top-level block to BLOCK_BEGIN (index 0, i.e. right below the title).

    `lark-cli docs +media-insert` appends to the document end by default, which
    silently breaks the pipeline contract "ZIP 文件块必须挂在标题正下方".
    Anchoring `block_move_after` on the document root token moves the block to
    index 0. Failure raises: no silent drift allowed.
    """

    document_id = parse_doc_token(doc_url)
    run_subprocess(
        [
            "lark-cli",
            "docs",
            "+update",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--command",
            "block_move_after",
            "--block-id",
            document_id,
            "--src-block-ids",
            block_id,
        ],
        "move zip file block to doc begin",
    )


def assert_zip_block_at_doc_begin(doc_url: str, block_id: str) -> None:
    """Runtime gate: the first body block MUST be the newly attached file block."""

    output = run_subprocess(
        [
            "lark-cli",
            "docs",
            "+fetch",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--doc-format",
            "xml",
            "--detail",
            "with-ids",
            "-q",
            ".data.document.content",
        ],
        "verify zip file block position",
    )
    body = re.sub(r"^\s*<title\b[^>]*>.*?</title>", "", output.strip(), count=1, flags=re.S)
    first_block = re.search(r'<(\w+)\s+id="([^"]+)"', body.strip())
    if not first_block or first_block.group(2) != block_id:
        actual = first_block.group(2) if first_block else "<none>"
        raise RuntimeError(
            "ZIP file block position verification FAILED: expected the first body block "
            f"to be {block_id} (BLOCK_BEGIN, right below the title), got {actual}. "
            "文件块回挂位置不符合契约，拒绝宣称发布成功。"
        )


ZIP_FIGURE_RE = re.compile(
    r'<figure\s+id="(?P<block_id>[^"]+)"[^>]*>\s*<source\s+id="[^"]*"\s+name="(?P<file_name>[^"]*)"'
    r'[^>]*?token="(?P<file_token>[^"]+)"',
    re.S,
)


def list_doc_zip_file_blocks(doc_url: str) -> list[Dict[str, str]]:
    """Enumerate every ZIP file block (figure/source) in the doc via `docs +fetch`.

    The former `docx.v1.document_block.list` internal proxy path is no longer
    supported (`unsupported lark method_name`), and it degraded *silently* to an
    empty list — which is exactly how 8 stale ZIP blocks could pile up unnoticed.
    Parsing the authoritative `--doc-format xml --detail with-ids` payload keeps
    the enumeration on the same channel we already use for the position gate.
    """

    output = run_subprocess(
        [
            "lark-cli",
            "docs",
            "+fetch",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--doc-format",
            "xml",
            "--detail",
            "with-ids",
            "-q",
            ".data.document.content",
        ],
        "enumerate doc zip file blocks",
    )
    blocks: list[Dict[str, str]] = []
    for match in ZIP_FIGURE_RE.finditer(output):
        file_name = match.group("file_name") or ""
        if not file_name.lower().endswith(".zip"):
            continue
        blocks.append(
            {
                "block_id": match.group("block_id"),
                "file_name": file_name,
                "file_token": match.group("file_token"),
            }
        )
    return blocks


def is_own_skill_zip(file_name: str, skill_name: str) -> bool:
    """Does this ZIP file name belong to the skill currently being published?

    Accepts version-suffixed variants such as `skill-x.zip`, `skill-x_v1.2.zip`,
    `skill-x-1.2.zip`, `skill-x (1).zip`.
    """

    name = (file_name or "").strip()
    if not name.lower().endswith(".zip") or not skill_name:
        return False
    stem = name[: -len(".zip")]
    if stem == skill_name:
        return True
    return bool(re.fullmatch(re.escape(skill_name) + r"[\s_\-.(]*[vV]?[\d.\-_() ]*", stem))


def delete_doc_blocks(doc_url: str, block_ids: list[str]) -> str:
    """Physically delete top-level blocks by id (`block_delete` supports batch)."""

    if not block_ids:
        return ""
    return run_subprocess(
        [
            "lark-cli",
            "docs",
            "+update",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--command",
            "block_delete",
            "--block-id",
            ",".join(block_ids),
        ],
        "delete stale zip file blocks",
    )


def prune_stale_zip_blocks(
    doc_url: str,
    skill_name: str,
    new_block_id: str,
) -> Dict[str, Any]:
    """Idempotent replacement: keep exactly ONE ZIP block for this skill.

    Ordering is deliberate — the caller must have already inserted, relocated and
    asserted the NEW block before pruning, so a failed insert can never leave the
    doc without any package. Foreign ZIP blocks (other skills) are reported only,
    never auto-deleted: deleting someone else's asset is destructive and needs a
    human call.
    """

    report: Dict[str, Any] = {
        "enumerated": [],
        "deleted": [],
        "foreign": [],
        "degraded": "",
        "unique_ok": None,
        "residual": [],
    }
    try:
        blocks = list_doc_zip_file_blocks(doc_url)
    except Exception as exc:  # noqa: BLE001 - degrade, never break the pipeline
        report["degraded"] = f"enumerate failed: {exc}"
        print(
            "⚠️ WARNING: 无法枚举文档现有 ZIP 文件块，降级为「只插入不删除」；"
            f"旧版本文件块可能残留，请人工清理。原因：{exc}"
        )
        return report

    report["enumerated"] = blocks
    own_stale = [
        b for b in blocks if is_own_skill_zip(b["file_name"], skill_name) and b["block_id"] != new_block_id
    ]
    report["foreign"] = [b for b in blocks if not is_own_skill_zip(b["file_name"], skill_name)]

    if report["foreign"]:
        print("⚠️ 检测到非本技能的 ZIP 文件块（异物块，仅报告不自动删除，需人工确认）：")
        for b in report["foreign"]:
            print(f"   - block_id={b['block_id']} file={b['file_name']}")

    if own_stale:
        print(f"🧹 Deleting {len(own_stale)} stale ZIP file block(s) of {skill_name}...")
        try:
            delete_doc_blocks(doc_url, [b["block_id"] for b in own_stale])
            report["deleted"] = own_stale
        except Exception as exc:  # noqa: BLE001
            report["degraded"] = f"delete failed: {exc}"
            print(f"⚠️ WARNING: 旧 ZIP 文件块删除失败，需人工清理。原因：{exc}")
    else:
        print("ℹ️ No stale ZIP file block of this skill found; nothing to prune.")

    # RAW read-after-write uniqueness assertion (report-only, never silent).
    time.sleep(2)
    try:
        after = list_doc_zip_file_blocks(doc_url)
    except Exception as exc:  # noqa: BLE001
        report["degraded"] = f"{report['degraded']}; readback failed: {exc}".strip("; ")
        print(f"⚠️ WARNING: 删除后回读断言失败，无法确认唯一性。原因：{exc}")
        return report

    mine = [b for b in after if is_own_skill_zip(b["file_name"], skill_name)]
    report["residual"] = mine
    report["unique_ok"] = len(mine) == 1 and mine[0]["block_id"] == new_block_id
    if report["unique_ok"]:
        print(f"✅ ZIP block uniqueness verified: exactly 1 block ({new_block_id}) for {skill_name}.")
    else:
        print(
            "⚠️ WARNING: 本技能 ZIP 文件块唯一性断言未通过，残留清单如下（需人工清理）："
        )
        for b in mine:
            print(f"   - block_id={b['block_id']} file={b['file_name']} token={b['file_token']}")
    return report


def attach_zip_to_doc_via_mcp(doc_url: str, zip_path: Path, skill_name: str = "") -> str:
    result = subprocess.run(
        [
            "lark-cli",
            "docs",
            "+media-insert",
            "--as",
            "user",
            "--doc",
            doc_url,
            "--file",
            zip_path.name,
            "--type",
            "file",
        ],
        cwd=str(zip_path.resolve().parent),
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"insert native file block via lark-cli failed with exit code {result.returncode}\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )

    attach_output = result.stdout.strip()

    # `+media-insert` appends to the document END; relocate to BLOCK_BEGIN and verify.
    block_id = _parse_block_id_from_attach_output(attach_output)
    if not block_id:
        raise RuntimeError(
            "Unable to parse the new file block_id from lark-cli media-insert output; "
            f"cannot guarantee BLOCK_BEGIN placement.\nSTDOUT:\n{attach_output}"
        )

    print(f"🚚 Relocating zip file block {block_id} to BLOCK_BEGIN (below title)...")
    move_block_to_doc_begin(doc_url, block_id)
    time.sleep(2)
    assert_zip_block_at_doc_begin(doc_url, block_id)
    print("✅ Zip file block verified at BLOCK_BEGIN (right below the doc title).")

    # Idempotent replacement instead of blind append: only after the NEW block is
    # verified in place do we prune this skill's stale ZIP blocks.
    effective_skill_name = skill_name or zip_path.stem
    prune_stale_zip_blocks(doc_url, effective_skill_name, block_id)

    return attach_output


def download_doc_markdown(doc_url: str) -> Path:
    paths = mcp_lark_download(doc_url)
    for path in paths:
        if path.suffix.lower() in {".md", ".lark.md"} or path.name.endswith(".lark.md"):
            return path.resolve()
    if paths:
        return paths[0].resolve()
    raise RuntimeError(f"Unable to download markdown for doc: {doc_url}")


def sync_version_to_skill_doc_via_mcp(doc_url: str, new_version: str) -> None:
    """Sync point #3: replace version marker inside the Feishu skill doc.

    This is a best-effort update:
    - We download the doc as .lark.md with block markers
    - Find the first block that contains an explicit version marker
    - Update that block via lark MCP

    If no marker is found, we will log a warning and skip.
    """

    md_path = download_doc_markdown(doc_url)
    text = md_path.read_text(encoding="utf-8", errors="ignore")

    block_re = re.compile(r"^<!--\s*(BLOCK_\d+)\s*\|\s*([^\s]+)\s*-->\s*$")
    end_re = re.compile(r"^<!--\s*END_BLOCK_\d+\s*-->\s*$")
    version_re = re.compile(r"(?i)((?:version|版本号|版本)\s*[:：]\s*)([vV]?\d+(?:\.\d+){1,2})")

    lines = text.splitlines()
    i = 0
    while i < len(lines):
        m = block_re.match(lines[i])
        if not m:
            i += 1
            continue

        block_number = m.group(1)
        block_id = m.group(2)
        i += 1

        content_lines: list[str] = []
        while i < len(lines) and not end_re.match(lines[i]):
            content_lines.append(lines[i])
            i += 1

        content = "\n".join(content_lines).strip("\n")
        if version_re.search(content):
            updated_content = version_re.sub(r"\\1" + new_version, content)
            print(f"📝 Syncing doc version marker via lark-cli: {block_number} | {block_id}")
            run_subprocess(
                [
                    "lark-cli",
                    "docs",
                    "+update",
                    "--api-version",
                    "v2",
                    "--as",
                    "user",
                    "--doc",
                    doc_url,
                    "--command",
                    "str_replace",
                    "--doc-format",
                    "markdown",
                    "--pattern",
                    content,
                    "--content",
                    updated_content,
                ],
                "sync doc version marker",
            )
            return

        # skip end marker
        i += 1

    print("⚠️ No explicit version marker found in doc; skip SSOT doc sync.")


def verify_file_block_attached(doc_url: str, zip_name: str) -> bool:
    markdown_path = download_doc_markdown(doc_url)
    content = markdown_path.read_text(encoding="utf-8")
    zip_stem = Path(zip_name).stem
    return f"file_{zip_stem}" in content or zip_name in content or zip_stem in content


def move_doc_to_wiki_via_mcp(doc_url: str, wiki_node_token: str) -> Dict[str, str]:
    doc_token = doc_url.rstrip("/").split("/")[-1].split("?")[0]
    node_info_output = run_subprocess(
        [
            "lark-cli",
            "wiki",
            "+node-get",
            "--as",
            "user",
            "--node-token",
            wiki_node_token,
            "--format",
            "json",
        ],
        "resolve target wiki node",
    )
    try:
        node_info = json.loads(node_info_output[node_info_output.find("{"):])
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Unable to parse target wiki node info: {node_info_output}") from exc
    data = node_info.get("data", {}) if isinstance(node_info, dict) else {}
    node = data.get("node", data) if isinstance(data, dict) else {}
    space_id = str(node.get("space_id") or data.get("space_id") or "").strip()
    if not space_id:
        raise RuntimeError(f"Unable to resolve target wiki space id: {node_info_output}")

    output = run_subprocess(
        [
            "lark-cli",
            "wiki",
            "+move",
            "--as",
            "user",
            "--obj-type",
            "docx",
            "--obj-token",
            doc_token,
            "--target-space-id",
            space_id,
            "--target-parent-token",
            wiki_node_token,
            "--format",
            "json",
        ],
        "lark move doc to wiki",
    )

    urls = re.findall(r'https?://[^\s"\'<>]+', output)
    moved_doc_url = next((url for url in urls if "/docx/" in url or "/docs/" in url), doc_url)
    moved_wiki_url = next((url for url in urls if "/wiki/" in url), "")
    if not moved_wiki_url:
        try:
            payload = json.loads(output[output.find("{"):])
            payload_text = json.dumps(payload, ensure_ascii=False)
            urls = re.findall(r'https?://[^\s"\'<>]+', payload_text)
            moved_wiki_url = next((url for url in urls if "/wiki/" in url), "")
            moved_doc_url = next((url for url in urls if "/docx/" in url or "/docs/" in url), moved_doc_url)
        except json.JSONDecodeError:
            pass

    return {
        "doc_link": moved_doc_url,
        "wiki_url": moved_wiki_url or f"https://bytedance.larkoffice.com/wiki/{wiki_node_token}",
        "wiki_node_token": wiki_node_token,
        "raw_output": output,
    }


def parse_doc_token(doc_url: str) -> str:
    path = urlparse(doc_url).path or ""
    parts = [part for part in path.split("/") if part]
    if len(parts) >= 2 and parts[0] in {"docx", "docs", "doc"}:
        return parts[1]
    raise ValueError(f"Unsupported doc url: {doc_url}")


def call_lark_proxy(method_name: str, *, path: Optional[Dict[str, Any]] = None, query: Optional[Dict[str, Any]] = None, body: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    api_host = (os.environ.get("IRIS_RUNTIME_AIME_API_HOST") or "aime.bytedance.net").strip()
    if api_host.startswith("http://") or api_host.startswith("https://"):
        proxy_url = f"{api_host.rstrip('/')}/api/agents/v2/internal/proxy/lark_api"
    else:
        proxy_url = f"https://{api_host}/api/agents/v2/internal/proxy/lark_api"

    response = requests.post(
        proxy_url,
        headers={"Authorization": f"Byte-Cloud-JWT {get_raw_cloud_jwt()}"},
        json={
            "method_name": method_name,
            "path": path or {},
            "query": query or {},
            "body": body or {},
            "use_user_token": True,
        },
        timeout=60,
    )
    data = response.json()
    if response.status_code != 200:
        raise RuntimeError(f"lark proxy failed: http={response.status_code}, resp={data}")
    return data


def list_doc_file_blocks(doc_url: str) -> list[Dict[str, str]]:
    document_id = parse_doc_token(doc_url)
    response = call_lark_proxy(
        "docx.v1.document_block.list",
        path={"document_id": document_id},
        query={"page_size": "500"},
    )

    if isinstance(response, dict) and response.get("code") not in (0, None):
        # Never degrade silently: an unsupported/failed proxy call used to return an
        # empty list, which made stale-block pruning a no-op without any signal.
        raise RuntimeError(f"docx.v1.document_block.list failed: {response}")

    items = response.get("data", {}).get("items") or response.get("items") or []
    file_blocks: list[Dict[str, str]] = []
    for item in items:
        file_info = item.get("file") or {}
        file_token = file_info.get("file_token") or file_info.get("token") or ""
        file_name = file_info.get("name") or file_info.get("file_name") or ""
        if file_token:
            file_blocks.append(
                {
                    "block_id": item.get("block_id", ""),
                    "file_token": file_token,
                    "file_name": file_name,
                }
            )
    return file_blocks


def parse_file_token_from_attach_output(attach_output: str) -> str:
    if not attach_output:
        return ""

    json_like_matches = re.findall(r'\{[^{}]*"file_token"\s*:\s*"([^"]+)"[^{}]*\}', attach_output)
    if json_like_matches:
        return json_like_matches[0]

    dict_like_matches = re.findall(r"['\"]file_token['\"]\s*[:=]\s*['\"]([^'\"\s]+)['\"]", attach_output)
    if dict_like_matches:
        return dict_like_matches[0]

    # Some MCP / download outputs only expose a raw token-like field or a
    # canonical /file/<token> URL instead of a JSON object.
    url_match = re.search(r"/file/([A-Za-z0-9_-]+)", attach_output)
    if url_match:
        return url_match.group(1)

    loose_token_match = re.search(
        r"(?:^|[\s'\"=:,])(file_[A-Za-z0-9_-]{10,}|boxcn[A-Za-z0-9_-]{10,}|[A-Za-z0-9_-]{20,})(?=$|[\s'\",}])",
        attach_output,
    )
    if loose_token_match:
        return loose_token_match.group(1)

    return ""


def resolve_attached_drive_file_token(
    attach_output: str,
    before_blocks: list[Dict[str, str]],
    doc_url: str,
    file_name: str,
    before_blocks_error: str = "",
) -> str:
    """Resolve the Drive file_token for the newly attached ZIP.

    Resolution order (fail-fast, but with an extra zero-trust fallback):
    1) Parse stdout/stderr from the MCP attach script (preferred, includes raw file_token).
    2) If that fails, call docx.v1.document_block.list via internal proxy and diff
       before/after file blocks to locate the new block.
    3) If the OpenAPI path fails (e.g. missing scope, doc type not supported),
       fall back to inspecting the latest downloaded .lark.md and reusing the
       same parsing heuristics we use for attach_output (searching for /file/<token>
       or JSON blobs containing "file_token").

    Only after all three paths fail do we hard-stop, to avoid producing an
    orphan ZIP without a recorded Drive token / permission grant.
    """

    # 1) Primary: parse attach_output directly (most structured / least fragile).
    parsed_token = parse_file_token_from_attach_output(attach_output)
    if parsed_token:
        return parsed_token

    attach_summary = "attach_output missing parseable file_token"
    list_failure_summary = before_blocks_error or "before attachment snapshot ok"

    # 2) Secondary: try listing file blocks via internal Lark proxy and diffing.
    try:
        after_blocks = list_doc_file_blocks(doc_url)
        return find_new_file_token(before_blocks, after_blocks, file_name)
    except Exception as exc:
        list_failure_summary = f"{list_failure_summary}; after attachment lookup failed: {exc}"

    # 3) Tertiary: zero-trust markdown introspection — download latest doc as
    # .lark.md and reuse the same token parsing heuristics on its content.
    try:
        md_path = download_doc_markdown(doc_url)
        md_text = md_path.read_text(encoding="utf-8", errors="ignore")
        markdown_token = parse_file_token_from_attach_output(md_text)
        if markdown_token:
            return markdown_token
    except Exception as exc:
        list_failure_summary = f"{list_failure_summary}; markdown introspection failed: {exc}"

    # All paths failed: refuse to continue in order to avoid delivering an
    # untracked, potentially permissionless orphan Drive file.
    raise RuntimeError(
        "Unable to resolve attached drive file token; refuse to continue because this would "
        "deliver a potentially permissionless orphan file. "
        f"attach_output path failed: {attach_summary}. "
        f"fallbacks failed: {list_failure_summary}."
    )


def find_new_file_token(before_blocks: list[Dict[str, str]], after_blocks: list[Dict[str, str]], file_name: str) -> str:
    before_tokens = {item.get("file_token", "") for item in before_blocks}
    named_candidates = [
        item for item in after_blocks if item.get("file_name") == file_name and item.get("file_token") not in before_tokens
    ]
    if named_candidates:
        return named_candidates[0]["file_token"]

    new_candidates = [item for item in after_blocks if item.get("file_token") not in before_tokens]
    if new_candidates:
        return new_candidates[0]["file_token"]

    raise RuntimeError(f"Unable to locate the newly inserted File Block token for {file_name}")


def extract_metadata(
    name: str,
    desc: str,
    doc_link: str,
    version: str = "",
    skill_id: Optional[str] = None,
    skill_dir: Optional[Path] = None,
    zip_path: Optional[Path] = None,
    drive_file_token: Optional[str] = None,
    drive_file_url: Optional[str] = None,
    wiki_url: str = "",
    wiki_node_token: str = "",
) -> Dict[str, Any]:
    now = datetime.datetime.now()
    return {
        "skill_id": skill_id or f"SKILL-{now.strftime('%m%d%H%M')}",
        "name": name,
        "description": desc,
        "version": version,
        "doc_link": doc_link,
        "wiki_url": wiki_url,
        "wiki_node_token": wiki_node_token,
        "date": now.strftime("%Y-%m-%d"),
        "updated_at": now.strftime("%Y-%m-%d %H:%M"),
        "skill_dir": str(skill_dir.resolve()) if skill_dir else "",
        "zip_path": str(zip_path.resolve()) if zip_path else "",
        "zip_name": zip_path.name if zip_path else "",
        "drive_file_token": drive_file_token or "",
        "drive_file_url": drive_file_url or "",
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--name", required=True, help="目标技能名称")
    parser.add_argument("--desc", required=True, help="目标技能功能描述")
    parser.add_argument("--path", required=True, help="目标技能说明飞书文档 URL")
    parser.add_argument("--skill-dir", help="目标技能目录，例如 user_skills/xxx")
    parser.add_argument("--id", required=False, help="技能编号或包 ID")
    parser.add_argument("--user-email", default=DEFAULT_EMAIL, help="默认授予 Full Access 的邮箱")
    parser.add_argument("--output-zip", help="可选，自定义 zip 输出路径")
    parser.add_argument(
        "--wiki-node-token",
        default=DEFAULT_WIKI_NODE_TOKEN,
        help="技能说明文档默认迁入的 Wiki 节点 token（默认 Aime 技能库根节点）。",
    )
    parser.add_argument(
        "--skip-wiki-mount",
        action="store_true",
        help="调试用：跳过 Wiki Mount Phase（正式发布默认必须执行）。",
    )

    # SSOT version sync bus
    parser.add_argument(
        "--bump",
        choices=["major", "minor"],
        help="SSOT 版本升迁：major(+1.0) / minor(+0.1)。若未提供且为交互式终端，会提示选择；非交互式将直接报错。",
    )
    parser.add_argument(
        "--new-version",
        help="直接指定新版本号（标准 X.Y 两段式，如 5.2）。优先级高于 --bump。",
    )
    parser.add_argument(
        "--initial-version",
        default=DEFAULT_INITIAL_VERSION,
        help=(
            "首次发布时的起始版本号（默认 1.1）。"
            "当 SKILL.md 当前版本仍处于 0.x 脚手架阶段时，"
            "本流水线会忽略 --bump，直接将版本设为该值。"
        ),
    )
    parser.add_argument(
        "--skip-ssot",
        action="store_true",
        help="调试用：跳过 SSOT 版本同步（不会改动目标技能的 SKILL.md，也不会同步台账/文档版本标识）。",
    )
    parser.add_argument(
        "--ssot-spreadsheet-token",
        default=DEFAULT_SSOT_SPREADSHEET_TOKEN,
        help="飞书技能台账 Spreadsheet Token 或完整 URL（默认使用主台账）。",
    )
    parser.add_argument(
        "--ssot-sheet-name",
        default=DEFAULT_SSOT_SHEET_NAME,
        help="飞书技能台账工作表名称（默认：专属技能清单）。",
    )
    parser.add_argument(
        "--skip-ssot-sheet-sync",
        action="store_true",
        help="调试用：跳过技能台账整行 upsert（omni-asset-archiver）。",
    )
    parser.add_argument(
        "--skip-ssot-doc-sync",
        action="store_true",
        help="调试用：跳过说明文档版本标识同步。",
    )

    parser.add_argument(
        "--skip-remote",
        action="store_true",
        help="仅执行本地打包与 metadata 生成，跳过飞书云盘上传、权限赋予、文档挂载与 SSOT 远端同步",
    )
    args = parser.parse_args()

    workspace_root = get_workspace_root()
    skill_dir = (workspace_root / args.skill_dir).resolve() if args.skill_dir else None
    zip_path: Optional[Path] = None
    drive_file_token = ""
    drive_file_url = ""
    final_doc_link = args.path
    wiki_url = ""
    wiki_node_token = ""

    # SSOT version (X.Y)
    ssot_version = ""
    if skill_dir:
        if args.skip_ssot:
            ssot_version = read_skill_version_from_skill_md(skill_dir)
            print(f"🚌 SSOT disabled (--skip-ssot). Keep current version: {ssot_version}")
        else:
            current_version = read_skill_version_from_skill_md(skill_dir)
            new_version = ""

            if args.new_version:
                major, minor = _normalize_version_to_int_pair(args.new_version)
                new_version = _format_version_pair(major, minor)
            elif is_initial_version(current_version):
                # First publish: ignore --bump, jump straight to initial version (default 1.1).
                major, minor = _normalize_version_to_int_pair(args.initial_version)
                new_version = _format_version_pair(major, minor)
                print(
                    f"🚌 SSOT initial publish detected (current={current_version}). "
                    f"Forcing initial version: {new_version}"
                )
            else:
                bump_type = args.bump
                if not bump_type:
                    if sys.stdin.isatty():
                        print("\n=== SSOT Version Bump ===")
                        print(f"current: {current_version}")
                        print("choose bump type:")
                        print("  1) minor (+0.1)")
                        print("  2) major (+1.0)")
                        choice = (input("Select (default=1): ") or "1").strip()
                        bump_type = "major" if choice == "2" else "minor"
                    else:
                        raise RuntimeError(
                            "SSOT version sync requires --bump {major|minor} or --new-version X.Y in non-interactive mode."
                        )
                new_version = bump_version(current_version, bump_type)

            ssot_version = new_version
            if ssot_version != current_version:
                print(f"🚌 SSOT bump: {current_version} -> {ssot_version}")
                write_skill_md_version(skill_dir, ssot_version)
            else:
                print(f"🚌 SSOT version unchanged: {ssot_version}")

        # --- CDA Guardrails Checkpoint (fail fast) ---
        print("🚧 Running CDA-Guardrails-Selfcheck before packaging...")
        print(run_cda_guardrails_selfcheck(skill_dir))

        output_zip = Path(args.output_zip).resolve() if args.output_zip else skill_dir.parent / f"{skill_dir.name}.zip"
        print(f"🚀 Packaging skill directory: {skill_dir}")
        zip_path = create_skill_zip(skill_dir, output_zip)
        print(f"✅ Skill package created: {zip_path}")

        if not args.skip_remote:
            # Per system guardrails: authenticate bytedcli before any Feishu MCP operations.
            ensure_bytedcli_auth()
            bearer_token = ensure_bearer_token()

            print("🚧 Validating skill doc template markers...")
            call_with_retry(
                "validate skill doc template markers",
                lambda: validate_doc_template_markers(args.path),
            )

            before_blocks: list[Dict[str, str]] = []
            before_blocks_error = ""
            try:
                before_blocks = list_doc_file_blocks(args.path)
            except Exception as exc:
                before_blocks_error = f"before attachment snapshot failed: {exc}"
                print(f"⚠️ Unable to list doc blocks before attachment: {exc}")

            print("🚀 Inserting native File Block into skill doc via Lark MCP...")
            attach_output = call_with_retry(
                "insert file block into skill doc",
                lambda: attach_zip_to_doc_via_mcp(args.path, zip_path, skill_dir.name),
            )
            verified = call_with_retry(
                "verify attached file block via lark MCP download",
                lambda: verify_file_block_attached(args.path, zip_path.name),
            )
            if not verified:
                raise RuntimeError(f"File Block verification failed for {zip_path.name}")

            print("✅ Skill doc attachment sync finished and verified.")
            if attach_output:
                print(attach_output)

            if ssot_version and (not args.skip_ssot) and (not args.skip_ssot_doc_sync):
                print("🚌 Syncing SSOT version marker to skill doc...")
                call_with_retry(
                    "sync ssot version marker",
                    lambda: sync_version_to_skill_doc_via_mcp(args.path, ssot_version),
                )

            if args.skip_wiki_mount:
                print("⚠️ Wiki Mount Phase skipped by --skip-wiki-mount (debug only).")
            else:
                print("🚀 Wiki Mount Phase: moving skill doc into target wiki node...")
                wiki_mount_result = call_with_retry(
                    "move skill doc to wiki",
                    lambda: move_doc_to_wiki_via_mcp(args.path, args.wiki_node_token),
                )
                final_doc_link = wiki_mount_result.get("doc_link") or args.path
                wiki_url = wiki_mount_result.get("wiki_url") or DEFAULT_WIKI_URL
                wiki_node_token = wiki_mount_result.get("wiki_node_token") or args.wiki_node_token
                print("✅ Wiki Mount Phase finished.")
                if wiki_mount_result.get("raw_output"):
                    print(wiki_mount_result["raw_output"])

            drive_file_token = resolve_attached_drive_file_token(
                attach_output=attach_output,
                before_blocks=before_blocks,
                doc_url=final_doc_link,
                file_name=zip_path.name,
                before_blocks_error=before_blocks_error,
            )
            drive_file_url = build_drive_file_url(final_doc_link, drive_file_token)
            print(f"🚀 Repairing drive asset access for {args.user_email} via MCP personal flow...")
            repair_output = call_with_retry(
                "repair drive asset access via MCP personal flow",
                lambda: ensure_drive_asset_access_via_mcp(
                    drive_file_url,
                    email=args.user_email,
                ),
            )
            print("✅ Drive asset access repaired.")
            if repair_output:
                print(repair_output)

    metadata = extract_metadata(
        args.name,
        args.desc,
        final_doc_link,
        ssot_version,
        args.id,
        skill_dir,
        zip_path,
        drive_file_token,
        drive_file_url,
        wiki_url,
        wiki_node_token,
    )

    if (
        skill_dir
        and ssot_version
        and (not args.skip_ssot)
        and (not args.skip_remote)
        and (not args.skip_ssot_sheet_sync)
    ):
        sync_result = call_with_retry(
            "sync skill inventory via omni-asset-archiver",
            lambda: sync_skill_inventory_via_omni(metadata),
        )
        row_number = int(sync_result.get("row_number") or 0)
        call_with_retry(
            "sync skill inventory updated_at formatter",
            lambda: ensure_skill_inventory_updated_at_formatter(row_number, metadata.get("updated_at", "")),
        )

    metadata_path = Path.cwd() / "metadata.json"
    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=4), encoding="utf-8")

    print("🚀 Metadata packaged for omni-asset-archiver:")
    print(json.dumps(metadata, indent=4, ensure_ascii=False))
    print(f"\n✅ Metadata written to {metadata_path.resolve()}")

    if skill_dir:
        print("🚀 Running post-forge git push hook...")
        print(run_post_forge_git_push(workspace_root, args.name, ssot_version or "latest"))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
