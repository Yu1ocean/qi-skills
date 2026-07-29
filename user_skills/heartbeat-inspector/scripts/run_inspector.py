#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List
from urllib.parse import parse_qs, urlparse

from scripts.chat_task_extractor import (
    TaskExtractionError,
    extract_status_updates_from_messages,
    extract_tasks_from_chat_messages,
)
from scripts.diff_engine import diff_feishu_messages, diff_sheet
from scripts.dlq import append_dlq
from scripts.heartbeat_config import ConfigError, Target, load_heartbeat_config, validate_and_normalize_config
from scripts.state_store import load_state, save_state

# 双轨写入（落盘：Aime日志 + 任务库）
from scripts.dual_write import DualTrackWriter
from scripts.lark_sheets_cli import LarkSheetsCLI


class FetchError(RuntimeError):
    pass


def _workspace_root() -> Path:
    env = os.environ.get("IRIS_WORKSPACE_PATH")
    if env:
        return Path(env).resolve()
    return Path(__file__).resolve().parents[3]


def _run_cmd(cmd: List[str], timeout: int = 60) -> str:
    p = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, timeout=timeout)
    if p.returncode != 0:
        raise FetchError(f"Command failed (code={p.returncode}): {' '.join(cmd)}\n{p.stdout}")
    return p.stdout


def _parse_json_from_stdout(text: str) -> Any:
    """Parse JSON from tool stdout."""
    text = text.strip()
    if not text:
        raise FetchError("empty stdout")

    try:
        return json.loads(text)
    except Exception:
        pass

    import json
    decoder = json.JSONDecoder()
    starts = [m.start() for m in re.finditer(r'\{|\[', text)]
    candidates = []
    for start in starts:
        try:
            obj, _ = decoder.raw_decode(text[start:])
            candidates.append(obj)
        except Exception:
            continue

    if not candidates:
        raise FetchError(f"stdout is not valid json:\n{text[-2000:]}")

    # Prioritize dicts with known keys (messages, data, items)
    for obj in reversed(candidates):
        if isinstance(obj, dict) and any(k in obj for k in ("messages", "data", "items")):
            return obj

    # Fallback to the last successfully parsed object
    return candidates[-1]


def _try_bytedcli_auth(verbose: bool, dlq_path: Path) -> bool:
    """Best-effort bytedcli auth.

    - 拉取数据（读）阶段：鉴权失败不应让整次巡检直接挂掉，所以这里依旧是 best-effort。
    - 写入（dual_write）阶段：由调用方决定是否必须严格依赖鉴权结果。

    返回：
    - True：鉴权成功
    - False：鉴权失败（已写入 DLQ）
    """

    root = _workspace_root()
    script = root / "inner_skills" / "bytedcli-auth" / "scripts" / "bytedcli_auth.sh"
    if not script.exists():
        append_dlq(dlq_path, {"type": "bytedcli_auth", "error": f"script_not_found: {script}"})
        return False

    try:
        out = _run_cmd(["bash", str(script)], timeout=60)
        if verbose:
            print(out, file=sys.stderr)
        return True
    except Exception as e:
        append_dlq(dlq_path, {"type": "bytedcli_auth", "error": str(e)})
        return False


def _message_text_full(msg: Dict[str, Any]) -> str:
    """Extract message text without truncation or whitespace normalization."""

    v = msg.get("content")
    if isinstance(v, str) and v:
        return v

    if isinstance(v, dict):
        for k in ["text", "content", "body", "value"]:
            vv = v.get(k)
            if isinstance(vv, str) and vv:
                return vv

    # Fallback: do NOT flatten/trim; keep full JSON as a lossless representation.
    if v is not None:
        return json.dumps(v, ensure_ascii=False, default=str)

    return json.dumps(msg, ensure_ascii=False, default=str)


def _guess_sender_name(msg: Dict[str, Any]) -> str:
    for k in ["sender_name", "sender", "user_name", "name"]:
        v = msg.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
        if isinstance(v, dict):
            n = v.get("name") or v.get("sender_name")
            if isinstance(n, str) and n.strip():
                return n.strip()
    return ""


def _guess_chat_name(msg: Dict[str, Any]) -> str:
    for k in ["chat_name", "group_name", "conversation_name"]:
        v = msg.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _load_chat_registry() -> Dict[str, str]:
    """Load verified chat_id -> chat name mapping from workspace CHAT_REGISTRY.json."""
    registry_path = _workspace_root() / "CHAT_REGISTRY.json"
    if not registry_path.exists():
        return {}

    try:
        data = json.loads(registry_path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    chats = data.get("chats") if isinstance(data, dict) else None
    if not isinstance(chats, dict):
        return {}

    mapping: Dict[str, str] = {}
    for item in chats.values():
        if not isinstance(item, dict):
            continue
        chat_id = item.get("chat_id")
        name = item.get("name")
        if isinstance(chat_id, str) and chat_id.startswith("oc_") and isinstance(name, str) and name.strip():
            mapping[chat_id] = name.strip()
    return mapping


def _verified_chat_name(chat_id: str, chat_registry: Dict[str, str]) -> str:
    """Return only registry-verified chat names; never trust natural-language names from messages/config."""
    if isinstance(chat_id, str) and chat_id.startswith("oc_"):
        name = chat_registry.get(chat_id)
        if isinstance(name, str) and name.strip():
            return name.strip()
        return f"未知群聊 (chat_id: {chat_id})"
    return "未知群聊"


def _guess_chat_id(msg: Dict[str, Any]) -> str:
    for k in ["chat_id", "open_chat_id", "openChatId"]:
        v = msg.get(k)
        if isinstance(v, str) and v.startswith("oc_"):
            return v

    chat = msg.get("chat")
    if isinstance(chat, dict):
        for k in ["chat_id", "open_chat_id", "openChatId"]:
            v = chat.get(k)
            if isinstance(v, str) and v.startswith("oc_"):
                return v

    return ""


def _extract_chat_id_from_applink(link: Any) -> str:
    normalized = _normalize_http_link(link)
    if not normalized:
        return ""

    try:
        parsed = urlparse(normalized)
        params = parse_qs(parsed.query)
    except Exception:
        return ""

    for key in ("open_chat_id", "openchatid"):
        values = params.get(key) or []
        if values:
            candidate = str(values[0] or "").strip()
            if candidate.startswith("oc_"):
                return candidate
    return ""


def _normalize_http_link(v: Any) -> str:
    if not isinstance(v, str):
        return ""
    s = v.strip()
    if s.startswith("https://") or s.startswith("http://"):
        return s
    return ""


def _build_feishu_chat_link(chat_id: str) -> str:
    if isinstance(chat_id, str) and chat_id.startswith("oc_"):
        return f"https://applink.larkoffice.com/client/chat/open?openChatId={chat_id}"
    return ""


def _guess_message_link(msg: Dict[str, Any]) -> str:
    for k in ["message_link", "message_url", "message_permalink", "permalink"]:
        link = _normalize_http_link(msg.get(k))
        if link:
            return link
    return ""


def _guess_chat_link(msg: Dict[str, Any]) -> str:
    for k in ["chat_link", "chat_url", "conversation_link", "group_link"]:
        link = _normalize_http_link(msg.get(k))
        if link:
            return link

    return _build_feishu_chat_link(_guess_chat_id(msg))


def _enrich_chat_message_context(
    msg: Dict[str, Any], *, default_chat_name: str = "", default_chat_id: str = "", chat_registry: Dict[str, str] | None = None
) -> Dict[str, Any]:
    out = dict(msg)

    chat_id = default_chat_id or _guess_chat_id(out)
    registry = chat_registry if isinstance(chat_registry, dict) else {}
    chat_name = _verified_chat_name(chat_id, registry)
    chat_link = _guess_chat_link(out) or _build_feishu_chat_link(chat_id)
    message_link = _guess_message_link(out)
    jump_link = message_link or chat_link

    if chat_id:
        out["chat_id"] = chat_id
    out["chat_name"] = chat_name
    if chat_link:
        out["chat_link"] = chat_link
    if message_link:
        out["message_link"] = message_link
    if jump_link:
        out["jump_link"] = jump_link

    return out


def _looks_like_system_broadcast(msg: Dict[str, Any]) -> bool:
    sender = _guess_sender_name(msg).strip().lower()
    sender_markers = {
        "system",
        "system message",
        "system notification",
        "系统",
        "系统消息",
        "系统通知",
        "系统广播",
        "飞书提醒",
        "lark notification",
    }
    if sender in sender_markers:
        return True

    for key in ("message_type", "msg_type", "type", "sender_type", "scene"):
        value = msg.get(key)
        if not isinstance(value, str):
            continue
        normalized = value.strip().lower()
        if normalized in {"system", "system_message", "system_notice", "broadcast", "notification"}:
            return True

    return False


def _contains_programmatic_broadcast_mention(text: str) -> bool:
    if not isinstance(text, str) or not text:
        return False
    return bool(re.search(r"(?:@_all|@all|@所有人|<at\s+id=all></at>)", text, flags=re.IGNORECASE))


def _should_filter_zero_trust_message(msg: Dict[str, Any]) -> bool:
    text = _message_text_full(msg)
    return _looks_like_system_broadcast(msg) or _contains_programmatic_broadcast_mention(text)


def _first_source_message(payload: Dict[str, Any]) -> Dict[str, Any]:
    items = payload.get("source_messages_full")
    if isinstance(items, list):
        for item in items:
            if isinstance(item, dict):
                return item
    return {}


def _pick_message_lookup_entry(message_id: str, payload: Any) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return {}

    messages = payload.get("messages")
    if not isinstance(messages, list):
        return {}

    fallback: Dict[str, Any] = {}
    for item in messages:
        if not isinstance(item, dict):
            continue
        if str(item.get("message_id") or "") == message_id:
            return item
        if not fallback and (
            _guess_chat_id(item)
            or _extract_chat_id_from_applink(item.get("message_app_link"))
            or _extract_chat_id_from_applink(item.get("message_link"))
        ):
            fallback = item
    return fallback


def _lookup_message_meta_by_id(message_id: str) -> Dict[str, str]:
    cmd = [
        "lark-cli",
        "im",
        "+messages-mget",
        "--message-ids",
        message_id,
        "--format",
        "json",
        "--as",
        "user",
    ]
    out = _run_cmd(cmd, timeout=90)
    data = _parse_json_from_stdout(out)
    picked = _pick_message_lookup_entry(message_id, data)
    if not picked:
        raise FetchError(f"message_id 未找到详情：{message_id}")

    chat_id = (
        _guess_chat_id(picked)
        or _extract_chat_id_from_applink(picked.get("message_app_link"))
        or _extract_chat_id_from_applink(picked.get("message_link"))
    )
    if not chat_id:
        raise FetchError(f"message_id 详情缺少 chat_id：{message_id}")

    message_link = _normalize_http_link(picked.get("message_app_link")) or _normalize_http_link(picked.get("message_link"))
    return {
        "chat_id": chat_id,
        "chat_link": _build_feishu_chat_link(chat_id),
        "message_link": message_link,
        "jump_link": message_link or _build_feishu_chat_link(chat_id),
    }


def _search_chat_meta_by_name(chat_name: str) -> Dict[str, str]:
    root = _workspace_root()
    script = root / "inner_skills" / "feishu-im-read" / "scripts" / "feishu_im_user_search_chats.js"
    if not script.exists():
        raise FetchError(f"script_not_found: {script}")

    inp = {"query": chat_name, "page_size": 50}
    out = _run_cmd(["node", str(script), "--input", json.dumps(inp, ensure_ascii=False)], timeout=90)
    data = _parse_json_from_stdout(out)

    if not isinstance(data, dict):
        raise FetchError(f"unexpected search_chats schema: {type(data)}")

    exact = data.get("exact_match")
    if isinstance(exact, dict):
        chat_id = exact.get("chat_id")
        if isinstance(chat_id, str) and chat_id.startswith("oc_"):
            return {"chat_id": chat_id, "name": str(exact.get("name") or chat_name).strip()}

    chats = data.get("chats")
    if isinstance(chats, list):
        candidates = [c for c in chats if isinstance(c, dict) and isinstance(c.get("chat_id"), str)]
        if len(candidates) == 1:
            item = candidates[0]
            return {"chat_id": str(item["chat_id"]), "name": str(item.get("name") or chat_name).strip()}
        if len(candidates) > 1:
            names = [str(c.get("name") or "") for c in candidates[:5]]
            raise FetchError(f"chat_name 命中多个群聊，无法唯一确定：{chat_name} candidates={names}")

    raise FetchError(f"chat_name 未找到匹配群聊：{chat_name}")


def _search_chat_id_by_name(chat_name: str) -> str:
    return _search_chat_meta_by_name(chat_name)["chat_id"]


def _recover_mention_chat_meta(
    msg: Dict[str, Any], chat_registry: Dict[str, str], runtime_cache: Dict[str, Any], dlq_path: Path
) -> Dict[str, str]:
    """Recover chat_id/chat_name for global mentions when search result omits chat context."""

    chat_id = _guess_chat_id(msg)
    if chat_id:
        return {"chat_id": chat_id, "name": _verified_chat_name(chat_id, chat_registry)}

    message_id = str(msg.get("message_id") or "")
    message_cache = runtime_cache.setdefault("mention_message_meta_cache", {})
    if message_id and isinstance(message_cache, dict):
        cached = message_cache.get(message_id)
        if isinstance(cached, dict):
            cached_id = str(cached.get("chat_id") or "")
            if cached_id.startswith("oc_"):
                cached_name = str(cached.get("name") or chat_registry.get(cached_id) or "").strip()
                out = dict(cached)
                out["chat_id"] = cached_id
                out["name"] = cached_name or _verified_chat_name(cached_id, chat_registry)
                return out

    if message_id:
        try:
            meta = _lookup_message_meta_by_id(message_id)
            recovered_id = meta["chat_id"]
            recovered_name = _verified_chat_name(recovered_id, chat_registry)
            out = dict(meta)
            out["name"] = recovered_name
            if isinstance(message_cache, dict):
                message_cache[message_id] = out
            return out
        except Exception as e:
            append_dlq(
                dlq_path,
                {
                    "type": "recover_mention_chat_by_message_id",
                    "message_id": message_id,
                    "error": str(e),
                },
            )

    raw_name = _guess_chat_name(msg)
    if not raw_name:
        return {"chat_id": "", "name": "未知群聊"}

    cache = runtime_cache.setdefault("mention_chat_name_cache", {})
    if isinstance(cache, dict):
        cached = cache.get(raw_name)
        if isinstance(cached, dict):
            cached_id = cached.get("chat_id")
            cached_name = cached.get("name")
            if isinstance(cached_id, str) and cached_id.startswith("oc_"):
                return {"chat_id": cached_id, "name": str(cached_name or raw_name).strip()}

    try:
        meta = _search_chat_meta_by_name(raw_name)
        recovered_id = meta["chat_id"]
        recovered_name = str(chat_registry.get(recovered_id) or meta.get("name") or raw_name).strip()
        if isinstance(cache, dict):
            cache[raw_name] = {"chat_id": recovered_id, "name": recovered_name}
        return {"chat_id": recovered_id, "name": recovered_name}
    except Exception as e:
        append_dlq(
            dlq_path,
            {
                "type": "recover_mention_chat",
                "message_id": message_id,
                "chat_name": raw_name,
                "error": str(e),
            },
        )
        return {"chat_id": "", "name": "未知群聊"}


def _resolve_chat_id(target: Target, prev_state: Dict[str, Any], dlq_path: Path) -> str:
    chat_id = target.raw.get("chat_id")
    if isinstance(chat_id, str) and chat_id.startswith("oc_"):
        return chat_id

    # Cache from previous successful resolution
    cached = prev_state.get("resolved_chat_id")
    if isinstance(cached, str) and cached.startswith("oc_"):
        return cached

    chat_name = target.raw.get("chat_name")
    if isinstance(chat_name, str) and chat_name.strip():
        try:
            resolved = _search_chat_id_by_name(chat_name.strip())
            return resolved
        except Exception as e:
            append_dlq(
                dlq_path,
                {
                    "type": "resolve_chat",
                    "target_id": target.id,
                    "chat_name": chat_name,
                    "error": str(e),
                },
            )
            raise

    raise FetchError("feishu_chat target requires chat_id (oc_xxx) or chat_name (群聊名称)")


def fetch_feishu_chat_messages(chat_id: str, relative_time: str, page_size: int) -> List[Dict[str, Any]]:
    root = _workspace_root()
    script = root / "inner_skills" / "feishu-im-read" / "scripts" / "feishu_im_user_get_messages.js"
    if not script.exists():
        raise FetchError(f"script_not_found: {script}")

    inp = {
        "chat_id": chat_id,
        "relative_time": relative_time,
        "page_size": page_size,
        "sort_rule": "create_time_asc",
    }

    out = _run_cmd(["node", str(script), "--input", json.dumps(inp, ensure_ascii=False)], timeout=90)
    data = _parse_json_from_stdout(out)

    if isinstance(data, dict):
        for key in ("messages", "data", "items"):
            msgs = data.get(key)
            if isinstance(msgs, list):
                return [m for m in msgs if isinstance(m, dict)]

    raise FetchError(f"unexpected feishu messages schema: {type(data)}")


def _get_self_email() -> str:
    # Prefer runtime env
    for k in ["AIME_CURRENT_USER_EMAIL", "IRIS_CURRENT_USER_EMAIL"]:
        v = os.environ.get(k)
        if v and "@" in v:
            return v

    # Last resort: current user name often exists but not email. Fail fast.
    raise FetchError(
        "无法获取当前用户邮箱（需要环境变量 AIME_CURRENT_USER_EMAIL 或 IRIS_CURRENT_USER_EMAIL）。"
    )


def _get_self_open_id(runtime_cache: Dict[str, Any], dlq_path: Path) -> str:
    cached = runtime_cache.get("self_open_id")
    if isinstance(cached, str) and cached.startswith("ou_"):
        return cached

    root = _workspace_root()
    script = root / "inner_skills" / "lark" / "mcp_lark_lark_user_info.py"
    if not script.exists():
        raise FetchError(f"script_not_found: {script}")

    email = _get_self_email()

    out = _run_cmd(["python3", str(script), json.dumps({"emails": [email]}, ensure_ascii=False)], timeout=60)
    data = _parse_json_from_stdout(out)

    # Be tolerant about schema; try to find open_id in common shapes.
    open_id: str | None = None
    if isinstance(data, list) and data:
        u0 = data[0]
        if isinstance(u0, dict) and isinstance(u0.get("open_id"), str):
            open_id = u0["open_id"]
    elif isinstance(data, dict):
        if isinstance(data.get("open_id"), str):
            open_id = data["open_id"]
        elif isinstance(data.get("user"), dict) and isinstance(data["user"].get("open_id"), str):
            open_id = data["user"]["open_id"]
        elif isinstance(data.get("users"), list) and data["users"]:
            u0 = data["users"][0]
            if isinstance(u0, dict) and isinstance(u0.get("open_id"), str):
                open_id = u0["open_id"]
        elif isinstance(data.get("data"), list) and data["data"]:
            u0 = data["data"][0]
            if isinstance(u0, dict) and isinstance(u0.get("open_id"), str):
                open_id = u0["open_id"]

    if not open_id or not open_id.startswith("ou_"):
        append_dlq(dlq_path, {"type": "self_open_id", "error": f"unexpected schema: {data}"})
        raise FetchError("无法解析当前用户 open_id（可能是不可见/未授权）")

    runtime_cache["self_open_id"] = open_id
    return open_id


def _relative_time_to_iso_window(relative_time: str) -> Dict[str, str]:
    value = (relative_time or "last_6_hours").strip().lower()
    now = datetime.now().astimezone().replace(microsecond=0)

    presets = {
        "last_30_minutes": timedelta(minutes=30),
        "last_1_hour": timedelta(hours=1),
        "last_3_hours": timedelta(hours=3),
        "last_6_hours": timedelta(hours=6),
        "last_12_hours": timedelta(hours=12),
        "last_24_hours": timedelta(hours=24),
        "last_2_days": timedelta(days=2),
        "last_3_days": timedelta(days=3),
        "last_7_days": timedelta(days=7),
    }
    delta = presets.get(value)

    if delta is None:
        m = re.fullmatch(r"last_(\d+)_(minute|minutes|hour|hours|day|days|week|weeks)", value)
        if not m:
            raise FetchError(f"unsupported relative_time: {relative_time}")
        amount = int(m.group(1))
        unit = m.group(2)
        if "minute" in unit:
            delta = timedelta(minutes=amount)
        elif "hour" in unit:
            delta = timedelta(hours=amount)
        elif "day" in unit:
            delta = timedelta(days=amount)
        else:
            delta = timedelta(weeks=amount)

    start = now - delta
    return {
        "start": start.isoformat(timespec="seconds"),
        "end": now.isoformat(timespec="seconds"),
    }



def fetch_feishu_mentions_global(relative_time: str, page_size: int) -> List[Dict[str, Any]]:
    window = _relative_time_to_iso_window(relative_time)
    cmd = [
        "lark-cli",
        "im",
        "+messages-search",
        "--chat-type",
        "group",
        "--is-at-me",
        "--start",
        window["start"],
        "--end",
        window["end"],
        "--page-size",
        str(page_size),
        "--as",
        "user",
        "--format",
        "json",
    ]

    out = _run_cmd(cmd, timeout=120)
    data = _parse_json_from_stdout(out)

    if isinstance(data, dict):
        payload = data.get("data") if isinstance(data.get("data"), dict) else data
        if isinstance(payload, dict):
            for key in ("messages", "items"):
                msgs = payload.get(key)
                if isinstance(msgs, list):
                    return [m for m in msgs if isinstance(m, dict)]

    raise FetchError(f"unexpected search_messages schema: {type(data)}")


def _excel_col_to_idx(col: str) -> int:
    col = col.upper().strip()
    n = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(col)
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _parse_a1_range(a1: str):
    # e.g. A1:Z200
    m = re.fullmatch(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", a1.strip())
    if not m:
        raise ValueError(f"invalid range: {a1}")
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    col1 = _excel_col_to_idx(c1)
    col2 = _excel_col_to_idx(c2)
    if r1 <= 0 or r2 <= 0 or col1 <= 0 or col2 <= 0:
        raise ValueError(a1)
    if r2 < r1 or col2 < col1:
        raise ValueError(a1)
    return r1, col1, r2, col2


def fetch_lark_sheet_matrix(target: Target) -> List[List[Any]]:
    root = _workspace_root()
    downloader = root / "inner_skills" / "lark" / "mcp_lark_lark_download.py"
    if not downloader.exists():
        raise FetchError(f"script_not_found: {downloader}")

    doc_url = target.raw.get("document_url")
    sheet_name = target.raw.get("sheet_name")
    rng = target.raw.get("range")
    if not isinstance(doc_url, str) or "larkoffice.com" not in doc_url:
        raise FetchError("lark_sheet_range target requires document_url")
    if not isinstance(sheet_name, str) or not sheet_name:
        raise FetchError("lark_sheet_range target requires sheet_name")
    if not isinstance(rng, str) or not rng:
        raise FetchError("lark_sheet_range target requires range")

    out = _run_cmd(["python3", str(downloader), json.dumps({"document_url": doc_url}, ensure_ascii=False)], timeout=120)
    data = _parse_json_from_stdout(out)

    if not (isinstance(data, list) and data):
        raise FetchError(f"unexpected download result: {data}")

    xlsx_path = None
    for p in data:
        if isinstance(p, str) and p.endswith(".xlsx"):
            xlsx_path = p
            break
    if not xlsx_path:
        xlsx_path = str(data[0])

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise FetchError(f"openpyxl_not_available: {e}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise FetchError(f"sheet_name_not_found: {sheet_name} (available={wb.sheetnames})")

    ws = wb[sheet_name]
    r1, c1, r2, c2 = _parse_a1_range(rng)

    matrix: List[List[Any]] = []
    for r in range(r1, r2 + 1):
        row: List[Any] = []
        for c in range(c1, c2 + 1):
            row.append(ws.cell(row=r, column=c).value)
        matrix.append(row)
    return matrix


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--heartbeat", default="HEARTBEAT.md")
    ap.add_argument("--state", default=".heartbeat_state.json")
    ap.add_argument("--dlq", default=".heartbeat_dlq.jsonl")
    ap.add_argument("--verbose", action="store_true")

    # 可选：把增量事件直接双写落盘到任务台账（Aime日志 + 任务库）
    ap.add_argument(
        "--dual-write-spreadsheet",
        default=None,
        help="可选：启用双轨写入。传入飞书表格 URL 或 token（支持 wiki URL / sheets URL / spreadsheet_token）",
    )
    ap.add_argument("--dual-write-log-sheet-title", default="Aime日志")
    ap.add_argument("--dual-write-task-sheet-title", default="任务库")
    ap.add_argument("--dual-write-dry-run", action="store_true", help="只计算写入行数，不实际写表")
    ap.add_argument(
        "--dual-write-no-raw-verify",
        action="store_true",
        help="禁用 dual_write RAW 写后即读校验（默认开启）",
    )

    args = ap.parse_args()

    root = _workspace_root()
    heartbeat_path = (root / args.heartbeat).resolve()
    state_path = (root / args.state).resolve()
    dlq_path = (root / args.dlq).resolve()

    # Best-effort auth once per run
    auth_ok = _try_bytedcli_auth(verbose=args.verbose, dlq_path=dlq_path)

    try:
        cfg = load_heartbeat_config(heartbeat_path)
        targets = validate_and_normalize_config(cfg)
    except ConfigError as e:
        append_dlq(dlq_path, {"type": "config", "error": str(e)})
        raise

    state = load_state(state_path)
    runtime_cache = state.setdefault("runtime_cache", {})
    chat_registry = _load_chat_registry()
    state_targets = state.setdefault("targets", {})

    alerts: List[str] = []
    # 仅包含 JSON 事件（可用于落盘双写）。非 JSON 的纯文本告警（如表格范围变化）不会进入该列表。
    alert_events: List[Dict[str, Any]] = []
    any_success = False

    for t in targets:
        prev = state_targets.get(t.id) if isinstance(state_targets, dict) else None
        if not isinstance(prev, dict):
            prev = {}

        for attempt in range(2):
            try:
                if t.type == "feishu_chat":
                    chat_id = _resolve_chat_id(t, prev, dlq_path)

                    relative_time = str(t.raw.get("relative_time") or "last_6_hours")
                    page_size = int(t.raw.get("page_size") or 50)

                    msgs = fetch_feishu_chat_messages(chat_id, relative_time=relative_time, page_size=page_size)
                    default_chat_name = str(t.raw.get("chat_name") or t.title or "")
                    msgs = [
                        _enrich_chat_message_context(
                            m,
                            default_chat_name=default_chat_name,
                            default_chat_id=chat_id,
                            chat_registry=chat_registry,
                        )
                        for m in msgs
                    ]

                    new_msgs, new_frag = diff_feishu_messages(prev, msgs)

                    # First run baseline: do not alert.
                    is_first_run = not bool(prev.get("last_seen_message_id"))
                    if new_msgs and not is_first_run:
                        # 1) Emit raw new messages (100% preserved)
                        for m in new_msgs:
                            mid = str(m.get("message_id") or "")
                            sender = _guess_sender_name(m)
                            msg_text = _message_text_full(m)
                            chat_name = _verified_chat_name(chat_id, chat_registry)
                            chat_link = _guess_chat_link(m)
                            message_link = _guess_message_link(m)
                            jump_link = message_link or chat_link
                            obj = {
                                "type": "chat_message_new",
                                "target_id": t.id,
                                "target_title": t.title,
                                "chat_id": chat_id,
                                "chat_name": chat_name,
                                "chat_link": chat_link,
                                "message_link": message_link,
                                "jump_link": jump_link,
                                "message_id": mid,
                                "sender": sender,
                                "create_time": m.get("create_time"),
                                "text": msg_text,
                            }
                            alert_events.append(obj)
                            alerts.append(
                                json.dumps(
                                    obj,
                                    ensure_ascii=False,
                                    separators=(",", ":"),
                                    default=str,
                                )
                            )

                        # 2) Extract tasks via LLM (batch)
                        try:
                            new_ids = [str(mm.get("message_id") or "") for mm in new_msgs if mm.get("message_id")]
                            tasks = extract_tasks_from_chat_messages(
                                chat_title=t.title,
                                messages=msgs,
                                new_message_ids=new_ids,
                            )
                            for task in tasks:
                                source0 = _first_source_message(task)
                                chat_name = _verified_chat_name(chat_id, chat_registry)
                                chat_link = str(source0.get("chat_link") or _build_feishu_chat_link(chat_id))
                                message_link = str(source0.get("message_link") or source0.get("jump_link") or "")
                                jump_link = message_link or chat_link
                                obj = {
                                    "type": "chat_task",
                                    "target_id": t.id,
                                    "target_title": t.title,
                                    "chat_id": chat_id,
                                    "chat_name": chat_name,
                                    "chat_link": chat_link,
                                    "message_link": message_link,
                                    "jump_link": jump_link,
                                    "message_id": str(source0.get("message_id") or ""),
                                    "task": task,
                                }
                                alert_events.append(obj)
                                alerts.append(
                                    json.dumps(
                                        obj,
                                        ensure_ascii=False,
                                        separators=(",", ":"),
                                        default=str,
                                    )
                                )

                            # State-Triggers: emit status updates for downstream dashboard sync
                            status_updates = extract_status_updates_from_messages(
                                view_messages=[
                                    {
                                        "message_id": str(mm.get("message_id") or ""),
                                        "create_time": mm.get("create_time"),
                                        "sender": _guess_sender_name(mm),
                                        "chat_id": chat_id,
                                        "chat_name": _guess_chat_name(mm) or default_chat_name,
                                        "chat_link": _guess_chat_link(mm),
                                        "message_link": _guess_message_link(mm),
                                        "jump_link": str(mm.get("jump_link") or _guess_message_link(mm) or _guess_chat_link(mm) or ""),
                                        "text": _message_text_full(mm),
                                    }
                                    for mm in msgs
                                ],
                                known_task_names=[str(tk.get("task_name") or "") for tk in tasks],
                            )
                            for su in status_updates:
                                chat_name = _verified_chat_name(chat_id, chat_registry)
                                chat_link = str(su.get("chat_link") or _build_feishu_chat_link(chat_id))
                                message_link = str(su.get("message_link") or su.get("jump_link") or "")
                                jump_link = message_link or chat_link
                                obj = {
                                    "type": "task_status_update",
                                    "target_id": t.id,
                                    "target_title": t.title,
                                    "chat_id": chat_id,
                                    "chat_name": chat_name,
                                    "chat_link": chat_link,
                                    "message_link": message_link,
                                    "jump_link": jump_link,
                                    "message_id": str(su.get("message_id") or ""),
                                    "status_update": su,
                                }
                                alert_events.append(obj)
                                alerts.append(
                                    json.dumps(
                                        obj,
                                        ensure_ascii=False,
                                        separators=(",", ":"),
                                        default=str,
                                    )
                                )
                        except TaskExtractionError as e:
                            append_dlq(
                                dlq_path,
                                {
                                    "type": "task_extract",
                                    "target_id": t.id,
                                    "target_type": t.type,
                                    "chat_id": chat_id,
                                    "error": str(e),
                                },
                            )

                    prev.update(new_frag)
                    prev["resolved_chat_id"] = chat_id
                    state_targets[t.id] = prev
                    any_success = True

                elif t.type == "feishu_mentions_global":
                    relative_time = str(t.raw.get("relative_time") or "last_6_hours")
                    page_size = int(t.raw.get("page_size") or 50)

                    msgs = fetch_feishu_mentions_global(relative_time=relative_time, page_size=page_size)
                    enriched_msgs = []
                    for m in msgs:
                        recovered_meta = _recover_mention_chat_meta(m, chat_registry, runtime_cache, dlq_path)
                        merged_msg = dict(m)
                        for k, v in recovered_meta.items():
                            if isinstance(v, str) and v:
                                merged_msg[k] = v
                        recovered_chat_id = recovered_meta.get("chat_id") or _guess_chat_id(merged_msg)
                        recovered_chat_name = recovered_meta.get("name") or ""
                        if recovered_chat_id and recovered_chat_name:
                            chat_registry.setdefault(recovered_chat_id, recovered_chat_name)
                        enriched = _enrich_chat_message_context(
                            merged_msg,
                            default_chat_name=recovered_chat_name,
                            chat_registry=chat_registry,
                            default_chat_id=recovered_chat_id,
                        )
                        enriched_msgs.append(enriched)
                    msgs = enriched_msgs
                    new_msgs, new_frag = diff_feishu_messages(prev, msgs)
                    actionable_new_msgs = [m for m in new_msgs if not _should_filter_zero_trust_message(m)]
                    is_first_run = not bool(prev.get("last_seen_message_id"))
                    if actionable_new_msgs:
                        mention_view_messages = []
                        new_ids: List[str] = []

                        for m in actionable_new_msgs:
                            chat_name = _verified_chat_name(_guess_chat_id(m), chat_registry)
                            chat_id = _guess_chat_id(m)
                            chat_link = _guess_chat_link(m)
                            message_link = _guess_message_link(m)
                            jump_link = message_link or chat_link
                            mid = str(m.get("message_id") or "")
                            sender = _guess_sender_name(m)
                            msg_text = _message_text_full(m)

                            mention_view_messages.append(
                                {
                                    "message_id": mid,
                                    "create_time": m.get("create_time"),
                                    "sender": sender,
                                    "chat_id": chat_id,
                                    "chat_name": chat_name,
                                    "chat_link": chat_link,
                                    "message_link": message_link,
                                    "jump_link": jump_link,
                                    "text": msg_text,
                                }
                            )
                            if mid:
                                new_ids.append(mid)

                            # 1) Raw message (100% preserved)
                            obj = {
                                "type": "mention_message_new",
                                "target_id": t.id,
                                "target_title": t.title,
                                "chat_id": chat_id,
                                "chat_name": chat_name,
                                "chat_link": chat_link,
                                "message_link": message_link,
                                "jump_link": jump_link,
                                "message_id": mid,
                                "sender": sender,
                                "create_time": m.get("create_time"),
                                "text": msg_text,
                            }
                            alert_events.append(obj)
                            if args.verbose:
                                print(f"[Debug] Appended mention, alert_events size: {len(alert_events)}", file=sys.stderr)
                            alerts.append(
                                json.dumps(
                                    obj,
                                    ensure_ascii=False,
                                    separators=(",", ":"),
                                    default=str,
                                )
                            )

                        # 2) Task extraction (batch)
                        try:
                            tasks = extract_tasks_from_chat_messages(
                                chat_title=t.title,
                                messages=msgs,
                                new_message_ids=new_ids,
                            )
                            for task in tasks:
                                source0 = _first_source_message(task)
                                event_chat_id = str(source0.get("chat_id") or "")
                                event_chat_name = _verified_chat_name(_guess_chat_id(source0), chat_registry)
                                event_chat_link = str(source0.get("chat_link") or _build_feishu_chat_link(event_chat_id))
                                event_message_link = str(source0.get("message_link") or source0.get("jump_link") or "")
                                event_jump_link = event_message_link or event_chat_link
                                obj = {
                                    "type": "chat_task",
                                    "target_id": t.id,
                                    "target_title": t.title,
                                    "chat_id": event_chat_id,
                                    "chat_name": event_chat_name,
                                    "chat_link": event_chat_link,
                                    "message_link": event_message_link,
                                    "jump_link": event_jump_link,
                                    "message_id": str(source0.get("message_id") or ""),
                                    "task": task,
                                }
                                alert_events.append(obj)

                                if args.verbose:
                                    md_lines = []
                                    task_name = str(task.get("task_name") or "未知任务")
                                    md_lines.append(f"**【新增群聊任务】** {task_name}")
                                    if event_chat_name:
                                        md_lines.append(f"🎯 群聊：{event_chat_name}")
                                    if event_jump_link:
                                        md_lines.append(f"🔗 跳转：[点击进入群聊]({event_jump_link})")
                                    print("\n".join(md_lines), file=sys.stderr)

                                alerts.append(
                                    json.dumps(
                                        obj,
                                        ensure_ascii=False,
                                        separators=(",", ":"),
                                        default=str,
                                    )
                                )

                            status_updates = extract_status_updates_from_messages(
                                view_messages=mention_view_messages,
                                known_task_names=[str(tk.get("task_name") or "") for tk in tasks],
                            )
                            for su in status_updates:
                                event_chat_id = str(su.get("chat_id") or "")
                                event_chat_name = _verified_chat_name(_guess_chat_id(su), chat_registry)
                                event_chat_link = str(su.get("chat_link") or _build_feishu_chat_link(event_chat_id))
                                event_message_link = str(su.get("message_link") or su.get("jump_link") or "")
                                event_jump_link = event_message_link or event_chat_link
                                obj = {
                                    "type": "task_status_update",
                                    "target_id": t.id,
                                    "target_title": t.title,
                                    "chat_id": event_chat_id,
                                    "chat_name": event_chat_name,
                                    "chat_link": event_chat_link,
                                    "message_link": event_message_link,
                                    "jump_link": event_jump_link,
                                    "message_id": str(su.get("message_id") or ""),
                                    "status_update": su,
                                }
                                alert_events.append(obj)

                                if args.verbose:
                                    md_lines = []
                                    su_action = str(su.get("action") or "状态更新")
                                    su_task = str(su.get("task_name") or "未知任务")
                                    md_lines.append(f"**【{su_action}】** {su_task}")
                                    if event_chat_name:
                                        md_lines.append(f"🎯 群聊：{event_chat_name}")
                                    if event_jump_link:
                                        md_lines.append(f"🔗 跳转：[点击进入群聊]({event_jump_link})")
                                    print("\n".join(md_lines), file=sys.stderr)

                                alerts.append(
                                    json.dumps(
                                        obj,
                                        ensure_ascii=False,
                                        separators=(",", ":"),
                                        default=str,
                                    )
                                )
                        except TaskExtractionError as e:
                            append_dlq(
                                dlq_path,
                                {
                                    "type": "task_extract",
                                    "target_id": t.id,
                                    "target_type": t.type,
                                    "error": str(e),
                                    "message_count": len(new_msgs),
                                },
                            )

                    prev.update(new_frag)
                    state_targets[t.id] = prev
                    any_success = True

                elif t.type == "lark_sheet_range":
                    matrix = fetch_lark_sheet_matrix(t)
                    changed, new_frag = diff_sheet(prev, matrix)

                    # First run baseline: do not alert.
                    is_first_run = prev.get("digest") is None
                    if changed and not is_first_run:
                        readable = f"[Heartbeat][{t.title}] 表格范围发生变化（{new_frag.get('rows')}x{new_frag.get('cols')}）"
                        if args.verbose:
                            print(readable, file=sys.stderr)
                        obj = {
                            "type": "sheet_range_change",
                            "target_id": t.id,
                            "target_title": t.title,
                            "rows": new_frag.get("rows"),
                            "cols": new_frag.get("cols"),
                            "text": readable,
                        }
                        alerts.append(
                            json.dumps(
                                obj,
                                ensure_ascii=False,
                                separators=(",", ":"),
                                default=str,
                            )
                        )

                    prev.update(new_frag)
                    state_targets[t.id] = prev
                    any_success = True

                else:
                    raise FetchError(f"unsupported target type: {t.type}")

                break
            except Exception as e:
                append_dlq(
                    dlq_path,
                    {
                        "type": "fetch",
                        "target_id": t.id,
                        "target_type": t.type,
                        "attempt": attempt + 1,
                        "error": str(e),
                    },
                )
                if attempt == 0:
                    continue

    # Persist state as long as we had any successful target fetch.
    if any_success:
        save_state(state_path, state)

    # Optional: dual-write JSON events into the task ledger sheets.
    # - 只对 JSON 事件生效（alert_events）
    # - 鉴权失败/写入失败：写入 DLQ，但不影响本次 stdout 告警输出
    if args.dual_write_spreadsheet and alert_events:
        if not auth_ok:
            append_dlq(
                dlq_path,
                {
                    "type": "dual_write",
                    "error": "bytedcli_auth_failed",
                    "spreadsheet": args.dual_write_spreadsheet,
                },
            )
        else:
            try:
                cli = LarkSheetsCLI()
                spreadsheet_token = cli.resolve_spreadsheet_token(args.dual_write_spreadsheet)
                writer = DualTrackWriter(
                    spreadsheet_token=spreadsheet_token,
                    log_sheet_title=args.dual_write_log_sheet_title,
                    task_sheet_title=args.dual_write_task_sheet_title,
                    cli=cli,
                )
                result = writer.write_events(
                    alert_events,
                    dry_run=bool(args.dual_write_dry_run),
                    raw_verify=not bool(args.dual_write_no_raw_verify),
                )

                # 若 RAW 校验被跳过，必须显式进入 DLQ（Zero-Trust：拒绝静默失败）
                if (not result.raw_verified) and result.raw_verify_skipped_reason:
                    append_dlq(
                        dlq_path,
                        {
                            "type": "dual_write_verify",
                            "warning": "raw_verify_skipped",
                            "reason": result.raw_verify_skipped_reason,
                            "spreadsheet": args.dual_write_spreadsheet,
                            "batch_id": result.batch_id,
                        },
                    )

                if args.verbose:
                    print(
                        f"[Heartbeat][dual_write] ok batch_id={result.batch_id} log_rows={result.written_log_rows} task_rows={result.written_task_rows}",
                        file=sys.stderr,
                    )
            except Exception as e:
                append_dlq(
                    dlq_path,
                    {
                        "type": "dual_write",
                        "error": str(e),
                        "spreadsheet": args.dual_write_spreadsheet,
                    },
                )

    # Output alerts (one line per alert; caller can forward as atomic messages)
    if alerts:
        for a in alerts:
            print(a)
    else:
        if args.verbose:
            print("[Heartbeat] 无新增，已静默退出", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
